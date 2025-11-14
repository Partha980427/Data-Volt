import streamlit as st
import pandas as pd
import os
from fractions import Fraction
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import tempfile
import re
from datetime import datetime
import plotly.express as px
import time
import json
import numpy as np
import math
import warnings
import logging
from typing import Dict, List, Optional, Any, Tuple
import io
import requests
from io import BytesIO
import openpyxl.styles
warnings.filterwarnings('ignore')

# ======================================================
# LOGGING CONFIGURATION
# ======================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler('fastener_app.log')
    ]
)
logger = logging.getLogger(__name__)

# ======================================================
# PATHS & FILES - UPDATED WITH GOOGLE SHEETS LINKS
# ======================================================
url = "https://docs.google.com/spreadsheets/d/11Icre8F3X8WA5BVwkJx75NOH3VzF6G7b/export?format=xlsx"
local_excel_path = r"G:\My Drive\Streamlite\ASME B18.2.1 Hex Bolt and Heavy Hex Bolt.xlsx"

# Mechanical and Chemical Properties paths
me_chem_google_url = "https://docs.google.com/spreadsheets/d/12lBzI67Wb0yZyJKYxpDCLzHF9zvS2Fha/export?format=xlsx"
me_chem_path = r"G:\My Drive\Streamlite\Mechanical and Chemical.xlsx"

# ISO 4014 paths - local and Google Sheets
iso4014_local_path = r"G:\My Drive\Streamlite\ISO 4014 Hex Bolt.xlsx"
iso4014_file_url = "https://docs.google.com/spreadsheets/d/1d2hANwoMhuzwyKJ72c125Uy0ujB6QsV_/export?format=xlsx"

# DIN-7991 paths - local and Google Sheets
din7991_local_path = r"G:\My Drive\Streamlite\DIN-7991.xlsx"
din7991_file_url = "https://docs.google.com/spreadsheets/d/1PjptIbFfebdF1h_Aj124fNgw5jNBWlvn/export?format=xlsx"

# ASME B18.3 paths - local and Google Sheets
asme_b18_3_local_path = r"G:\My Drive\Streamlite\ASME B18.3.xlsx"
asme_b18_3_file_url = "https://docs.google.com/spreadsheets/d/1dPNGwf7bv5A77rMSPpl11dhcJTXQfob1/export?format=xlsx"

# Thread files - UPDATED WITH GOOGLE SHEETS LINKS
thread_files = {
    "ASME B1.1": "https://docs.google.com/spreadsheets/d/1YHgUloNsFudxxqhWQV66D2DtSSKWFP_w/export?format=xlsx",
    "ISO 965-2-98 Coarse": "https://docs.google.com/spreadsheets/d/1be5eEy9hbVfMg2sl1-Cz1NNCGGF8EB-L/export?format=xlsx",
    "ISO 965-2-98 Fine": "https://docs.google.com/spreadsheets/d/1QGQ6SMWBSTsah-vq3zYnhOC3NXaBdKPe/export?format=xlsx",
}

# ======================================================
# LOADING INDICATORS MANAGEMENT
# ======================================================
class LoadingManager:
    """Manage loading states and progress indicators"""
    
    @staticmethod
    def show_loading_spinner(text="Processing..."):
        """Display a loading spinner"""
        return st.spinner(text)
    
    @staticmethod
    def show_progress_bar(step, total_steps, text="Processing"):
        """Display a progress bar"""
        progress = st.progress(0)
        for i in range(total_steps):
            progress.progress((i + 1) / total_steps, text=f"{text}... {i+1}/{total_steps}")
            time.sleep(0.1)
        progress.empty()
    
    @staticmethod
    def log_operation(operation_name, success=True, details=""):
        """Log operations with details"""
        status = "SUCCESS" if success else "FAILED"
        logger.info(f"{operation_name} - {status} - {details}")

# ======================================================
# ENHANCED CONFIGURATION & ERROR HANDLING
# ======================================================
@st.cache_data(ttl=3600, show_spinner=False)
def safe_load_excel_file_enhanced(path_or_url, max_retries=3, timeout=30):
    """Enhanced loading with better caching, validation and retry mechanism"""
    for attempt in range(max_retries):
        try:
            if path_or_url.startswith('http'):
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
                }
                response = requests.get(path_or_url, headers=headers, timeout=timeout)
                response.raise_for_status()
                
                if len(response.content) < 100:
                    st.warning(f"File seems too small: {path_or_url}")
                    continue
                    
                df = pd.read_excel(BytesIO(response.content))
            else:
                if os.path.exists(path_or_url):
                    file_size = os.path.getsize(path_or_url)
                    if file_size < 100:
                        st.warning(f"File seems too small: {path_or_url}")
                        continue
                    df = pd.read_excel(path_or_url)
                else:
                    st.error(f"File not found: {path_or_url}")
                    return pd.DataFrame()
            
            if df.empty:
                st.warning(f"Empty dataframe loaded from: {path_or_url}")
                return pd.DataFrame()
                
            if len(df.columns) < 2:
                st.warning(f"Dataframe has too few columns: {path_or_url}")
                return pd.DataFrame()
                
            LoadingManager.log_operation(f"Load Excel File: {path_or_url}", True, f"Rows: {len(df)}, Columns: {len(df.columns)}")
            return df
            
        except Exception as e:
            if attempt == max_retries - 1:
                st.error(f"Error loading {path_or_url}: {str(e)}")
                LoadingManager.log_operation(f"Load Excel File: {path_or_url}", False, str(e))
                return pd.DataFrame()
            time.sleep(1)

def validate_dataframe(df, required_columns=[]):
    """Validate dataframe structure"""
    if df.empty:
        return False, "DataFrame is empty"
    
    missing_cols = [col for col in required_columns if col not in df.columns]
    if missing_cols:
        return False, f"Missing columns: {missing_cols}"
    
    return True, "Valid"

def load_config():
    """Load configuration from session state with defaults"""
    if 'app_config' not in st.session_state:
        st.session_state.app_config = {
            'data_sources': {
                'main_data': url,
                'me_chem_data': me_chem_google_url,
                'iso4014': iso4014_file_url,
                'din7991': din7991_file_url,
                'asme_b18_3': asme_b18_3_file_url,
                'thread_files': thread_files
            },
            'ui': {
                'theme': 'oracle11g',
                'page_title': 'JSC Industries - Fastener Intelligence'
            },
            'features': {
                'batch_processing': True,
                'analytics': True
            }
        }
    return st.session_state.app_config

def save_user_preferences():
    """Save user preferences to session state"""
    if 'user_prefs' not in st.session_state:
        st.session_state.user_prefs = {
            'default_standard': 'ASME B1.1',
            'preferred_units': 'metric',
            'recent_searches': [],
            'favorite_filters': {},
            'theme_preference': 'oracle11g'
        }

def initialize_session_state():
    """Initialize all session state variables"""
    defaults = {
        "selected_section": None,
        "batch_result_df": None,
        "current_filters": {},
        "recent_searches": [],
        "favorite_products": [],
        "calculation_history": [],
        "export_format": "csv",
        "multi_search_products": [],
        "current_filters_dimensional": {},
        "current_filters_thread": {},
        "current_filters_material": {},
        "product_intelligence_filters": {},
        "me_chem_columns": [],
        "property_classes": [],
        "din7991_loaded": False,
        "asme_b18_3_loaded": False,
        "dimensional_standards_count": 0,
        "available_products": {},
        "available_series": {},
        "debug_mode": False,
        "section_a_view": True,
        "section_b_view": True,
        "section_c_view": True,
        "thread_independent_mode": True,
        "section_a_results": pd.DataFrame(),
        "section_b_results": pd.DataFrame(),
        "section_c_results": pd.DataFrame(),
        "combined_results": pd.DataFrame(),
        "section_a_filters": {},
        "section_b_filters": {},
        "section_c_filters": {},
        "section_a_current_product": "All",
        "section_a_current_series": "All",
        "section_a_current_standard": "All",
        "section_a_current_size": "All",
        "section_a_current_grade": "All",
        "section_b_current_standard": "All",
        "section_b_current_size": "All",
        "section_b_current_class": "All",
        "section_c_current_class": "All",
        "section_c_current_standard": "All",
        "thread_data_cache": {},
        "show_professional_card": False,
        "selected_product_details": None,
        "batch_calculation_results": pd.DataFrame(),
        # Weight calculator session states
        "weight_calc_product": "Select Product",
        "weight_calc_series": "Select Series",
        "weight_calc_standard": "Select Standard",
        "weight_calc_size": "Select Size",
        "weight_calc_grade": "Select Grade",
        "weight_calc_diameter_type": "Blank Diameter",
        "weight_calc_blank_diameter": 10.0,
        "weight_calc_blank_dia_unit": "mm",
        "weight_calc_thread_standard": "Select Thread Standard",
        "weight_calc_thread_size": "All",
        "weight_calc_thread_class": "2A",
        "weight_calc_length": 50.0,
        "weight_calc_length_unit": "mm",
        "weight_calc_material": "Carbon Steel",
        "weight_calc_result": None,
        "weight_calculation_performed": False,
        "pitch_diameter_value": None,
        "weight_form_submitted": False,
        # Mobile view state
        "mobile_view_optimized": False,
        # Batch calculator session states - NEW
        "batch_uploaded_file": None,
        "batch_processing": False,
        "batch_results": None,
        "batch_summary": None,
        "batch_errors": [],
        "batch_processing_complete": False,
        "batch_mode": "basic",  # 'basic' or 'advanced'
        "batch_diameter_type": "Blank Diameter",  # NEW: Store diameter type for batch
    }
    
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value
    
    load_config()
    save_user_preferences()

# ======================================================
# MOBILE OPTIMIZATION
# ======================================================
def detect_mobile_device():
    """Detect if the user is on a mobile device"""
    try:
        user_agent = st.query_params.get('user_agent', '')
        mobile_indicators = ['Mobile', 'Android', 'iPhone', 'iPad']
        return any(indicator in user_agent for indicator in mobile_indicators)
    except:
        return False

def optimize_for_mobile():
    """Apply mobile-specific optimizations"""
    if detect_mobile_device() or st.session_state.mobile_view_optimized:
        st.session_state.mobile_view_optimized = True
        st.markdown("""
        <style>
        @media (max-width: 768px) {
            .oracle11g-header {
                padding: 1rem !important;
            }
            .oracle11g-header h1 {
                font-size: 1.5rem !important;
            }
            .stButton > button {
                padding: 0.5rem 1rem !important;
                font-size: 0.9rem !important;
            }
            .stSelectbox, .stTextInput, .stNumberInput {
                font-size: 0.9rem !important;
            }
            .metric-card {
                padding: 1rem !important;
                margin-bottom: 0.5rem !important;
            }
            .quick-action {
                height: 100px !important;
                padding: 1rem 0.5rem !important;
            }
        }
        </style>
        """, unsafe_allow_html=True)

# ======================================================
# ENHANCED DATA EXPORT TEMPLATES
# ======================================================
class ExportTemplateManager:
    """Manage export templates for different data types"""
    
    @staticmethod
    def get_weight_calc_template():
        """Get template for weight calculations"""
        template_data = {
            'Product_Type': ['Hex Bolt', 'Heavy Hex Bolt', 'Threaded Rod', 'Hex Cap Screws', 'Hex Bolt', 'Hexagon Socket Head Cap Screws', 'Hexagon Socket Countersunk Head Cap Screw'],
            'Series': ['Inch', 'Inch', 'Inch', 'Inch', 'Metric', 'Inch', 'Metric'],
            'Standard': ['ASME B18.2.1', 'ASME B18.2.1', 'Not Required', 'ASME B18.2.1', 'ISO 4014', 'ASME B18.3', 'DIN-7991'],
            'Size': ['1/4', '5/16', 'Not Required', '3/8', 'M10', '1/4', 'M6'],
            'Grade': ['N/A', 'N/A', 'N/A', 'N/A', 'A', 'N/A', 'N/A'],
            'Diameter_Type': ['Blank Diameter', 'Pitch Diameter', 'Pitch Diameter', 'Blank Diameter', 'Blank Diameter', 'Blank Diameter', 'Blank Diameter'],
            'Blank_Diameter': [6.35, 0, 0, 9.525, 10.0, 6.35, 6.0],
            'Blank_Diameter_Unit': ['mm', 'mm', 'mm', 'mm', 'mm', 'mm', 'mm'],
            'Thread_Standard': ['N/A', 'ASME B1.1', 'ASME B1.1', 'N/A', 'N/A', 'N/A', 'N/A'],
            'Thread_Size': ['N/A', '1/4', '1/2', 'N/A', 'N/A', 'N/A', 'N/A'],
            'Thread_Class': ['N/A', '2A', '2A', 'N/A', 'N/A', 'N/A', 'N/A'],
            'Length': [50, 100, 200, 75, 60, 50, 40],
            'Length_Unit': ['mm', 'mm', 'ft', 'mm', 'mm', 'mm', 'mm'],
            'Material': ['Carbon Steel', 'Carbon Steel', 'Stainless Steel', 'Carbon Steel', 'Carbon Steel', 'Carbon Steel', 'Carbon Steel']
        }
        return pd.DataFrame(template_data)
    
    @staticmethod
    def get_product_database_template():
        """Get template for product database exports"""
        template_data = {
            'Product': ['Hex Bolt', 'Heavy Hex Bolt', 'Hex Cap Screw'],
            'Standard': ['ASME B18.2.1', 'ASME B18.2.1', 'ASME B18.3'],
            'Size': ['1/4', '5/16', '1/4'],
            'Thread': ['1/4-20-UNC-2A', '5/16-18-UNC-2A', '1/4-20-UNC-2A'],
            'Material_Grade': ['N/A', 'N/A', 'N/A'],
            'Quantity': [100, 50, 200],
            'Notes': ['Standard hex bolt', 'Heavy hex pattern', 'Socket head cap screw']
        }
        return pd.DataFrame(template_data)
    
    @staticmethod
    def export_to_pdf(data, title="Fastener Report"):
        """Export data to PDF format (placeholder for future implementation)"""
        st.info("PDF export feature will be available in the next update")
        return None

# ======================================================
# ENHANCED BATCH CALCULATOR TEMPLATES WITH DIAMETER TYPE SUPPORT
# ======================================================
class BatchTemplateManager:
    """Manage batch calculator templates with diameter type support"""
    
    @staticmethod
    def get_basic_template(diameter_type="Blank Diameter"):
        """Get basic template with Product_Type column and diameter type support"""
        if diameter_type == "Blank Diameter":
            template_data = {
                'Product_Type': ['Hex Bolt', 'Hex Bolt', 'Threaded Rod', 'Hex Bolt'],
                'Product_Code': ['HB-001', 'HB-002', 'TR-001', 'HB-003'],
                'Size': ['1/4', 'M10', '1/2', '5/16'],
                'Length': [50, 75, 200, 100],
                'Length_Unit': ['mm', 'mm', 'mm', 'mm'],
                'Diameter_Value': [6.35, 10.0, 12.7, 7.94],
                'Diameter_Unit': ['mm', 'mm', 'mm', 'mm'],
                'Material': ['Carbon Steel', 'Carbon Steel', 'Stainless Steel', 'Carbon Steel'],
                'Quantity': [100, 50, 25, 200]
            }
        else:  # Pitch Diameter
            template_data = {
                'Product_Type': ['Hex Bolt', 'Hex Bolt', 'Threaded Rod', 'Hex Bolt'],
                'Product_Code': ['HB-001', 'HB-002', 'TR-001', 'HB-003'],
                'Size': ['1/4', '5/16', '1/2', '3/8'],
                'Length': [50, 75, 200, 100],
                'Length_Unit': ['mm', 'mm', 'mm', 'mm'],
                'Thread_Standard': ['ASME B1.1', 'ASME B1.1', 'ASME B1.1', 'ASME B1.1'],
                'Product_Standard': ['ASME B18.2.1', 'ASME B18.2.1', 'Not Required', 'ASME B18.2.1'],
                'Thread_Class': ['2A', '2A', '2A', '2A'],
                'Material': ['Carbon Steel', 'Carbon Steel', 'Stainless Steel', 'Carbon Steel'],
                'Quantity': [100, 50, 25, 200]
            }
        return pd.DataFrame(template_data)
    
    @staticmethod
    def get_advanced_template(diameter_type="Blank Diameter"):
        """Get advanced template with all parameters and diameter type support"""
        if diameter_type == "Blank Diameter":
            template_data = {
                'Product_Type': ['Hex Bolt', 'Hex Bolt', 'Threaded Rod', 'Hexagon Socket Head Cap Screws', 'Hex Bolt'],
                'Product_Code': ['HB-001', 'HB-002', 'TR-001', 'SHS-001', 'HB-003'],
                'Series': ['Inch', 'Metric', 'Inch', 'Inch', 'Metric'],
                'Standard': ['ASME B18.2.1', 'ISO 4014', 'Not Required', 'ASME B18.3', 'ISO 4014'],
                'Size': ['1/4', 'M10', '1/2', '1/4', 'M12'],
                'Grade': ['N/A', 'A', 'N/A', 'N/A', 'B'],
                'Diameter_Type': ['Blank Diameter', 'Blank Diameter', 'Blank Diameter', 'Blank Diameter', 'Blank Diameter'],
                'Diameter_Value': [6.35, 10.0, 12.7, 6.35, 12.0],
                'Diameter_Unit': ['mm', 'mm', 'mm', 'mm', 'mm'],
                'Thread_Standard': ['N/A', 'N/A', 'ASME B1.1', 'N/A', 'N/A'],
                'Thread_Size': ['N/A', 'N/A', '1/2', 'N/A', 'N/A'],
                'Thread_Class': ['N/A', 'N/A', '2A', 'N/A', 'N/A'],
                'Length': [50, 75, 200, 50, 100],
                'Length_Unit': ['mm', 'mm', 'mm', 'mm', 'mm'],
                'Material': ['Carbon Steel', 'Carbon Steel', 'Stainless Steel', 'Carbon Steel', 'Carbon Steel'],
                'Quantity': [100, 50, 25, 75, 150]
            }
        else:  # Pitch Diameter
            template_data = {
                'Product_Type': ['Hex Bolt', 'Hex Bolt', 'Threaded Rod', 'Hex Bolt'],
                'Product_Code': ['HB-001', 'HB-002', 'TR-001', 'HB-003'],
                'Series': ['Inch', 'Inch', 'Inch', 'Inch'],
                'Standard': ['ASME B18.2.1', 'ASME B18.2.1', 'Not Required', 'ASME B18.2.1'],
                'Size': ['1/4', '5/16', '1/2', '3/8'],
                'Grade': ['N/A', 'N/A', 'N/A', 'N/A'],
                'Diameter_Type': ['Pitch Diameter', 'Pitch Diameter', 'Pitch Diameter', 'Pitch Diameter'],
                'Thread_Standard': ['ASME B1.1', 'ASME B1.1', 'ASME B1.1', 'ASME B1.1'],
                'Thread_Size': ['1/4', '5/16', '1/2', '3/8'],
                'Thread_Class': ['2A', '2A', '2A', '2A'],
                'Length': [50, 75, 200, 100],
                'Length_Unit': ['mm', 'mm', 'mm', 'mm'],
                'Material': ['Carbon Steel', 'Carbon Steel', 'Stainless Steel', 'Carbon Steel'],
                'Quantity': [100, 50, 25, 200]
            }
        return pd.DataFrame(template_data)
    
    @staticmethod
    def detect_input_mode(row):
        """Detect whether row is in basic or advanced mode"""
        basic_columns = ['Product_Type', 'Size', 'Length']
        advanced_columns = ['Product_Type', 'Series', 'Standard', 'Diameter_Type']
        
        # Check if advanced columns are populated
        advanced_mode = any(pd.notna(row.get(col, None)) for col in advanced_columns if col in row)
        
        # Check if basic columns are populated
        basic_mode = all(pd.notna(row.get(col, None)) for col in basic_columns if col in row)
        
        if advanced_mode:
            return "advanced"
        elif basic_mode:
            return "basic"
        else:
            return "invalid"
    
    @staticmethod
    def infer_parameters_basic_mode(row, diameter_type="Blank Diameter"):
        """Intelligently infer parameters from basic mode with diameter type support"""
        try:
            product_type = row.get('Product_Type', 'Hex Bolt')
            product_code = row.get('Product_Code', '')
            size = row.get('Size')
            length = row.get('Length', 50.0)
            length_unit = row.get('Length_Unit', 'mm')
            material = row.get('Material', 'Carbon Steel')
            quantity = row.get('Quantity', 1)
            thread_standard = row.get('Thread_Standard', 'ASME B1.1')
            product_standard = row.get('Product_Standard', 'ASME B18.2.1')
            thread_class = row.get('Thread_Class', '2A')
            
            # Initialize default parameters
            params = {
                'product_type': product_type,
                'product_code': product_code,
                'size': str(size),
                'material': material,
                'length': length,
                'length_unit': length_unit,
                'quantity': quantity,
                'diameter_type': diameter_type,
                'thread_standard': thread_standard,
                'product_standard': product_standard,
                'thread_class': thread_class
            }
            
            # Analyze size pattern
            size_str = str(size).strip().upper()
            
            # Metric detection (M10, M12, M16, etc.)
            if size_str.startswith('M'):
                params['series'] = 'Metric'
                params['standard'] = 'ISO 4014' if product_standard == 'All' else product_standard
                
                # Extract diameter from metric size (M10 -> 10.0 mm)
                try:
                    diameter_value = float(size_str[1:])
                    params['diameter_value'] = diameter_value
                    params['diameter_unit'] = 'mm'
                    params['grade'] = 'A'  # Default grade for metric
                except ValueError:
                    params['diameter_value'] = 10.0  # Default fallback
                    params['diameter_unit'] = 'mm'
                    params['grade'] = 'A'
            
            # Inch detection (fractions or numbers)
            elif '/' in size_str or any(char.isdigit() for char in size_str):
                params['series'] = 'Inch'
                params['standard'] = 'ASME B18.2.1' if product_standard == 'All' else product_standard
                params['grade'] = 'N/A'
                
                try:
                    # Handle fractions
                    if '/' in size_str:
                        if '-' in size_str:
                            # Handle cases like "1-1/2"
                            parts = size_str.split('-')
                            whole = float(parts[0]) if parts[0] else 0
                            fraction = float(Fraction(parts[1]))
                            decimal_inches = whole + fraction
                        else:
                            decimal_inches = float(Fraction(size_str))
                    else:
                        decimal_inches = float(size_str)
                    
                    # Convert inches to mm for calculation
                    params['diameter_value'] = decimal_inches * 25.4
                    params['diameter_unit'] = 'mm'
                    
                except (ValueError, ZeroDivisionError):
                    params['diameter_value'] = 6.35  # 1/4" default
                    params['diameter_unit'] = 'mm'
            
            else:
                # Default fallback
                params['series'] = 'Inch'
                params['standard'] = 'ASME B18.2.1'
                params['diameter_value'] = 10.0
                params['diameter_unit'] = 'mm'
                params['grade'] = 'N/A'
            
            # Handle diameter type specific parameters
            if diameter_type == "Pitch Diameter":
                # For pitch diameter, we need thread information
                params['thread_standard'] = row.get('Thread_Standard', 'ASME B1.1')
                params['thread_size'] = row.get('Thread_Size', size)
                params['thread_class'] = row.get('Thread_Class', '2A')
                
                # Get pitch diameter from database
                pitch_diameter = get_pitch_diameter_from_thread_data(
                    params['thread_standard'],
                    params['thread_size'],
                    params['thread_class']
                )
                
                if pitch_diameter is not None:
                    params['diameter_value'] = pitch_diameter
                    params['diameter_unit'] = 'inch' if params['series'] == 'Inch' else 'mm'
                else:
                    # Fallback to blank diameter calculation
                    st.warning(f"Pitch diameter not found for {params['thread_size']}, using blank diameter")
                    params['diameter_type'] = 'Blank Diameter'
            
            elif diameter_type == "Blank Diameter":
                # Use the provided diameter values
                params['diameter_value'] = row.get('Diameter_Value', params.get('diameter_value', 10.0))
                params['diameter_unit'] = row.get('Diameter_Unit', params.get('diameter_unit', 'mm'))
            
            return params
            
        except Exception as e:
            st.error(f"Error inferring parameters for size {size}: {str(e)}")
            # Return safe defaults
            return {
                'product_type': 'Hex Bolt',
                'product_code': product_code,
                'series': 'Inch',
                'standard': 'ASME B18.2.1',
                'size': str(size),
                'grade': 'N/A',
                'diameter_type': diameter_type,
                'diameter_value': 10.0,
                'diameter_unit': 'mm',
                'material': material,
                'length': length,
                'length_unit': length_unit,
                'quantity': quantity,
                'thread_standard': thread_standard,
                'product_standard': product_standard,
                'thread_class': thread_class
            }

# ======================================================
# BATCH PROCESSING ENGINE WITH DIAMETER TYPE SUPPORT
# ======================================================
class BatchProcessor:
    """Handle batch weight calculations with diameter type support"""
    
    @staticmethod
    def validate_batch_file(df, mode="basic", diameter_type="Blank Diameter"):
        """Validate batch file structure and data with diameter type support"""
        errors = []
        warnings = []
        
        if df.empty:
            errors.append("Uploaded file is empty")
            return False, errors, warnings
        
        # Basic mode validation
        if mode == "basic":
            required_columns = ['Product_Type', 'Size', 'Length']
            
            if diameter_type == "Blank Diameter":
                required_columns.extend(['Diameter_Value', 'Diameter_Unit'])
            else:  # Pitch Diameter
                required_columns.extend(['Thread_Standard', 'Product_Standard', 'Thread_Class'])
            
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                errors.append(f"Missing required columns for basic mode with {diameter_type}: {', '.join(missing_columns)}")
            
            # Check for empty values in required columns
            for col in required_columns:
                if col in df.columns and df[col].isna().any():
                    empty_count = df[col].isna().sum()
                    warnings.append(f"Column '{col}' has {empty_count} empty values")
        
        # Advanced mode validation
        elif mode == "advanced":
            required_columns = ['Product_Type', 'Size', 'Length', 'Diameter_Type']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                errors.append(f"Missing required columns for advanced mode: {', '.join(missing_columns)}")
        
        # Check data types
        if 'Length' in df.columns:
            try:
                pd.to_numeric(df['Length'], errors='coerce')
            except:
                errors.append("Length column contains non-numeric values")
        
        if 'Quantity' in df.columns:
            try:
                pd.to_numeric(df['Quantity'], errors='coerce')
            except:
                warnings.append("Quantity column contains non-numeric values - will use default of 1")
        
        if 'Diameter_Value' in df.columns:
            try:
                pd.to_numeric(df['Diameter_Value'], errors='coerce')
            except:
                errors.append("Diameter_Value column contains non-numeric values")
        
        return len(errors) == 0, errors, warnings
    
    @staticmethod
    def process_batch_calculations(batch_df, diameter_type="Blank Diameter", progress_callback=None):
        """Process batch calculations for all rows with diameter type support"""
        results = []
        errors = []
        summary = {
            'total_rows': len(batch_df),
            'successful_calculations': 0,
            'failed_calculations': 0,
            'total_weight_kg': 0.0,
            'total_weight_lb': 0.0,
            'start_time': datetime.now(),
            'diameter_type_used': diameter_type
        }
        
        for index, row in batch_df.iterrows():
            try:
                # Determine input mode and prepare parameters
                input_mode = BatchTemplateManager.detect_input_mode(row)
                
                if input_mode == "invalid":
                    errors.append({
                        'row_index': index,
                        'input_data': row.to_dict(),
                        'error': 'Invalid input - missing required columns',
                        'status': 'failed'
                    })
                    continue
                
                # Prepare calculation parameters based on mode
                if input_mode == "basic":
                    params = BatchTemplateManager.infer_parameters_basic_mode(row, diameter_type)
                else:  # advanced mode
                    params = {
                        'product_type': row.get('Product_Type', 'Hex Bolt'),
                        'product_code': row.get('Product_Code', ''),
                        'series': row.get('Series', 'Inch'),
                        'standard': row.get('Standard', 'ASME B18.2.1'),
                        'size': row.get('Size'),
                        'grade': row.get('Grade', 'N/A'),
                        'diameter_type': row.get('Diameter_Type', diameter_type),
                        'diameter_value': row.get('Diameter_Value', 10.0),
                        'diameter_unit': row.get('Diameter_Unit', 'mm'),
                        'thread_standard': row.get('Thread_Standard', 'N/A'),
                        'thread_size': row.get('Thread_Size', 'N/A'),
                        'thread_class': row.get('Thread_Class', 'N/A'),
                        'length': row.get('Length'),
                        'length_unit': row.get('Length_Unit', 'mm'),
                        'material': row.get('Material', 'Carbon Steel'),
                        'quantity': row.get('Quantity', 1)
                    }
                
                # Handle pitch diameter thread data
                if params['diameter_type'] == 'Pitch Diameter' and params.get('thread_standard') != 'N/A':
                    pitch_diameter = get_pitch_diameter_from_thread_data(
                        params['thread_standard'],
                        params['thread_size'],
                        params['thread_class']
                    )
                    
                    if pitch_diameter is not None:
                        params['diameter_value'] = pitch_diameter
                        params['diameter_unit'] = 'inch' if params.get('series') == 'Inch' else 'mm'
                    else:
                        errors.append({
                            'row_index': index,
                            'input_data': row.to_dict(),
                            'error': f"Pitch diameter not found for {params['thread_size']}",
                            'status': 'failed'
                        })
                        summary['failed_calculations'] += 1
                        continue
                
                # Perform calculation
                calculation_result = calculate_weight_rectified(params)
                
                if calculation_result:
                    # Add batch-specific information
                    result_record = {
                        'row_index': index,
                        'input_data': row.to_dict(),
                        'calculation_result': calculation_result,
                        'status': 'success',
                        'input_mode': input_mode,
                        'quantity': params.get('quantity', 1)
                    }
                    
                    results.append(result_record)
                    summary['successful_calculations'] += 1
                    summary['total_weight_kg'] += calculation_result['weight_kg'] * result_record['quantity']
                    summary['total_weight_lb'] += calculation_result['weight_lb'] * result_record['quantity']
                    
                    # Update progress
                    if progress_callback and index % max(1, len(batch_df) // 10) == 0:
                        progress = (index + 1) / len(batch_df)
                        progress_callback(progress, f"Processed {index + 1}/{len(batch_df)} rows")
                
                else:
                    errors.append({
                        'row_index': index,
                        'input_data': row.to_dict(),
                        'error': 'Calculation returned no result',
                        'status': 'failed',
                        'input_mode': input_mode
                    })
                    summary['failed_calculations'] += 1
                    
            except Exception as e:
                errors.append({
                    'row_index': index,
                    'input_data': row.to_dict(),
                    'error': str(e),
                    'status': 'failed',
                    'input_mode': input_mode if 'input_mode' in locals() else 'unknown'
                })
                summary['failed_calculations'] += 1
        
        summary['end_time'] = datetime.now()
        summary['processing_time'] = (summary['end_time'] - summary['start_time']).total_seconds()
        
        return results, errors, summary

# ======================================================
# BATCH RESULTS DISPLAY
# ======================================================
class BatchResultsDisplay:
    """Display batch calculation results"""
    
    @staticmethod
    def show_processing_summary(summary):
        """Show batch processing summary"""
        st.markdown("### 📊 Batch Processing Summary")
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("Total Records", summary['total_rows'])
        with col2:
            success_rate = (summary['successful_calculations'] / summary['total_rows']) * 100 if summary['total_rows'] > 0 else 0
            st.metric("Successful", f"{summary['successful_calculations']} ({success_rate:.1f}%)")
        with col3:
            st.metric("Failed", summary['failed_calculations'])
        with col4:
            st.metric("Processing Time", f"{summary['processing_time']:.2f}s")
        
        st.markdown(f"**Diameter Type Used:** {summary.get('diameter_type_used', 'Blank Diameter')}")
        st.markdown("---")
    
    @staticmethod
    def show_weight_summary(summary):
        """Show weight summary"""
        st.markdown("### ⚖️ Total Weight Summary")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.metric("Total Weight (kg)", f"{summary['total_weight_kg']:.4f}")
        with col2:
            st.metric("Total Weight (lb)", f"{summary['total_weight_lb']:.4f}")
        with col3:
            st.metric("Total Weight (grams)", f"{summary['total_weight_kg'] * 1000:.2f}")
    
    @staticmethod
    def show_detailed_results(results):
        """Show detailed results table"""
        if not results:
            return
        
        st.markdown("### 📋 Detailed Results")
        
        # Prepare data for display
        display_data = []
        for result in results:
            calc = result['calculation_result']
            input_data = result['input_data']
            quantity = result.get('quantity', 1)
            
            display_record = {
                'Row': result['row_index'] + 1,
                'Product': input_data.get('Product_Type', 'Auto-detected'),
                'Product_Code': input_data.get('Product_Code', ''),
                'Size': input_data.get('Size', 'N/A'),
                'Length': f"{input_data.get('Length', 'N/A')} {input_data.get('Length_Unit', 'mm')}",
                'Material': input_data.get('Material', 'Carbon Steel'),
                'Diameter Type': input_data.get('Diameter_Type', 'Blank Diameter'),
                'Input Mode': result.get('input_mode', 'basic').title(),
                'Weight (kg)': f"{calc['weight_kg']:.4f}",
                'Weight (lb)': f"{calc['weight_lb']:.4f}",
                'Quantity': quantity,
                'Total Weight (kg)': f"{calc['weight_kg'] * quantity:.4f}",
                'Total Weight (lb)': f"{calc['weight_lb'] * quantity:.4f}",
                'Status': '✅ Success'
            }
            display_data.append(display_record)
        
        results_df = pd.DataFrame(display_data)
        st.dataframe(results_df, use_container_width=True)
    
    @staticmethod
    def show_error_report(errors):
        """Show error report"""
        if not errors:
            return
        
        st.markdown("### ❌ Error Report")
        st.warning(f"Found {len(errors)} calculation errors")
        
        for error in errors[:10]:  # Show first 10 errors
            with st.expander(f"Row {error['row_index'] + 1} - {error['error']}"):
                st.write("**Input Data:**", error['input_data'])
                st.write("**Error:**", error['error'])
        
        if len(errors) > 10:
            st.info(f"Showing first 10 of {len(errors)} errors. Download full report for complete details.")
    
    @staticmethod
    def export_batch_results(results, errors, summary, filename_prefix="batch_weight_results"):
        """Export batch results to Excel"""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{filename_prefix}_{timestamp}.xlsx"
            
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                with pd.ExcelWriter(tmp.name, engine='openpyxl') as writer:
                    # Sheet 1: Summary
                    summary_df = pd.DataFrame([summary])
                    summary_df.to_excel(writer, sheet_name='Summary', index=False)
                    
                    # Sheet 2: Detailed Results
                    if results:
                        detailed_data = []
                        for result in results:
                            calc = result['calculation_result']
                            input_data = result['input_data']
                            quantity = result.get('quantity', 1)
                            
                            detailed_record = {
                                'Row_Index': result['row_index'] + 1,
                                'Product_Type': input_data.get('Product_Type', 'Auto-detected'),
                                'Product_Code': input_data.get('Product_Code', ''),
                                'Series': input_data.get('Series', 'Auto-detected'),
                                'Standard': input_data.get('Standard', 'Auto-detected'),
                                'Size': input_data.get('Size', 'N/A'),
                                'Grade': input_data.get('Grade', 'N/A'),
                                'Diameter_Type': input_data.get('Diameter_Type', 'Auto-detected'),
                                'Diameter_Value': input_data.get('Diameter_Value', 'Auto-calculated'),
                                'Diameter_Unit': input_data.get('Diameter_Unit', 'mm'),
                                'Length': input_data.get('Length', 'N/A'),
                                'Length_Unit': input_data.get('Length_Unit', 'mm'),
                                'Material': input_data.get('Material', 'Carbon Steel'),
                                'Input_Mode': result.get('input_mode', 'basic'),
                                'Weight_kg': calc['weight_kg'],
                                'Weight_lb': calc['weight_lb'],
                                'Quantity': quantity,
                                'Total_Weight_kg': calc['weight_kg'] * quantity,
                                'Total_Weight_lb': calc['weight_lb'] * quantity,
                                'Status': 'Success'
                            }
                            detailed_data.append(detailed_record)
                        
                        detailed_df = pd.DataFrame(detailed_data)
                        detailed_df.to_excel(writer, sheet_name='Detailed_Results', index=False)
                    
                    # Sheet 3: Error Report
                    if errors:
                        error_df = pd.DataFrame(errors)
                        error_df.to_excel(writer, sheet_name='Error_Report', index=False)
                    
                    # Sheet 4: Processing Log
                    log_data = {
                        'Timestamp': [datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                        'Total_Rows': [summary['total_rows']],
                        'Successful': [summary['successful_calculations']],
                        'Failed': [summary['failed_calculations']],
                        'Success_Rate': [f"{(summary['successful_calculations']/summary['total_rows'])*100:.2f}%"],
                        'Total_Weight_kg': [summary['total_weight_kg']],
                        'Total_Weight_lb': [summary['total_weight_lb']],
                        'Processing_Time_seconds': [summary['processing_time']],
                        'Diameter_Type': [summary.get('diameter_type_used', 'Blank Diameter')]
                    }
                    log_df = pd.DataFrame(log_data)
                    log_df.to_excel(writer, sheet_name='Processing_Log', index=False)
                
                return tmp.name, filename
                
        except Exception as e:
            st.error(f"Error exporting results: {str(e)}")
            return None, None

# ======================================================
# ENHANCED BATCH CALCULATOR UI WITH DIAMETER TYPE SELECTION
# ======================================================
def show_batch_weight_calculator():
    """Enhanced Batch Weight Calculator with Diameter Type Selection"""
    
    st.markdown("""
    <div class="oracle11g-header">
        <h1>Industrial Batch Weight Calculator</h1>
        <p>Process 1000+ products simultaneously • Smart Diameter Handling</p>
        <div>
            <span class="oracle11g-badge">Batch Processing</span>
            <span class="oracle11g-badge-orange">Smart Diameter Handling</span>
            <span class="oracle11g-badge-green">Auto Pitch Diameter</span>
            <span class="oracle11g-badge-yellow">Product Code Column</span>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # Diameter Type Selection - ADDED AT THE TOP
    st.markdown("### 🎯 Select Diameter Type")
    diameter_type = st.radio(
        "Choose how to specify diameters:",
        ["Blank Diameter", "Pitch Diameter"],
        horizontal=True,
        key="batch_diameter_type"
    )
    
    st.info(f"""
    **{diameter_type} Mode Selected:**
    - **Blank Diameter**: Provide diameter values directly in the template
    - **Pitch Diameter**: System automatically fetches pitch diameter values from thread database
    """)
    
    # Mode selection
    st.markdown("### 🎯 Select Input Mode")
    mode_col1, mode_col2 = st.columns(2)
    
    with mode_col1:
        basic_mode = st.checkbox(
            "Basic Mode (Auto-Detection)", 
            value=st.session_state.batch_mode == "basic",
            help="Provide Product Type, Size and Length - system auto-detects other parameters"
        )
    
    with mode_col2:
        advanced_mode = st.checkbox(
            "Advanced Mode (Manual Specification)", 
            value=st.session_state.batch_mode == "advanced",
            help="Provide complete product specifications for precise control"
        )
    
    # Set mode
    if basic_mode and not advanced_mode:
        st.session_state.batch_mode = "basic"
    elif advanced_mode and not basic_mode:
        st.session_state.batch_mode = "advanced"
    elif not basic_mode and not advanced_mode:
        st.session_state.batch_mode = "basic"  # Default
    
    st.markdown("---")
    
    # Template download section - UPDATED WITH DIAMETER TYPE
    st.markdown("### 📥 Download Professional Template")
    
    st.info(f"**Template will include:** Product_Type, Product_Code, Size, Length, and {'Diameter_Value, Diameter_Unit' if diameter_type == 'Blank Diameter' else 'Thread_Standard, Product_Standard, Thread_Class'} columns")
    
    col1, col2 = st.columns(2)
    
    with col1:
        if st.button("Download Basic Template", use_container_width=True):
            template_df = BatchTemplateManager.get_basic_template(diameter_type)
            
            # Create professional Excel file with formatting
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                with pd.ExcelWriter(tmp.name, engine='openpyxl') as writer:
                    template_df.to_excel(writer, sheet_name='Batch Template', index=False)
                    
                    # Get workbook and worksheet
                    workbook = writer.book
                    worksheet = writer.sheets['Batch Template']
                    
                    # Set column widths for better readability
                    column_widths = {
                        'A': 20,  # Product_Type
                        'B': 15,  # Product_Code
                        'C': 12,  # Size
                        'D': 12,  # Length
                        'E': 12,  # Length_Unit
                    }
                    
                    if diameter_type == "Blank Diameter":
                        column_widths.update({
                            'F': 15,  # Diameter_Value
                            'G': 15,  # Diameter_Unit
                            'H': 15,  # Material
                            'I': 12   # Quantity
                        })
                    else:  # Pitch Diameter
                        column_widths.update({
                            'F': 18,  # Thread_Standard
                            'G': 18,  # Product_Standard
                            'H': 15,  # Thread_Class
                            'I': 15,  # Material
                            'J': 12   # Quantity
                        })
                    
                    for col, width in column_widths.items():
                        worksheet.column_dimensions[col].width = width
                    
                    # Add header formatting
                    header_fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    header_font = openpyxl.styles.Font(color="FFFFFF", bold=True)
                    
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                
                # Read the file and create download button
                with open(tmp.name, 'rb') as f:
                    st.download_button(
                        label="Download Professional Excel Template",
                        data=f.read(),
                        file_name=f"batch_weight_basic_{diameter_type.lower().replace(' ', '_')}_template.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
    
    with col2:
        if st.button("Download Advanced Template", use_container_width=True):
            template_df = BatchTemplateManager.get_advanced_template(diameter_type)
            
            # Create professional Excel file with formatting
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                with pd.ExcelWriter(tmp.name, engine='openpyxl') as writer:
                    template_df.to_excel(writer, sheet_name='Batch Template', index=False)
                    
                    # Get workbook and worksheet
                    workbook = writer.book
                    worksheet = writer.sheets['Batch Template']
                    
                    # Set column widths for better readability
                    column_widths = {
                        'A': 20,  # Product_Type
                        'B': 15,  # Product_Code
                        'C': 12,  # Series
                        'D': 18,  # Standard
                        'E': 12,  # Size
                        'F': 12,  # Grade
                        'G': 15,  # Diameter_Type
                    }
                    
                    if diameter_type == "Blank Diameter":
                        column_widths.update({
                            'H': 15,  # Diameter_Value
                            'I': 15,  # Diameter_Unit
                            'J': 18,  # Thread_Standard
                            'K': 15,  # Thread_Size
                            'L': 15,  # Thread_Class
                            'M': 12,  # Length
                            'N': 12,  # Length_Unit
                            'O': 15,  # Material
                            'P': 12   # Quantity
                        })
                    else:  # Pitch Diameter
                        column_widths.update({
                            'H': 18,  # Thread_Standard
                            'I': 18,  # Thread_Size
                            'J': 15,  # Thread_Class
                            'K': 12,  # Length
                            'L': 12,  # Length_Unit
                            'M': 15,  # Material
                            'N': 12   # Quantity
                        })
                    
                    for col, width in column_widths.items():
                        worksheet.column_dimensions[col].width = width
                    
                    # Add header formatting
                    header_fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    header_font = openpyxl.styles.Font(color="FFFFFF", bold=True)
                    
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                
                # Read the file and create download button
                with open(tmp.name, 'rb') as f:
                    st.download_button(
                        label="Download Professional Excel Template", 
                        data=f.read(),
                        file_name=f"batch_weight_advanced_{diameter_type.lower().replace(' ', '_')}_template.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
    
    st.info(f"""
    **{'Basic Mode' if st.session_state.batch_mode == 'basic' else 'Advanced Mode'} Selected with {diameter_type}:**
    - **Basic Mode**: Upload CSV with Product_Type, Product_Code, Size, Length, {'Diameter_Value, Diameter_Unit' if diameter_type == 'Blank Diameter' else 'Thread_Standard, Product_Standard, Thread_Class'} → System auto-detects other parameters
    - **Advanced Mode**: Upload CSV with complete product specifications for precise control
    - **Diameter Type**: {diameter_type} - {'Provide diameter values directly' if diameter_type == 'Blank Diameter' else 'System fetches pitch diameter from database'}
    - **New Column**: Product_Code for internal reference
    """)
    
    st.markdown("---")
    
    # File upload section
    st.markdown("### 📤 Upload Batch File")
    
    uploaded_file = st.file_uploader(
        f"Upload your {'Basic' if st.session_state.batch_mode == 'basic' else 'Advanced'} CSV/Excel file",
        type=['csv', 'xlsx'],
        key="batch_file_uploader"
    )
    
    if uploaded_file:
        st.session_state.batch_uploaded_file = uploaded_file
        
        try:
            # Load the file
            if uploaded_file.name.endswith('.xlsx'):
                batch_df = pd.read_excel(uploaded_file)
            else:
                batch_df = pd.read_csv(uploaded_file)
            
            st.success(f"✅ File uploaded successfully! Loaded {len(batch_df)} records")
            
            # Show preview
            with st.expander("📋 Preview Uploaded Data"):
                st.dataframe(batch_df.head(10), use_container_width=True)
                st.write(f"Total rows: {len(batch_df)}")
                st.write(f"Columns: {list(batch_df.columns)}")
            
            # Validate file with diameter type
            is_valid, validation_errors, validation_warnings = BatchProcessor.validate_batch_file(
                batch_df, st.session_state.batch_mode, diameter_type
            )
            
            if validation_warnings:
                for warning in validation_warnings:
                    st.warning(warning)
            
            if not is_valid:
                for error in validation_errors:
                    st.error(error)
                st.stop()
            
            # Show inferred parameters example for basic mode
            if st.session_state.batch_mode == "basic" and len(batch_df) > 0:
                with st.expander("🔍 Auto-Detection Preview"):
                    sample_row = batch_df.iloc[0]
                    inferred_params = BatchTemplateManager.infer_parameters_basic_mode(sample_row, diameter_type)
                    st.write("**Sample Auto-detected Parameters:**")
                    st.json(inferred_params)
                    st.caption("The system will automatically determine these parameters for all rows")
            
            # Process batch button
            st.markdown("---")
            st.markdown("### ⚙️ Process Batch Calculations")
            
            if st.button(
                f"🚀 Process {len(batch_df)} Records", 
                type="primary", 
                use_container_width=True,
                key="process_batch_calculations"
            ):
                st.session_state.batch_processing = True
                st.session_state.batch_processing_complete = False
                
                # Process batch calculations with diameter type
                with st.spinner(f"Processing {len(batch_df)} records with {diameter_type}..."):
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    
                    def update_progress(progress, status):
                        progress_bar.progress(progress)
                        status_text.text(status)
                    
                    results, errors, summary = BatchProcessor.process_batch_calculations(
                        batch_df, diameter_type, update_progress
                    )
                    
                    # Store results in session state
                    st.session_state.batch_results = results
                    st.session_state.batch_errors = errors
                    st.session_state.batch_summary = summary
                    st.session_state.batch_processing = False
                    st.session_state.batch_processing_complete = True
                
                progress_bar.empty()
                status_text.empty()
                
                st.rerun()
        
        except Exception as e:
            st.error(f"Error processing file: {str(e)}")
            LoadingManager.log_operation("Batch File Processing", False, str(e))
    
    # Display results if processing is complete
    if st.session_state.batch_processing_complete and st.session_state.batch_summary:
        st.markdown("---")
        st.markdown("## 📊 Batch Processing Results")
        
        # Show summary
        BatchResultsDisplay.show_processing_summary(st.session_state.batch_summary)
        
        # Show weight summary
        BatchResultsDisplay.show_weight_summary(st.session_state.batch_summary)
        
        # Show detailed results
        if st.session_state.batch_results:
            BatchResultsDisplay.show_detailed_results(st.session_state.batch_results)
            
            # Add Weight column to results and create downloadable file
            st.markdown("### 💾 Download Results with Weight Column")
            
            # Prepare results with weight column
            results_with_weight = []
            for result in st.session_state.batch_results:
                input_data = result['input_data']
                calc = result['calculation_result']
                quantity = result.get('quantity', 1)
                
                result_row = {
                    'Product_Type': input_data.get('Product_Type', ''),
                    'Product_Code': input_data.get('Product_Code', ''),
                    'Size': input_data.get('Size', ''),
                    'Length': input_data.get('Length', ''),
                    'Length_Unit': input_data.get('Length_Unit', 'mm'),
                    'Thread_Standard': input_data.get('Thread_Standard', ''),
                    'Product_Standard': input_data.get('Product_Standard', ''),
                    'Thread_Class': input_data.get('Thread_Class', ''),
                    'Material': input_data.get('Material', 'Carbon Steel'),
                    'Quantity': quantity,
                    'Weight_kg': calc['weight_kg'],
                    'Weight_lb': calc['weight_lb'],
                    'Total_Weight_kg': calc['weight_kg'] * quantity,
                    'Total_Weight_lb': calc['weight_lb'] * quantity
                }
                results_with_weight.append(result_row)
            
            results_df = pd.DataFrame(results_with_weight)
            
            # Create professional Excel file with results
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                with pd.ExcelWriter(tmp.name, engine='openpyxl') as writer:
                    results_df.to_excel(writer, sheet_name='Weight Results', index=False)
                    
                    # Get workbook and worksheet
                    workbook = writer.book
                    worksheet = writer.sheets['Weight Results']
                    
                    # Set column widths for better readability
                    column_widths = {
                        'A': 20,  # Product_Type
                        'B': 15,  # Product_Code
                        'C': 12,  # Size
                        'D': 12,  # Length
                        'E': 12,  # Length_Unit
                        'F': 18,  # Thread_Standard
                        'G': 18,  # Product_Standard
                        'H': 15,  # Thread_Class
                        'I': 15,  # Material
                        'J': 12,  # Quantity
                        'K': 15,  # Weight_kg
                        'L': 15,  # Weight_lb
                        'M': 18,  # Total_Weight_kg
                        'N': 18   # Total_Weight_lb
                    }
                    
                    for col, width in column_widths.items():
                        worksheet.column_dimensions[col].width = width
                    
                    # Add header formatting
                    header_fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    header_font = openpyxl.styles.Font(color="FFFFFF", bold=True)
                    
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                
                # Download button for professional results
                with open(tmp.name, 'rb') as f:
                    st.download_button(
                        label="📥 Download Professional Results with Weight Column",
                        data=f.read(),
                        file_name="batch_weight_results_professional.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
        
        # Show errors
        if st.session_state.batch_errors:
            BatchResultsDisplay.show_error_report(st.session_state.batch_errors)
        
        # Export section
        st.markdown("---")
        st.markdown("### 💾 Export Results")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            if st.button("📥 Export to Excel", use_container_width=True):
                with st.spinner("Generating Excel report..."):
                    file_path, filename = BatchResultsDisplay.export_batch_results(
                        st.session_state.batch_results,
                        st.session_state.batch_errors,
                        st.session_state.batch_summary
                    )
                    
                    if file_path:
                        with open(file_path, 'rb') as f:
                            st.download_button(
                                label="Download Excel Report",
                                data=f,
                                file_name=filename,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True
                            )
        
        with col2:
            # Export successful results only
            if st.session_state.batch_results:
                successful_df = pd.DataFrame([
                    {
                        'Product_Type': r['input_data'].get('Product_Type', 'Auto-detected'),
                        'Product_Code': r['input_data'].get('Product_Code', ''),
                        'Size': r['input_data'].get('Size', 'N/A'),
                        'Length': f"{r['input_data'].get('Length', 'N/A')} {r['input_data'].get('Length_Unit', 'mm')}",
                        'Material': r['input_data'].get('Material', 'Carbon Steel'),
                        'Diameter_Type': r['input_data'].get('Diameter_Type', 'Blank Diameter'),
                        'Weight_kg': r['calculation_result']['weight_kg'],
                        'Weight_lb': r['calculation_result']['weight_lb'],
                        'Quantity': r.get('quantity', 1),
                        'Total_Weight_kg': r['calculation_result']['weight_kg'] * r.get('quantity', 1),
                        'Total_Weight_lb': r['calculation_result']['weight_lb'] * r.get('quantity', 1)
                    }
                    for r in st.session_state.batch_results
                ])
                
                csv_data = successful_df.to_csv(index=False)
                st.download_button(
                    label="📊 Download CSV Summary",
                    data=csv_data,
                    file_name="batch_weight_results.csv",
                    mime="text/csv",
                    use_container_width=True
                )
        
        with col3:
            if st.button("🔄 Process New Batch", use_container_width=True):
                # Reset batch state
                st.session_state.batch_uploaded_file = None
                st.session_state.batch_processing = False
                st.session_state.batch_results = None
                st.session_state.batch_errors = []
                st.session_state.batch_summary = None
                st.session_state.batch_processing_complete = False
                st.rerun()

# ======================================================
# FIXED THREAD DATA LOADING - PROPER DATA TYPES
# ======================================================
@st.cache_data(ttl=3600)
def load_thread_data_enhanced(standard_name):
    """Enhanced thread data loading with proper data type handling"""
    if standard_name not in thread_files:
        return pd.DataFrame()
    
    file_path = thread_files[standard_name]
    try:
        df_thread = safe_load_excel_file_enhanced(file_path)
        if df_thread.empty:
            st.warning(f"Thread data for {standard_name} is empty")
            return pd.DataFrame()
        
        # Clean column names
        df_thread.columns = [str(col).strip() for col in df_thread.columns]
        
        # Debug: Show column info
        if st.session_state.debug_mode:
            st.sidebar.write(f"Columns {standard_name}:", df_thread.columns.tolist())
            st.sidebar.write(f"Shape {standard_name}:", df_thread.shape)
        
        # Handle different column naming patterns
        thread_col = None
        class_col = None
        
        # Find thread size column
        possible_thread_cols = ['Thread', 'Size', 'Thread Size', 'Nominal Size', 'Basic Major Diameter']
        for col in df_thread.columns:
            col_lower = str(col).lower()
            for possible in possible_thread_cols:
                if possible.lower() in col_lower:
                    thread_col = col
                    break
            if thread_col:
                break
        
        # Find class/tolerance column
        possible_class_cols = ['Class', 'Tolerance', 'Tolerance Class', 'Thread Class']
        for col in df_thread.columns:
            col_lower = str(col).lower()
            for possible in possible_class_cols:
                if possible.lower() in col_lower:
                    class_col = col
                    break
            if class_col:
                break
        
        # If no specific class column found, check for columns containing tolerance info
        if not class_col:
            for col in df_thread.columns:
                if 'tolerance' in str(col).lower() or 'class' in str(col).lower():
                    class_col = col
                    break
        
        # Standardize column names for consistent processing
        if thread_col:
            df_thread = df_thread.rename(columns={thread_col: 'Thread'})
        
        if class_col:
            df_thread = df_thread.rename(columns={class_col: 'Class'})
        
        # Clean data - convert all to string and handle NaN
        if 'Thread' in df_thread.columns:
            df_thread['Thread'] = df_thread['Thread'].astype(str).str.strip()
            df_thread = df_thread[df_thread['Thread'] != 'nan']
            df_thread = df_thread[df_thread['Thread'] != '']
        
        if 'Class' in df_thread.columns:
            df_thread['Class'] = df_thread['Class'].astype(str).str.strip()
            df_thread = df_thread[df_thread['Class'] != 'nan']
            df_thread = df_thread[df_thread['Class'] != '']
        
        # Add standard identifier
        df_thread['Standard'] = standard_name
        
        LoadingManager.log_operation(f"Load Thread Data: {standard_name}", True, f"Records: {len(df_thread)}")
        return df_thread
        
    except Exception as e:
        st.error(f"Error loading thread data for {standard_name}: {str(e)}")
        LoadingManager.log_operation(f"Load Thread Data: {standard_name}", False, str(e))
        return pd.DataFrame()

def get_thread_data_enhanced(standard, thread_size=None, thread_class=None):
    """Enhanced thread data retrieval with proper filtering"""
    df_thread = load_thread_data_enhanced(standard)
    
    if df_thread.empty:
        return pd.DataFrame()
    
    # Apply filters if provided
    result_df = df_thread.copy()
    
    if thread_size and thread_size != "All" and "Thread" in result_df.columns:
        result_df = result_df[result_df["Thread"].astype(str).str.strip() == str(thread_size).strip()]
    
    if thread_class and thread_class != "All" and "Class" in result_df.columns:
        result_df = result_df[
            result_df["Class"].astype(str).str.strip().str.upper() == 
            str(thread_class).strip().upper()
        ]
    
    return result_df

def get_thread_sizes_enhanced(standard):
    """Get available thread sizes with proper data handling"""
    df_thread = load_thread_data_enhanced(standard)
    
    if df_thread.empty or "Thread" not in df_thread.columns:
        return ["All"]
    
    try:
        # Get unique sizes and handle NaN values properly
        unique_sizes = df_thread['Thread'].dropna().unique()
        
        # Convert all sizes to string and filter out empty strings
        unique_sizes = [str(size).strip() for size in unique_sizes if str(size).strip() != '']
        
        if len(unique_sizes) > 0:
            sorted_sizes = safe_sort_sizes(unique_sizes)
            return ["All"] + sorted_sizes
        else:
            return ["All"]
    except Exception as e:
        st.warning(f"Thread size processing warning for {standard}: {str(e)}")
        return ["All"]

def get_thread_classes_enhanced(standard):
    """Get available thread classes with proper data handling"""
    df_thread = load_thread_data_enhanced(standard)
    
    if df_thread.empty or "Class" not in df_thread.columns:
        return ["All"]
    
    try:
        # Get unique classes and handle NaN values properly
        unique_classes = df_thread['Class'].dropna().unique()
        
        # Convert all classes to string and filter out empty strings
        unique_classes = [str(cls).strip() for cls in unique_classes if str(cls).strip() != '']
        
        if len(unique_classes) > 0:
            sorted_classes = sorted(unique_classes)
            return ["All"] + sorted_classes
        else:
            return ["All"]
    except Exception as e:
        st.warning(f"Thread class processing warning for {standard}: {str(e)}")
        return ["All"]

# ======================================================
# ORACLE 11G STYLING - COMPLETE UI TRANSFORMATION
# ======================================================
st.set_page_config(
    page_title="JSC Industries - Fastener Intelligence", 
    layout="wide",
    page_icon="🔧",
    initial_sidebar_state="expanded"
)

# Oracle 11g Professional CSS
st.markdown("""
<style>
    :root {
        --oracle11g-blue: #1F4E78;
        --oracle11g-blue-dark: #0D2B4A;
        --oracle11g-blue-light: #2E75B6;
        --oracle11g-gray-dark: #333333;
        --oracle11g-gray: #666666;
        --oracle11g-gray-light: #F5F5F5;
        --oracle11g-gray-border: #CCCCCC;
        --oracle11g-orange: #E66C37;
        --oracle11g-green: #4CAF50;
        --oracle11g-yellow: #FFC107;
        --oracle11g-red: #D32F2F;
    }
    
    .stApp {
        background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%);
        font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
    }
    
    .oracle11g-header {
        background: linear-gradient(135deg, var(--oracle11g-blue) 0%, var(--oracle11g-blue-dark) 100%);
        padding: 2.5rem;
        border-radius: 8px;
        color: white;
        text-align: center;
        margin-bottom: 2rem;
        box-shadow: 0 6px 18px rgba(31, 78, 120, 0.3);
        border: 1px solid var(--oracle11g-blue-light);
        position: relative;
        overflow: hidden;
    }
    
    .oracle11g-header::before {
        content: '';
        position: absolute;
        top: 0;
        left: 0;
        right: 0;
        height: 4px;
        background: linear-gradient(90deg, var(--oracle11g-orange) 0%, var(--oracle11g-yellow) 100%);
    }
    
    .oracle11g-header h1 {
        font-size: 2.5rem;
        font-weight: 700;
        margin-bottom: 0.5rem;
        color: white;
        text-shadow: 0 2px 4px rgba(0,0,0,0.3);
    }
    
    .oracle11g-header p {
        font-size: 1.2rem;
        opacity: 0.95;
        margin-bottom: 1rem;
        color: white;
        font-weight: 300;
    }
    
    .oracle11g-card {
        background: white;
        padding: 1.8rem;
        border-radius: 8px;
        box-shadow: 0 4px 12px rgba(0, 0, 0, 0.1);
        border-left: 5px solid var(--oracle11g-blue);
        transition: all 0.3s ease;
        margin-bottom: 1.2rem;
        border: 1px solid var(--oracle11g-gray-border);
        position: relative;
        overflow: hidden;
    }
    
    .oracle11g-card::before {
        content: '';
        position: absolute;
        top: 0;
        left: 0;
        width: 100%;
        height: 3px;
        background: var(--oracle11g-blue);
    }
    
    .oracle11g-card:hover {
        transform: translateY(-3px);
        box-shadow: 0 8px 25px rgba(31, 78, 120, 0.15);
        border-left-color: var(--oracle11g-blue-light);
    }
    
    .metric-card {
        background: white;
        padding: 1.8rem;
        border-radius: 8px;
        box-shadow: 0 4px 12px rgba(0, 0, 0, 0.1);
        border-left: 5px solid var(--oracle11g-blue);
        transition: transform 0.3s ease;
        border: 1px solid var(--oracle11g-gray-border);
        text-align: center;
    }
    
    .metric-card:hover {
        transform: translateY(-3px);
        box-shadow: 0 8px 25px rgba(31, 78, 120, 0.15);
    }
    
    .oracle11g-badge {
        background: linear-gradient(135deg, var(--oracle11g-blue) 0%, var(--oracle11g-blue-light) 100%);
        color: white;
        padding: 0.4rem 1rem;
        border-radius: 20px;
        font-size: 0.8rem;
        font-weight: 600;
        margin: 0.2rem;
        display: inline-block;
        box-shadow: 0 2px 6px rgba(31, 78, 120, 0.3);
        border: 1px solid var(--oracle11g-blue-light);
    }
    
    .oracle11g-badge-orange {
        background: linear-gradient(135deg, var(--oracle11g-orange) 0%, #F57C00 100%);
        border: 1px solid var(--oracle11g-orange);
    }
    
    .oracle11g-badge-green {
        background: linear-gradient(135deg, var(--oracle11g-green) 0%, #388E3C 100%);
        border: 1px solid var(--oracle11g-green);
    }
    
    .oracle11g-badge-yellow {
        background: linear-gradient(135deg, var(--oracle11g-yellow) 0%, #FFA000 100%);
        color: #333;
        border: 1px solid var(--oracle11g-yellow);
    }
    
    .stButton>button {
        background: linear-gradient(135deg, var(--oracle11g-blue) 0%, var(--oracle11g-blue-light) 100%);
        color: white;
        border: none;
        padding: 0.8rem 1.5rem;
        border-radius: 6px;
        font-weight: 600;
        transition: all 0.3s ease;
        box-shadow: 0 4px 12px rgba(31, 78, 120, 0.3);
        font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
        border: 1px solid var(--oracle11g-blue-light);
    }
    
    .stButton>button:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 20px rgba(31, 78, 120, 0.4);
        background: linear-gradient(135deg, var(--oracle11g-blue-light) 0%, var(--oracle11g-blue) 100%);
    }
    
    .stButton>button:active {
        transform: translateY(0);
    }
    
    .stButton>button[kind="primary"] {
        background: linear-gradient(135deg, var(--oracle11g-blue) 0%, var(--oracle11g-blue-light) 100%);
        color: white;
    }
    
    .stButton>button[kind="primary"]:hover {
        background: linear-gradient(135deg, var(--oracle11g-blue-light) 0%, var(--oracle11g-blue) 100%);
    }
    
    .stButton>button[kind="secondary"] {
        background: white;
        color: var(--oracle11g-blue);
        border: 2px solid var(--oracle11g-blue);
    }
    
    .stButton>button[kind="secondary"]:hover {
        background: var(--oracle11g-blue);
        color: white;
    }
    
    .css-1d391kg, .css-1lcbmhc {
        background: linear-gradient(180deg, var(--oracle11g-gray-light) 0%, white 100%);
        border-right: 2px solid var(--oracle11g-gray-border);
    }
    
    .sidebar .sidebar-content {
        background: linear-gradient(180deg, white 0%, var(--oracle11g-gray-light) 100%);
    }
    
    .stTextInput>div>div>input, 
    .stNumberInput>div>div>input,
    .stSelectbox>div>div>select {
        border: 2px solid var(--oracle11g-gray-border);
        border-radius: 6px;
        padding: 0.7rem 1rem;
        transition: all 0.3s ease;
        font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
        background: white;
    }
    
    .stTextInput>div>div>input:focus, 
    .stNumberInput>div>div>input:focus,
    .stSelectbox>div>div>select:focus {
        border-color: var(--oracle11g-blue);
        box-shadow: 0 0 0 3px rgba(31, 78, 120, 0.1);
        background: white;
    }
    
    .dataframe {
        border-radius: 8px;
        overflow: hidden;
        box-shadow: 0 4px 12px rgba(0,0,0,0.1);
        border: 1px solid var(--oracle11g-gray-border);
    }
    
    .streamlit-expanderHeader {
        background: var(--oracle11g-gray-light);
        border-radius: 6px;
        border: 2px solid var(--oracle11g-gray-border);
        font-weight: 600;
        color: var(--oracle11g-blue);
        padding: 1rem;
    }
    
    .stTabs [data-baseweb="tab-list"] {
        gap: 1rem;
    }
    
    .stTabs [data-baseweb="tab"] {
        background: white;
        border-radius: 6px 6px 0 0;
        padding: 1rem 2rem;
        border: 2px solid var(--oracle11g-gray-border);
        border-bottom: none;
        font-weight: 600;
        color: var(--oracle11g-gray);
        transition: all 0.3s ease;
    }
    
    .stTabs [aria-selected="true"] {
        background: var(--oracle11g-blue);
        color: white;
        border-color: var(--oracle11g-blue);
    }
    
    .stProgress > div > div > div {
        background: linear-gradient(90deg, var(--oracle11g-blue) 0%, var(--oracle11g-blue-light) 100%);
    }
    
    .stAlert {
        border-radius: 8px;
        border: 2px solid;
        padding: 1rem;
    }
    
    .stAlert [data-testid="stMarkdownContainer"] {
        font-weight: 500;
    }
    
    .section-header {
        border-left: 5px solid var(--oracle11g-blue);
        padding-left: 1.2rem;
        margin: 2rem 0 1.5rem 0;
        color: var(--oracle11g-blue);
        font-weight: 700;
        font-size: 1.5rem;
        text-shadow: 0 1px 2px rgba(0,0,0,0.1);
    }
    
    .quick-action {
        background: white;
        padding: 1.5rem 1rem;
        border-radius: 8px;
        text-align: center;
        cursor: pointer;
        transition: all 0.3s ease;
        box-shadow: 0 4px 12px rgba(0,0,0,0.1);
        border: 2px solid var(--oracle11g-gray-border);
        height: 120px;
        display: flex;
        flex-direction: column;
        justify-content: center;
        align-items: center;
        position: relative;
        overflow: hidden;
    }
    
    .quick-action::before {
        content: '';
        position: absolute;
        top: 0;
        left: 0;
        width: 100%;
        height: 4px;
        background: var(--oracle11g-blue);
    }
    
    .quick-action:hover {
        transform: translateY(-3px);
        box-shadow: 0 8px 25px rgba(0,0,0,0.15);
        border-color: var(--oracle11g-blue);
    }
    
    .professional-card {
        background: linear-gradient(135deg, #ffffff 0%, var(--oracle11g-gray-light) 100%);
        border: 2px solid var(--oracle11g-blue);
        border-radius: 10px;
        padding: 2rem;
        margin: 2rem 0;
        box-shadow: 0 8px 30px rgba(31, 78, 120, 0.2);
        position: relative;
        overflow: hidden;
    }
    
    .professional-card::before {
        content: '';
        position: absolute;
        top: 0;
        left: 0;
        right: 0;
        height: 4px;
        background: linear-gradient(90deg, var(--oracle11g-blue) 0%, var(--oracle11g-orange) 100%);
    }
    
    .card-header {
        display: flex;
        justify-content: space-between;
        align-items: flex-start;
        margin-bottom: 2rem;
        padding-bottom: 1.5rem;
        border-bottom: 2px solid var(--oracle11g-gray-border);
    }
    
    .card-title {
        font-size: 1.8rem;
        font-weight: 700;
        color: var(--oracle11g-blue);
        margin: 0;
        text-shadow: 0 1px 2px rgba(0,0,0,0.1);
    }
    
    .card-subtitle {
        font-size: 1.1rem;
        color: var(--oracle11g-gray);
        margin: 0.5rem 0 0 0;
        font-weight: 400;
    }
    
    .card-company {
        background: var(--oracle11g-blue);
        color: white;
        padding: 0.6rem 1.2rem;
        border-radius: 6px;
        font-weight: 700;
        font-size: 0.9rem;
        box-shadow: 0 2px 8px rgba(31, 78, 120, 0.3);
    }
    
    .spec-row {
        display: grid;
        grid-template-columns: 1fr auto 1fr;
        gap: 1rem;
        align-items: center;
        margin: 0.8rem 0;
        padding: 0.8rem;
        border-radius: 6px;
        background: var(--oracle11g-gray-light);
        border: 1px solid var(--oracle11g-gray-border);
    }
    
    .spec-label-min, .spec-label-max {
        font-size: 0.9rem;
        color: var(--oracle11g-gray);
        text-align: center;
        font-weight: 600;
    }
    
    .spec-dimension {
        font-weight: 700;
        color: var(--oracle11g-blue);
        text-align: center;
        padding: 0.5rem 1rem;
        background: white;
        border-radius: 4px;
        border: 2px solid var(--oracle11g-gray-border);
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
    }
    
    .spec-value {
        font-weight: 700;
        color: var(--oracle11g-blue);
        text-align: center;
        padding: 0.5rem;
        background: white;
        border-radius: 4px;
        border: 2px solid var(--oracle11g-gray-border);
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
    }
    
    .card-footer {
        display: flex;
        justify-content: space-between;
        align-items: center;
        margin-top: 2rem;
        padding-top: 1.5rem;
        border-top: 2px solid var(--oracle11g-gray-border);
        font-size: 0.9rem;
        color: var(--oracle11g-gray);
    }
    
    .card-actions {
        display: flex;
        gap: 1rem;
        margin-top: 1.5rem;
        justify-content: center;
    }
    
    .action-button {
        background: linear-gradient(135deg, var(--oracle11g-blue) 0%, var(--oracle11g-blue-light) 100%);
        color: white;
        border: none;
        padding: 0.8rem 1.5rem;
        border-radius: 6px;
        font-weight: 600;
        cursor: pointer;
        transition: all 0.3s ease;
        box-shadow: 0 4px 12px rgba(31, 78, 120, 0.3);
        border: 1px solid var(--oracle11g-blue-light);
    }
    
    .action-button:hover {
        background: linear-gradient(135deg, var(--oracle11g-blue-light) 0%, var(--oracle11g-blue) 100%);
        transform: translateY(-2px);
        box-shadow: 0 6px 20px rgba(31, 78, 120, 0.4);
    }
    
    .action-button.secondary {
        background: white;
        color: var(--oracle11g-blue);
        border: 2px solid var(--oracle11g-blue);
    }
    
    .action-button.secondary:hover {
        background: var(--oracle11g-blue);
        color: white;
    }
    
    .filter-section {
        background: white;
        padding: 1.5rem;
        border-radius: 8px;
        box-shadow: 0 4px 12px rgba(0, 0, 0, 0.1);
        margin-bottom: 1.5rem;
        border: 2px solid var(--oracle11g-gray-border);
        position: relative;
        overflow: hidden;
    }
    
    .filter-section::before {
        content: '';
        position: absolute;
        top: 0;
        left: 0;
        width: 100%;
        height: 4px;
        background: var(--oracle11g-blue);
    }
    
    .filter-header {
        border-left: 4px solid var(--oracle11g-blue);
        padding-left: 1rem;
        margin-bottom: 1rem;
        color: var(--oracle11g-blue);
        font-weight: 700;
        font-size: 1.3rem;
    }
    
    .independent-section {
        border: 2px solid var(--oracle11g-blue);
        border-radius: 8px;
        padding: 1.5rem;
        margin-bottom: 2rem;
        background: linear-gradient(135deg, var(--oracle11g-gray-light) 0%, #ffffff 100%);
        position: relative;
        overflow: hidden;
    }
    
    .independent-section::before {
        content: '';
        position: absolute;
        top: 0;
        left: 0;
        width: 100%;
        height: 4px;
        background: var(--oracle11g-blue);
    }
    
    .section-results {
        border: 2px solid var(--oracle11g-green);
        border-radius: 8px;
        padding: 1.5rem;
        margin-bottom: 1.5rem;
        background: linear-gradient(135deg, #f0f8f0 0%, #ffffff 100%);
    }
    
    .combined-results {
        border: 2px solid var(--oracle11g-orange);
        border-radius: 8px;
        padding: 1.5rem;
        margin-bottom: 1.5rem;
        background: linear-gradient(135deg, #fff8f0 0%, #ffffff 100%);
    }
    
    .data-quality-indicator {
        padding: 0.8rem;
        border-radius: 6px;
        margin: 0.3rem 0;
        font-size: 0.9rem;
        border-left: 4px solid;
        background: white;
        border: 1px solid var(--oracle11g-gray-border);
    }
    
    .quality-good {
        background: #e8f5e8;
        color: #2e7d32;
        border-left-color: var(--oracle11g-green);
    }
    
    .quality-warning {
        background: #fff8e1;
        color: #f57c00;
        border-left-color: var(--oracle11g-yellow);
    }
    
    .quality-error {
        background: #ffebee;
        color: #c62828;
        border-left-color: var(--oracle11g-red);
    }
    
    .calculation-card {
        background: linear-gradient(135deg, var(--oracle11g-gray-light) 0%, #e9ecef 100%);
        padding: 1rem;
        border-radius: 8px;
        margin: 0.5rem 0;
        border-left: 4px solid var(--oracle11g-green);
        border: 1px solid var(--oracle11g-gray-border);
    }
    
    .spec-grid {
        display: grid;
        grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
        gap: 1rem;
        margin: 1rem 0;
    }
    
    .property-grid {
        display: grid;
        grid-template-columns: repeat(auto-fit, minmax(150px, 1fr));
        gap: 0.8rem;
        margin: 1rem 0;
    }
    
    .specification-grid {
        display: grid;
        grid-template-columns: repeat(auto-fit, minmax(300px, 1fr));
        gap: 1.5rem;
        margin: 1.5rem 0;
    }
    
    .oracle11g-footer {
        text-align: center;
        color: var(--oracle11g-gray);
        padding: 2rem;
        margin-top: 3rem;
        border-top: 2px solid var(--oracle11g-gray-border);
        background: var(--oracle11g-gray-light);
        border-radius: 8px;
    }
    
    /* Loading spinner enhancements */
    .stSpinner > div {
        border: 4px solid #f3f3f3;
        border-radius: 50%;
        border-top: 4px solid var(--oracle11g-blue);
        width: 40px;
        height: 40px;
        animation: spin 2s linear infinite;
        margin: 0 auto;
    }
    
    @keyframes spin {
        0% { transform: rotate(0deg); }
        100% { transform: rotate(360deg); }
    }
    
    /* Mobile optimizations */
    @media (max-width: 768px) {
        .oracle11g-header {
            padding: 1.5rem !important;
        }
        
        .oracle11g-header h1 {
            font-size: 2rem !important;
        }
        
        .spec-grid,
        .property-grid,
        .specification-grid {
            grid-template-columns: 1fr;
        }
        
        .card-header {
            flex-direction: column;
            gap: 1rem;
            text-align: center;
        }
        
        .stTabs [data-baseweb="tab"] {
            padding: 0.8rem 1rem;
        }
        
        .spec-row {
            grid-template-columns: 1fr;
            gap: 0.5rem;
        }
        
        .card-actions {
            flex-direction: column;
        }
        
        .stButton > button {
            padding: 0.7rem 1.2rem !important;
            font-size: 0.9rem !important;
        }
        
        .metric-card {
            padding: 1.2rem !important;
            margin-bottom: 0.8rem !important;
        }
        
        .quick-action {
            height: 100px !important;
            padding: 1rem 0.6rem !important;
        }
        
        .stSelectbox, .stTextInput, .stNumberInput {
            font-size: 0.9rem !important;
        }
    }
</style>
""", unsafe_allow_html=True)

# Initialize session state
initialize_session_state()
optimize_for_mobile()

# ======================================================
# ENHANCED DATA LOADING WITH PRODUCT MAPPING
# ======================================================

# Load main data
with LoadingManager.show_loading_spinner("Loading main fastener data..."):
    df = safe_load_excel_file_enhanced(url) if url else safe_load_excel_file_enhanced(local_excel_path)

# Load Mechanical and Chemical data
with LoadingManager.show_loading_spinner("Loading mechanical & chemical data..."):
    df_mechem = safe_load_excel_file_enhanced(me_chem_google_url)
    if df_mechem.empty:
        st.info("Online Mechanical & Chemical file not accessible, trying local version...")
        df_mechem = safe_load_excel_file_enhanced(me_chem_path)

# Load ISO 4014 data
with LoadingManager.show_loading_spinner("Loading ISO 4014 data..."):
    df_iso4014 = safe_load_excel_file_enhanced(iso4014_file_url)
    if df_iso4014.empty:
        st.info("Online ISO 4014 file not accessible, trying local version...")
        df_iso4014 = safe_load_excel_file_enhanced(iso4014_local_path)

# Load DIN-7991 data
with LoadingManager.show_loading_spinner("Loading DIN-7991 data..."):
    df_din7991 = safe_load_excel_file_enhanced(din7991_file_url)
    if df_din7991.empty:
        st.info("Online DIN-7991 file not accessible, trying local version...")
        df_din7991 = safe_load_excel_file_enhanced(din7991_local_path)

# Load ASME B18.3 data
with LoadingManager.show_loading_spinner("Loading ASME B18.3 data..."):
    df_asme_b18_3 = safe_load_excel_file_enhanced(asme_b18_3_file_url)
    if df_asme_b18_3.empty:
        st.info("Online ASME B18.3 file not accessible, trying local version...")
        df_asme_b18_3 = safe_load_excel_file_enhanced(asme_b18_3_local_path)

# ======================================================
# FIXED DATA PROCESSING - CORRECT PRODUCT NAMES
# ======================================================

def process_standard_data():
    """FIXED VERSION: Get ACTUAL product names from Excel files"""
    
    standard_products = {}
    standard_series = {}
    
    # Process ASME B18.2.1 - Get ACTUAL products from Excel
    if not df.empty:
        if 'Product' in df.columns:
            # Get ACTUAL unique products from the Excel data
            asme_products = df['Product'].dropna().unique().tolist()
            # Clean and sort the actual product names
            cleaned_products = [str(p).strip() for p in asme_products if p and str(p).strip() != '']
            standard_products['ASME B18.2.1'] = ["All"] + sorted(cleaned_products)
        else:
            # Fallback if no Product column
            standard_products['ASME B18.2.1'] = ["All", "Hex Bolt", "Heavy Hex Bolt", "Hex Cap Screws", "Heavy Hex Screws"]
        standard_series['ASME B18.2.1'] = "Inch"
    
    # Process ASME B18.3 Data - Get ACTUAL products
    if not df_asme_b18_3.empty:
        if 'Product' in df_asme_b18_3.columns:
            asme_b18_3_products = df_asme_b18_3['Product'].dropna().unique().tolist()
            cleaned_products = [str(p).strip() for p in asme_b18_3_products if p and str(p).strip() != '']
            standard_products['ASME B18.3'] = ["All"] + sorted(cleaned_products)
        else:
            standard_products['ASME B18.3'] = ["All", "Hexagon Socket Head Cap Screws"]
        standard_series['ASME B18.3'] = "Inch"
    
    # Process DIN-7991 Data - Get ACTUAL products
    if not df_din7991.empty:
        if 'Product' in df_din7991.columns:
            din_products = df_din7991['Product'].dropna().unique().tolist()
            cleaned_products = [str(p).strip() for p in din_products if p and str(p).strip() != '']
            standard_products['DIN-7991'] = ["All"] + sorted(cleaned_products)
        else:
            standard_products['DIN-7991'] = ["All", "Hexagon Socket Countersunk Head Cap Screw"]
        standard_series['DIN-7991'] = "Metric"
    
    # Process ISO 4014 Data - Get ACTUAL products
    if not df_iso4014.empty:
        product_col = None
        for col in df_iso4014.columns:
            if 'product' in col.lower():
                product_col = col
                break
        
        if product_col:
            iso_products = df_iso4014[product_col].dropna().unique().tolist()
            cleaned_products = [str(p).strip() for p in iso_products if p and str(p).strip() != '']
            standard_products['ISO 4014'] = ["All"] + sorted(cleaned_products)
        else:
            standard_products['ISO 4014'] = ["All", "Hex Bolt"]
        standard_series['ISO 4014'] = "Metric"
    
    # ADD THREADED ROD TO ALL STANDARDS
    for standard in standard_products:
        if "Threaded Rod" not in standard_products[standard]:
            standard_products[standard] = ["All", "Threaded Rod"] + [p for p in standard_products[standard] if p != "All" and p != "Threaded Rod"]
    
    # Store in session state
    st.session_state.available_products = standard_products
    st.session_state.available_series = standard_series
    
    # Count dimensional standards
    dimensional_standards_count = 0
    if not df.empty:
        dimensional_standards_count += 1
    if not df_iso4014.empty:
        dimensional_standards_count += 1
    if not df_din7991.empty:
        dimensional_standards_count += 1
    if not df_asme_b18_3.empty:
        dimensional_standards_count += 1
    
    st.session_state.dimensional_standards_count = dimensional_standards_count
    
    return standard_products, standard_series

# Process all standards data
with LoadingManager.show_loading_spinner("Processing standards data..."):
    standard_products, standard_series = process_standard_data()

# Process DIN-7991 data if loaded
if not df_din7991.empty:
    if 'Product' not in df_din7991.columns:
        df_din7991['Product'] = "Hexagon Socket Countersunk Head Cap Screw"
    if 'Standards' not in df_din7991.columns:
        df_din7991['Standards'] = "DIN-7991"
    st.session_state.din7991_loaded = True
else:
    st.session_state.din7991_loaded = False

# Process ASME B18.3 data if loaded
if not df_asme_b18_3.empty:
    if 'Product' not in df_asme_b18_3.columns:
        df_asme_b18_3['Product'] = "Hexagon Socket Head Cap Screws"
    if 'Standards' not in df_asme_b18_3.columns:
        df_asme_b18_3['Standards'] = "ASME B18.3"
    st.session_state.asme_b18_3_loaded = True
else:
    st.session_state.asme_b18_3_loaded = False

# Process ISO 4014 data
if not df_iso4014.empty:
    product_col = None
    for col in df_iso4014.columns:
        if 'product' in col.lower():
            product_col = col
            break
    
    if product_col:
        df_iso4014['Product'] = df_iso4014[product_col]
    else:
        df_iso4014['Product'] = "Hex Bolt"
    
    df_iso4014['Standards'] = "ISO-4014-2011"
    
    grade_col = None
    for col in df_iso4014.columns:
        if 'grade' in col.lower():
            grade_col = col
            break
    
    if grade_col and grade_col != 'Product Grade':
        df_iso4014['Product Grade'] = df_iso4014[grade_col]

# ======================================================
# ENHANCED MECHANICAL & CHEMICAL DATA PROCESSING - COMPLETELY FIXED
# ======================================================
def process_mechanical_chemical_data():
    """Process and extract ALL property classes from Mechanical & Chemical data - COMPLETELY FIXED"""
    if df_mechem.empty:
        return [], []
    
    try:
        me_chem_columns = df_mechem.columns.tolist()
        
        # Find ALL possible property class columns
        property_class_cols = []
        possible_class_cols = ['Grade', 'Class', 'Property Class', 'Material Grade', 'Type', 'Designation', 'Material']
        
        for col in me_chem_columns:
            col_lower = str(col).lower()
            for possible in possible_class_cols:
                if possible.lower() in col_lower:
                    property_class_cols.append(col)
                    break
        
        # If no specific class columns found, use first few columns that have string data
        if not property_class_cols:
            for col in me_chem_columns[:3]:  # Check first 3 columns
                if df_mechem[col].dtype == 'object':  # String/object type
                    property_class_cols.append(col)
                    break
        
        # Collect ALL unique property classes from ALL identified columns
        all_property_classes = set()
        
        for prop_col in property_class_cols:
            if prop_col in df_mechem.columns:
                unique_classes = df_mechem[prop_col].dropna().unique()
                # Clean and add all classes
                for cls in unique_classes:
                    if pd.notna(cls) and str(cls).strip() != '':
                        all_property_classes.add(str(cls).strip())
        
        # Convert to sorted list
        property_classes = sorted(list(all_property_classes))
        
        st.session_state.me_chem_columns = me_chem_columns
        st.session_state.property_classes = property_classes
        
        # Debug info
        if st.session_state.debug_mode:
            st.sidebar.write(f"Found {len(property_classes)} property classes")
            st.sidebar.write(f"Property class columns: {property_class_cols}")
        
        LoadingManager.log_operation("Process Mechanical & Chemical Data", True, f"Property Classes: {len(property_classes)}")
        return me_chem_columns, property_classes
        
    except Exception as e:
        st.error(f"Error processing Mechanical & Chemical data: {str(e)}")
        LoadingManager.log_operation("Process Mechanical & Chemical Data", False, str(e))
        return [], []

def get_standards_for_property_class(property_class):
    """Get available standards for a specific property class - COMPLETELY FIXED"""
    if df_mechem.empty or not property_class or property_class == "All":
        return []
    
    try:
        # Find ALL possible standard columns
        standard_cols = []
        possible_standard_cols = ['Standard', 'Specification', 'Norm', 'Type', 'Designation']
        
        for col in df_mechem.columns:
            col_lower = str(col).lower()
            for possible in possible_standard_cols:
                if possible.lower() in col_lower:
                    standard_cols.append(col)
                    break
        
        # If no standard columns found, look for any column that might contain standard info
        if not standard_cols:
            for col in df_mechem.columns:
                if any(word in col.lower() for word in ['iso', 'astm', 'asme', 'din', 'bs', 'jis', 'gb']):
                    standard_cols.append(col)
                    break
        
        # Find ALL possible property class columns
        property_class_cols = []
        possible_class_cols = ['Grade', 'Class', 'Property Class', 'Material Grade', 'Type', 'Designation', 'Material']
        
        for col in df_mechem.columns:
            col_lower = str(col).lower()
            for possible in possible_class_cols:
                if possible.lower() in col_lower:
                    property_class_cols.append(col)
                    break
        
        # If no specific class columns found, use first few columns
        if not property_class_cols:
            for col in df_mechem.columns[:3]:
                if df_mechem[col].dtype == 'object':
                    property_class_cols.append(col)
                    break
        
        # Try to find matching data using ALL property class columns
        matching_standards = set()
        
        for prop_col in property_class_cols:
            if prop_col in df_mechem.columns:
                # Try exact match first
                exact_match = df_mechem[df_mechem[prop_col] == property_class]
                if not exact_match.empty:
                    for std_col in standard_cols:
                        if std_col in exact_match.columns:
                            standards = exact_match[std_col].dropna().unique()
                            for std in standards:
                                if pd.notna(std) and str(std).strip() != '':
                                    matching_standards.add(str(std).strip())
                
                # Try string contains match for more flexible matching
                str_match = df_mechem[df_mechem[prop_col].astype(str).str.contains(str(property_class), na=False, case=False)]
                if not str_match.empty:
                    for std_col in standard_cols:
                        if std_col in str_match.columns:
                            standards = str_match[std_col].dropna().unique()
                            for std in standards:
                                if pd.notna(std) and str(std).strip() != '':
                                    matching_standards.add(str(std).strip())
        
        # If still no standards found, return some default/common standards
        if not matching_standards:
            common_standards = ['ASTM A193', 'ASTM A320', 'ASTM A194', 'ISO 898-1', 'ISO 3506', 'ASME B18.2.1']
            for std in common_standards:
                matching_standards.add(std)
        
        return sorted(list(matching_standards))
        
    except Exception as e:
        st.error(f"Error getting standards for {property_class}: {str(e)}")
        return []

def show_mechanical_chemical_details(property_class):
    """Show detailed mechanical and chemical properties for a selected property class"""
    if df_mechem.empty or not property_class:
        return
    
    try:
        # Find ALL possible property class columns
        property_class_cols = []
        possible_class_cols = ['Grade', 'Class', 'Property Class', 'Material Grade', 'Type', 'Designation', 'Material']
        
        for col in df_mechem.columns:
            col_lower = str(col).lower()
            for possible in possible_class_cols:
                if possible.lower() in col_lower:
                    property_class_cols.append(col)
                    break
        
        if not property_class_cols:
            st.info("No property class column found in the data")
            return
        
        # Try to find matching data using ALL property class columns
        filtered_data = pd.DataFrame()
        
        for prop_col in property_class_cols:
            if prop_col in df_mechem.columns:
                # Try exact match
                exact_match = df_mechem[df_mechem[prop_col] == property_class]
                if not exact_match.empty:
                    filtered_data = exact_match
                    break
                # Try string contains
                str_match = df_mechem[df_mechem[prop_col].astype(str).str.contains(str(property_class), na=False, case=False)]
                if not str_match.empty:
                    filtered_data = str_match
                    break
        
        if filtered_data.empty:
            st.info(f"No detailed data found for {property_class}")
            return
        
        st.markdown(f"### Detailed Properties for {property_class}")
        
        # Display the filtered data
        st.dataframe(
            filtered_data,
            use_container_width=True,
            height=400
        )
        
        # Show key properties in a structured way
        st.markdown("#### Key Properties")
        
        mechanical_props = []
        chemical_props = []
        other_props = []
        
        for col in filtered_data.columns:
            col_lower = str(col).lower()
            if col in property_class_cols:
                continue
            
            if any(keyword in col_lower for keyword in ['tensile', 'yield', 'hardness', 'strength', 'elongation', 'proof']):
                mechanical_props.append(col)
            elif any(keyword in col_lower for keyword in ['carbon', 'manganese', 'phosphorus', 'sulfur', 'chromium', 'nickel', 'chemical']):
                chemical_props.append(col)
            else:
                other_props.append(col)
        
        if mechanical_props:
            st.markdown("**Mechanical Properties:**")
            mech_cols = st.columns(min(3, len(mechanical_props)))
            for idx, prop in enumerate(mechanical_props):
                with mech_cols[idx % len(mech_cols)]:
                    value = filtered_data[prop].iloc[0] if not filtered_data[prop].isna().all() else "N/A"
                    st.metric(prop, value)
        
        if chemical_props:
            st.markdown("**Chemical Composition (%):**")
            chem_cols = st.columns(min(4, len(chemical_props)))
            for idx, prop in enumerate(chemical_props):
                with chem_cols[idx % len(chem_cols)]:
                    value = filtered_data[prop].iloc[0] if not filtered_data[prop].isna().all() else "N/A"
                    st.metric(prop, value)
                    
    except Exception as e:
        st.error(f"Error displaying mechanical/chemical details: {str(e)}")

# Initialize Mechanical & Chemical data processing
with LoadingManager.show_loading_spinner("Processing mechanical & chemical properties..."):
    me_chem_columns, property_classes = process_mechanical_chemical_data()

# ======================================================
# COMPLETELY BULLETPROOF SIZE HANDLING - FIXED VERSION
# ======================================================
def size_to_float(size_str):
    """Convert size string to float for sorting - ULTRA ROBUST VERSION"""
    try:
        if pd.isna(size_str) or not isinstance(size_str, (str, int, float)):
            return 0.0
        
        size_str = str(size_str).strip()
        if not size_str or size_str == "":
            return 0.0
        
        # Handle numeric sizes (0,1,2,3 etc)
        if size_str.isdigit():
            return float(size_str)
        
        if size_str.startswith('M'):
            match = re.match(r'M\s*([\d.]+)', size_str)
            if match:
                return float(match.group(1))
            return 0.0
        
        if "/" in size_str:
            try:
                if "-" in size_str:
                    # Handle cases like "1-1/2"
                    parts = size_str.split("-")
                    whole = float(parts[0]) if parts[0] else 0
                    fraction = float(Fraction(parts[1]))
                    return whole + fraction
                else:
                    return float(Fraction(size_str))
            except:
                return 0.0
        
        try:
            return float(size_str)
        except:
            return 0.0
        
    except Exception as e:
        return 0.0

def safe_sort_sizes(size_list):
    """Safely sort size list with multiple fallbacks"""
    if not size_list or len(size_list) == 0:
        return []
    
    try:
        return sorted(size_list, key=lambda x: (size_to_float(x), str(x)))
    except:
        try:
            return sorted(size_list, key=str)
        except:
            return list(size_list)

def get_safe_size_options(temp_df):
    """Completely safe way to get size options - FIXED VERSION"""
    size_options = ["All"]
    
    if temp_df is None or temp_df.empty:
        return size_options
    
    if 'Size' not in temp_df.columns:
        return size_options
    
    try:
        # Get unique sizes and handle NaN values properly
        unique_sizes = temp_df['Size'].dropna().unique()
        
        # Convert all sizes to string and filter out empty strings
        unique_sizes = [str(size) for size in unique_sizes if str(size).strip() != '']
        
        if len(unique_sizes) > 0:
            sorted_sizes = safe_sort_sizes(unique_sizes)
            size_options.extend(sorted_sizes)
        else:
            return ["All"]
    except Exception as e:
        st.warning(f"Size processing warning: {str(e)}")
        try:
            unique_sizes = temp_df['Size'].dropna().unique()
            unique_sizes = [str(size) for size in unique_sizes if str(size).strip() != '']
            size_options.extend(list(unique_sizes))
        except:
            pass
    
    return size_options

# ======================================================
# NEW: GRADE HANDLING FUNCTIONS FOR ISO 4014
# ======================================================
def get_available_grades_for_standard_product(standard, product):
    """Get available grades for specific standard and product"""
    grade_options = ["All"]
    
    if standard == "Select Standard" or product == "Select Product":
        return grade_options
    
    # Only ISO 4014 has product grades A and B
    if standard == "ISO 4014" and product == "Hex Bolt":
        # Get the appropriate dataframe
        temp_df = df_iso4014.copy()
        
        # Filter by product if specified
        if product != "All" and 'Product' in temp_df.columns:
            temp_df = temp_df[temp_df['Product'] == product]
        
        # Get grade options from the data
        if 'Product Grade' in temp_df.columns:
            unique_grades = temp_df['Product Grade'].dropna().unique()
            unique_grades = [str(grade).strip() for grade in unique_grades if str(grade).strip() != '']
            if len(unique_grades) > 0:
                grade_options.extend(sorted(unique_grades))
        else:
            # Default grades for ISO 4014 Hex Bolt
            grade_options.extend(["A", "B"])
    
    return grade_options

def get_sizes_for_standard_product_grade(standard, product, grade):
    """Get available sizes for specific standard, product and grade"""
    size_options = ["Select Size"]
    
    if standard == "Select Standard" or product == "Select Product":
        return size_options
    
    # Get the appropriate dataframe based on standard
    if standard == "ASME B18.2.1":
        temp_df = df.copy()
    elif standard == "ISO 4014":
        temp_df = df_iso4014.copy()
    elif standard == "DIN-7991":
        temp_df = df_din7991.copy()
    elif standard == "ASME B18.3":
        temp_df = df_asme_b18_3.copy()
    else:
        return size_options
    
    # Filter by product if specified
    if product != "All" and 'Product' in temp_df.columns:
        temp_df = temp_df[temp_df['Product'] == product]
    
    # Filter by grade if specified (only for ISO 4014)
    if standard == "ISO 4014" and grade != "All" and 'Product Grade' in temp_df.columns:
        temp_df = temp_df[temp_df['Product Grade'] == grade]
    
    # Get size options
    size_options = get_safe_size_options(temp_df)
    
    return ["Select Size"] + [size for size in size_options if size != "All"]

# ======================================================
# RECTIFIED UNIT CONVERSION FUNCTIONS - FIXED VERSION
# ======================================================
def convert_to_mm(value, from_unit):
    """Convert any unit to millimeters - FIXED: No unnecessary conversion"""
    try:
        if pd.isna(value):
            return 0.0
        
        value = float(value)
        
        # Only convert if the unit is NOT mm
        if from_unit == 'mm':
            return value  # Already in mm, no conversion needed
        elif from_unit == 'inch':
            return value * 25.4  # Convert inches to mm
        elif from_unit == 'ft':
            return value * 304.8  # Convert feet to mm
        elif from_unit == 'meter':
            return value * 1000  # Convert meters to mm
        else:
            return value  # Assume mm if unknown unit
    except Exception as e:
        st.warning(f"Unit conversion error: {str(e)}")
        return value

# ======================================================
# FIXED: SEPARATE DATA FETCHING FOR SOCKET HEAD PRODUCTS
# ======================================================

def get_asme_b18_3_dimensions(product, size):
    """FIXED VERSION: Get head diameter and head height for ASME B18.3 socket head cap screws"""
    try:
        temp_df = df_asme_b18_3.copy()
        original_unit = "inch"  # ASME B18.3 data is in inches
        
        # Filter by product and size
        if 'Product' in temp_df.columns and product != "All":
            temp_df = temp_df[temp_df['Product'].str.contains('Socket Head', na=False, case=False)]
        
        if 'Size' in temp_df.columns and size != "All":
            # Normalize size comparison - handle different formats
            temp_df['Size_Normalized'] = temp_df['Size'].ast(str).str.strip()
            size_normalized = str(size).strip()
            temp_df = temp_df[temp_df['Size_Normalized'] == size_normalized]
        
        if temp_df.empty:
            st.warning(f"No ASME B18.3 data found for {product} size {size}")
            return None, None, original_unit
        
        # SPECIFIC ASME B18.3 COLUMN MAPPING FOR HEAD DIAMETER (MIN) AND HEAD HEIGHT (MIN)
        head_dia_col = None
        head_height_col = None
        
        # Debug: Show available columns
        if st.session_state.debug_mode:
            st.sidebar.write(f"ASME B18.3 Debug - Size: {size}")
            st.sidebar.write(f"All columns: {temp_df.columns.tolist()}")
        
        # SPECIFIC COLUMN NAMES FOR ASME B18.3 - EXACT MATCHES
        # Head Diameter (Min) - Look for exact column names
        head_dia_columns = [
            'Head Diameter (Min)', 'Head_Diameter_Min', 'Head Dia Min', 
            'Head Diameter Min', 'Head_Dia_Min', 'dk_min', 'Head_D_Min'
        ]
        
        # Head Height (Min) - Look for exact column names
        head_height_columns = [
            'Head Height (Min)', 'Head_Height_Min', 'Head Height Min',
            'Head_Ht_Min', 'k_min', 'Head_H_Min'
        ]
        
        # Find Head Diameter (Min) column
        for col in temp_df.columns:
            col_clean = str(col).strip()
            for target_col in head_dia_columns:
                if target_col.lower() == col_clean.lower():
                    head_dia_col = col
                    break
            if head_dia_col:
                break
        
        # If not found with exact match, try partial match
        if not head_dia_col:
            for col in temp_df.columns:
                col_lower = str(col).lower()
                if 'head' in col_lower and 'diameter' in col_lower and 'min' in col_lower:
                    head_dia_col = col
                    break
        
        # Find Head Height (Min) column
        for col in temp_df.columns:
            col_clean = str(col).strip()
            for target_col in head_height_columns:
                if target_col.lower() == col_clean.lower():
                    head_height_col = col
                    break
            if head_height_col:
                break
        
        # If not found with exact match, try partial match
        if not head_height_col:
            for col in temp_df.columns:
                col_lower = str(col).lower()
                if 'head' in col_lower and 'height' in col_lower and 'min' in col_lower:
                    head_height_col = col
                    break
        
        # Debug: Show found columns
        if st.session_state.debug_mode:
            st.sidebar.write(f"Head Diameter Column: {head_dia_col}")
            st.sidebar.write(f"Head Height Column: {head_height_col}")
        
        head_diameter = None
        head_height = None
        
        # Get Head Diameter (Min) value
        if head_dia_col and head_dia_col in temp_df.columns:
            head_diameter_val = temp_df[head_dia_col].iloc[0]
            if pd.notna(head_diameter_val):
                try:
                    head_diameter = float(head_diameter_val)
                    if st.session_state.debug_mode:
                        st.sidebar.write(f"Head Diameter from {head_dia_col}: {head_diameter}")
                except (ValueError, TypeError) as e:
                    st.warning(f"Could not convert head diameter value: {head_diameter_val}")
        
        # Get Head Height (Min) value
        if head_height_col and head_height_col in temp_df.columns:
            head_height_val = temp_df[head_height_col].iloc[0]
            if pd.notna(head_height_val):
                try:
                    head_height = float(head_height_val)
                    if st.session_state.debug_mode:
                        st.sidebar.write(f"Head Height from {head_height_col}: {head_height}")
                except (ValueError, TypeError) as e:
                    st.warning(f"Could not convert head height value: {head_height_val}")
        
        # If still no values found, try alternative approaches
        if head_diameter is None:
            # Try to find any head diameter column
            for col in temp_df.columns:
                col_lower = str(col).lower()
                if 'head' in col_lower and 'diameter' in col_lower:
                    if 'thread' not in col_lower and 'body' not in col_lower:
                        try:
                            head_diameter = float(temp_df[col].iloc[0])
                            head_dia_col = col
                            break
                        except:
                            continue
        
        if head_height is None:
            # Try to find any head height column
            for col in temp_df.columns:
                col_lower = str(col).lower()
                if 'head' in col_lower and 'height' in col_lower:
                    try:
                        head_height = float(temp_df[col].iloc[0])
                        head_height_col = col
                        break
                    except:
                        continue
        
        # Final debug information
        if st.session_state.debug_mode:
            st.sidebar.write(f"ASME B18.3 Final Head Diameter: {head_diameter}")
            st.sidebar.write(f"ASME B18.3 Final Head Height: {head_height}")
            st.sidebar.write(f"Head Diameter Column Used: {head_dia_col}")
            st.sidebar.write(f"Head Height Column Used: {head_height_col}")
        
        LoadingManager.log_operation(f"Get ASME B18.3 Dimensions", True, f"Head Dia: {head_diameter}, Head Height: {head_height}")
        return head_diameter, head_height, original_unit
            
    except Exception as e:
        st.error(f"Error getting ASME B18.3 dimensions: {str(e)}")
        LoadingManager.log_operation("Get ASME B18.3 Dimensions", False, str(e))
        return None, None, "inch"

def get_din7991_dimensions(product, size):
    """SEPARATE FUNCTION: Get head diameter and head height for DIN-7991 socket countersunk head cap screws"""
    try:
        temp_df = df_din7991.copy()
        original_unit = "mm"  # DIN-7991 data is in mm
        
        # Filter by product and size
        if 'Product' in temp_df.columns and product != "All":
            temp_df = temp_df[temp_df['Product'] == product]
        
        if 'Size' in temp_df.columns and size != "All":
            # Normalize size comparison
            temp_df = temp_df[temp_df['Size'].astype(str).str.strip() == str(size).strip()]
        
        if temp_df.empty:
            return None, None, original_unit
        
        # SPECIFIC COLUMN MAPPING FOR DIN-7991
        # Look for Head Diameter (dk) - specifically dk column for DIN-7991
        head_dia_cols = [col for col in temp_df.columns if any(keyword in col.lower() for keyword in ['dk', 'head diameter', 'head_dia'])]
        head_dia_col = None
        
        # Prioritize 'dk' column for DIN-7991
        for col in head_dia_cols:
            if 'dk' in col.lower():
                head_dia_col = col
                break
        
        # If no 'dk' found, look for other head diameter columns
        if not head_dia_col and head_dia_cols:
            for col in head_dia_cols:
                if 'min' in col.lower():
                    head_dia_col = col
                    break
            if not head_dia_col:
                head_dia_col = head_dia_cols[0]
        
        # Look for Head Height (k) - specifically k column for DIN-7991
        head_height_cols = [col for col in temp_df.columns if any(keyword in col.lower() for keyword in ['k', 'head height', 'head_height'])]
        head_height_col = None
        
        # Prioritize 'k' column for DIN-7991
        for col in head_height_cols:
            if col.lower() == 'k' or 'head height' in col.lower():
                head_height_col = col
                break
        
        # If no 'k' found, look for other head height columns
        if not head_height_col and head_height_cols:
            for col in head_height_cols:
                if 'max' in col.lower():
                    head_height_col = col
                    break
            if not head_height_col:
                head_height_col = head_height_cols[0]
        
        head_diameter = None
        head_height = None
        
        if head_dia_col and head_dia_col in temp_df.columns:
            head_diameter = temp_df[head_dia_col].iloc[0]
            if pd.notna(head_diameter):
                head_diameter = float(head_diameter)
        
        if head_height_col and head_height_col in temp_df.columns:
            head_height = temp_df[head_height_col].iloc[0]
            if pd.notna(head_height):
                head_height = float(head_height)
        
        # Debug information
        if st.session_state.debug_mode:
            st.sidebar.write(f"DIN-7991 Debug - Size: {size}")
            st.sidebar.write(f"Head Diameter Column: {head_dia_col}, Value: {head_diameter}")
            st.sidebar.write(f"Head Height Column: {head_height_col}, Value: {head_height}")
            st.sidebar.write(f"Available columns: {temp_df.columns.tolist()}")
        
        LoadingManager.log_operation(f"Get DIN-7991 Dimensions", True, f"Head Dia: {head_diameter}, Head Height: {head_height}")
        return head_diameter, head_height, original_unit
        
    except Exception as e:
        st.warning(f"Error getting DIN-7991 dimensions: {str(e)}")
        LoadingManager.log_operation("Get DIN-7991 Dimensions", False, str(e))
        return None, None, "mm"

def get_socket_head_dimensions(standard, product, size, grade="All"):
    """MAIN FUNCTION: Route to appropriate socket head dimension function based on standard"""
    try:
        if standard == "ASME B18.3":
            return get_asme_b18_3_dimensions(product, size)
        elif standard == "DIN-7991":
            return get_din7991_dimensions(product, size)
        else:
            return None, None, "unknown"
    except Exception as e:
        st.warning(f"Error in get_socket_head_dimensions for {standard}: {str(e)}")
        return None, None, "unknown"

def get_hex_head_dimensions(standard, product, size, grade="All"):
    """RECTIFIED: Get width across flats and head height for hex products from database with proper unit tracking"""
    try:
        # Get the appropriate dataframe based on standard
        if standard == "ASME B18.2.1":
            temp_df = df.copy()
            original_unit = "inch"  # ASME B18.2.1 data is in inches
        elif standard == "ISO 4014":
            temp_df = df_iso4014.copy()
            original_unit = "mm"  # ISO 4014 data is in mm
        elif standard == "DIN-7991":
            temp_df = df_din7991.copy()
            original_unit = "mm"  # DIN-7991 data is in mm
        elif standard == "ASME B18.3":
            temp_df = df_asme_b18_3.copy()
            original_unit = "inch"  # ASME B18.3 data is in inches
        else:
            return None, None, "unknown"
        
        # Filter by product and size
        if 'Product' in temp_df.columns and product != "All":
            temp_df = temp_df[temp_df['Product'] == product]
        
        if 'Size' in temp_df.columns and size != "All":
            # Normalize size comparison
            temp_df = temp_df[temp_df['Size'].astype(str).str.strip() == str(size).strip()]
        
        # Filter by grade if specified (only for ISO 4014)
        if standard == "ISO 4014" and grade != "All" and 'Product Grade' in temp_df.columns:
            temp_df = temp_df[temp_df['Product Grade'] == grade]
        
        if temp_df.empty:
            return None, None, original_unit
        
        # Look for width across flats column
        width_cols = [col for col in temp_df.columns if any(keyword in col.lower() for keyword in ['width', 'across', 'flats', 'w_'])]
        width_col = None
        for col in width_cols:
            if 'min' in col.lower():
                width_col = col
                break
        if not width_col and width_cols:
            width_col = width_cols[0]
        
        # Look for head height column
        height_cols = [col for col in temp_df.columns if any(keyword in col.lower() for keyword in ['head', 'height', 'head_height'])]
        height_col = None
        for col in height_cols:
            if 'min' in col.lower():
                height_col = col
                break
        if not height_col and height_cols:
            height_col = height_cols[0]
        
        width_across_flats = None
        head_height = None
        
        if width_col and width_col in temp_df.columns:
            width_across_flats = temp_df[width_col].iloc[0]
            if pd.notna(width_across_flats):
                width_across_flats = float(width_across_flats)
        
        if height_col and height_col in temp_df.columns:
            head_height = temp_df[height_col].iloc[0]
            if pd.notna(head_height):
                head_height = float(head_height)
        
        LoadingManager.log_operation(f"Get Hex Head Dimensions", True, f"Width: {width_across_flats}, Height: {head_height}")
        return width_across_flats, head_height, original_unit
        
    except Exception as e:
        st.warning(f"Error getting hex head dimensions: {str(e)}")
        LoadingManager.log_operation("Get Hex Head Dimensions", False, str(e))
        return None, None, "unknown"

# ======================================================
# FIXED: VOLUME CALCULATION FUNCTIONS FOR SOCKET HEAD PRODUCTS
# ======================================================

def calculate_socket_head_volume_rectified(head_diameter_mm, head_height_mm):
    """Calculate volume for socket head using cylinder formula - FIXED: No head angle"""
    try:
        # For socket head, we use cylinder volume formula: V = 0.7853 × d² × h
        head_volume_mm3 = 0.7853 * (head_diameter_mm ** 2) * head_height_mm
        return head_volume_mm3
    except Exception as e:
        st.warning(f"Error calculating socket head volume: {str(e)}")
        return 0.0

def calculate_shank_volume_rectified(diameter_mm, length_mm):
    """Calculate shank volume using cylinder formula"""
    try:
        # Shank volume formula: V = 0.7853 × d² × L
        shank_volume_mm3 = 0.7853 * (diameter_mm ** 2) * length_mm
        return shank_volume_mm3
    except Exception as e:
        st.warning(f"Error calculating shank volume: {str(e)}")
        return 0.0

def calculate_socket_product_weight_rectified(parameters, head_diameter, head_height, original_unit):
    """FIXED: Calculate weight for Socket Head Products (ASME B18.3 and DIN-7991)"""
    try:
        # Extract parameters
        product_type = parameters.get('product_type', 'Hexagon Socket Head Cap Screws')
        diameter_type = parameters.get('diameter_type', 'Blank Diameter')
        diameter_value = parameters.get('diameter_value', 0.0)
        diameter_unit = parameters.get('diameter_unit', 'mm')
        length = parameters.get('length', 0.0)
        length_unit = parameters.get('length_unit', 'mm')
        material = parameters.get('material', 'Carbon Steel')
        
        # Convert dimensions to mm only if needed
        diameter_mm = convert_to_mm(diameter_value, diameter_unit)
        length_mm = convert_to_mm(length, length_unit)
        
        # Convert head dimensions to mm (they should already be in mm from database)
        if head_diameter is not None:
            head_diameter_mm = convert_to_mm(head_diameter, original_unit)
        else:
            head_diameter_mm = diameter_mm * 1.5  # Default ratio if not available
        
        if head_height is not None:
            head_height_mm = convert_to_mm(head_height, original_unit)
        else:
            head_height_mm = diameter_mm * 0.65  # Default ratio if not available
        
        # Get material density in g/cm³
        density_g_cm3 = get_material_density_rectified(material)
        
        # Calculate volumes in mm³
        shank_volume_mm3 = calculate_shank_volume_rectified(diameter_mm, length_mm)
        head_volume_mm3 = calculate_socket_head_volume_rectified(head_diameter_mm, head_height_mm)
        total_volume_mm3 = shank_volume_mm3 + head_volume_mm3
        
        # Convert mm³ to cm³ for weight calculation
        total_volume_cm3 = total_volume_mm3 / 1000
        
        # Calculate Weight in grams and kg
        weight_g = total_volume_cm3 * density_g_cm3
        weight_kg = weight_g / 1000
        weight_lb = weight_kg * 2.20462
        
        result = {
            'weight_kg': weight_kg,
            'weight_g': weight_g,
            'weight_lb': weight_lb,
            'shank_volume_mm3': shank_volume_mm3,
            'head_volume_mm3': head_volume_mm3,
            'total_volume_mm3': total_volume_mm3,
            'total_volume_cm3': total_volume_cm3,
            'diameter_mm': diameter_mm,
            'length_mm': length_mm,
            'head_diameter_mm': head_diameter_mm,
            'head_height_mm': head_height_mm,
            'density_g_cm3': density_g_cm3,
            'original_diameter': f"{diameter_value} {diameter_unit}",
            'original_length': f"{length} {length_unit}",
            'original_head_diameter': f"{head_diameter} {original_unit}" if head_diameter else "N/A",
            'original_head_height': f"{head_height} {original_unit}" if head_height else "N/A",
            'calculation_method': 'Socket Head Formula',
            'formula_details': {
                'shank_volume_formula': '0.7853 × (diameter)² × length (mm³)',
                'head_volume_formula': '0.7853 × (head_diameter_min)² × head_height_min (mm³)',
                'total_volume_formula': 'shank_volume + head_volume (mm³)',
                'volume_conversion': 'mm³ to cm³: divide by 1000',
                'weight_formula': 'total_volume_cm³ × density_g/cm³'
            },
            'dimensions_used': {
                'diameter_input': f"{diameter_value:.4f} {diameter_unit}",
                'diameter_calculation_mm': f"{diameter_mm:.4f}",
                'length_input': f"{length:.4f} {length_unit}",
                'length_calculation_mm': f"{length_mm:.4f}",
                'head_diameter_input': f"{head_diameter:.4f} {original_unit}" if head_diameter else "Estimated",
                'head_diameter_calculation_mm': f"{head_diameter_mm:.4f}",
                'head_height_input': f"{head_height:.4f} {original_unit}" if head_height else "Estimated",
                'head_height_calculation_mm': f"{head_height_mm:.4f}",
                'shank_volume_mm3': f"{shank_volume_mm3:.4f}",
                'head_volume_mm3': f"{head_volume_mm3:.4f}",
                'total_volume_mm3': f"{total_volume_mm3:.4f}",
                'total_volume_cm3': f"{total_volume_cm3:.4f}",
                'density_g_cm3': f"{density_g_cm3:.4f}"
            }
        }
        
        LoadingManager.log_operation("Socket Product Weight Calculation", True, f"Weight: {weight_kg:.4f} kg")
        return result
        
    except Exception as e:
        st.error(f"Socket product calculation error: {str(e)}")
        LoadingManager.log_operation("Socket Product Weight Calculation", False, str(e))
        return None

def calculate_hex_product_weight_rectified(parameters, width_across_flats, head_height, original_unit):
    """FIXED: Calculate weight for hex products with detailed parameters"""
    try:
        # Extract parameters
        product_type = parameters.get('product_type', 'Hex Bolt')
        diameter_type = parameters.get('diameter_type', 'Blank Diameter')
        diameter_value = parameters.get('diameter_value', 0.0)
        diameter_unit = parameters.get('diameter_unit', 'mm')
        length = parameters.get('length', 0.0)
        length_unit = parameters.get('length_unit', 'mm')
        material = parameters.get('material', 'Carbon Steel')
        
        # Convert dimensions to mm only if needed
        diameter_mm = convert_to_mm(diameter_value, diameter_unit)
        length_mm = convert_to_mm(length, length_unit)
        
        # Convert head dimensions to mm (they should already be in mm from database)
        if width_across_flats is not None:
            width_across_flats_mm = convert_to_mm(width_across_flats, original_unit)
        else:
            width_across_flats_mm = diameter_mm * 1.5  # Default ratio if not available
        
        if head_height is not None:
            head_height_mm = convert_to_mm(head_height, original_unit)
        else:
            head_height_mm = diameter_mm * 0.65  # Default ratio if not available
        
        # Get material density in g/cm³
        density_g_cm3 = get_material_density_rectified(material)
        
        # Calculate volumes in mm³
        shank_volume_mm3 = calculate_shank_volume_rectified(diameter_mm, length_mm)
        
        # Calculate Head Volume using the specific formula in mm³
        side_length_mm = width_across_flats_mm * 1.1547
        head_volume_mm3 = 0.65 * (side_length_mm**2) * head_height_mm
        
        # Total Volume in mm³
        total_volume_mm3 = shank_volume_mm3 + head_volume_mm3
        
        # Convert mm³ to cm³ for weight calculation
        total_volume_cm3 = total_volume_mm3 / 1000
        
        # Calculate Weight in grams and kg
        weight_g = total_volume_cm3 * density_g_cm3
        weight_kg = weight_g / 1000
        weight_lb = weight_kg * 2.20462
        
        result = {
            'weight_kg': weight_kg,
            'weight_g': weight_g,
            'weight_lb': weight_lb,
            'shank_volume_mm3': shank_volume_mm3,
            'head_volume_mm3': head_volume_mm3,
            'total_volume_mm3': total_volume_mm3,
            'total_volume_cm3': total_volume_cm3,
            'diameter_mm': diameter_mm,
            'length_mm': length_mm,
            'width_across_flats_mm': width_across_flats_mm,
            'head_height_mm': head_height_mm,
            'side_length_mm': side_length_mm,
            'density_g_cm3': density_g_cm3,
            'original_diameter': f"{diameter_value} {diameter_unit}",
            'original_length': f"{length} {length_unit}",
            'original_width_across_flats': f"{width_across_flats} {original_unit}" if width_across_flats else "N/A",
            'original_head_height': f"{head_height} {original_unit}" if head_height else "N/A",
            'calculation_method': 'Hex Product Formula',
            'formula_details': {
                'shank_volume_formula': '0.7853 × (diameter)² × length (mm³)',
                'head_volume_formula': '0.65 × side_length² × head_height (mm³)',
                'side_length_formula': 'width_across_flats × 1.1547 (mm)',
                'total_volume_formula': 'shank_volume + head_volume (mm³)',
                'volume_conversion': 'mm³ to cm³: divide by 1000',
                'weight_formula': 'total_volume_cm³ × density_g/cm³'
            },
            'dimensions_used': {
                'diameter_input': f"{diameter_value:.4f} {diameter_unit}",
                'diameter_calculation_mm': f"{diameter_mm:.4f}",
                'length_input': f"{length:.4f} {length_unit}",
                'length_calculation_mm': f"{length_mm:.4f}",
                'width_across_flats_input': f"{width_across_flats:.4f} {original_unit}" if width_across_flats else "Estimated",
                'width_across_flats_calculation_mm': f"{width_across_flats_mm:.4f}",
                'head_height_input': f"{head_height:.4f} {original_unit}" if head_height else "Estimated",
                'head_height_calculation_mm': f"{head_height_mm:.4f}",
                'side_length_calculation_mm': f"{side_length_mm:.4f}",
                'shank_volume_mm3': f"{shank_volume_mm3:.4f}",
                'head_volume_mm3': f"{head_volume_mm3:.4f}",
                'total_volume_mm3': f"{total_volume_mm3:.4f}",
                'total_volume_cm3': f"{total_volume_cm3:.4f}",
                'density_g_cm3': f"{density_g_cm3:.4f}"
            }
        }
        
        LoadingManager.log_operation("Hex Product Weight Calculation", True, f"Weight: {weight_kg:.4f} kg")
        return result
        
    except Exception as e:
        st.error(f"Hex product calculation error: {str(e)}")
        LoadingManager.log_operation("Hex Product Weight Calculation", False, str(e))
        return None

# ======================================================
# WEIGHT CALCULATION SECTION - COMPLETE WORKFLOW IMPLEMENTATION
# ======================================================

def get_available_products():
    """Get all available products from standards database"""
    all_products = set()
    for standard_products_list in st.session_state.available_products.values():
        all_products.update(standard_products_list)
    return ["Select Product"] + sorted([p for p in all_products if p != "All"])

def get_series_for_product(product):
    """Get available series for a specific product"""
    if product == "Select Product":
        return ["Select Series"]
    
    available_series = set()
    for standard, products in st.session_state.available_products.items():
        if product in products:
            series = st.session_state.available_series.get(standard, "")
            if series:
                available_series.add(series)
    
    return ["Select Series"] + sorted(list(available_series))

def get_standards_for_product_series(product, series):
    """Get available standards for specific product and series"""
    if product == "Select Product" or series == "Select Series":
        return ["Select Standard"]
    
    available_standards = []
    for standard, products in st.session_state.available_products.items():
        if product in products:
            std_series = st.session_state.available_series.get(standard, "")
            if std_series == series:
                available_standards.append(standard)
    
    return ["Select Standard"] + sorted(available_standards)

def get_sizes_for_standard_product(standard, product):
    """Get available sizes for specific standard and product in weight calculator"""
    if standard == "Select Standard" or product == "Select Product":
        return ["Select Size"]
    
    # Get the appropriate dataframe based on standard
    if standard == "ASME B18.2.1":
        temp_df = df.copy()
    elif standard == "ISO 4014":
        temp_df = df_iso4014.copy()
    elif standard == "DIN-7991":
        temp_df = df_din7991.copy()
    elif standard == "ASME B18.3":
        temp_df = df_asme_b18_3.copy()
    else:
        return ["Select Size"]
    
    # Filter by product if specified
    if product != "All" and 'Product' in temp_df.columns:
        temp_df = temp_df[temp_df['Product'] == product]
    
    # Get size options
    size_options = get_safe_size_options(temp_df)
    
    return ["Select Size"] + [size for size in size_options if size != "All"]

def get_thread_standards_for_series(series):
    """Get thread standards based on series"""
    if series == "Inch":
        return ["ASME B1.1"]
    elif series == "Metric":
        return ["ISO 965-2-98 Coarse", "ISO 965-2-98 Fine"]
    return ["Select Thread Standard"]

def get_material_density_rectified(material):
    """RECTIFIED: Get density for different materials in g/cm³"""
    density_map = {
        "Carbon Steel": 7.85,
        "Stainless Steel": 8.00,
        "Alloy Steel": 7.85,
        "Brass": 8.50,
        "Aluminum": 2.70,
        "Copper": 8.96,
        "Titanium": 4.50,
        "Bronze": 8.80,
        "Inconel": 8.20,
        "Monel": 8.80,
        "Nickel": 8.90
    }
    return density_map.get(material, 7.85)  # Default to carbon steel

def get_pitch_diameter_from_thread_data(thread_standard, thread_size, thread_class):
    """Get pitch diameter from thread data for threaded rod calculation - ENHANCED FOR THREADED ROD"""
    try:
        df_thread = get_thread_data_enhanced(thread_standard, thread_size, thread_class)
        
        if df_thread.empty:
            return None
        
        # Look for pitch diameter columns - prioritize minimum pitch diameter for threaded rod
        pitch_dia_cols = []
        
        # First priority: Pitch diameter minimum columns
        min_pitch_cols = [col for col in df_thread.columns if 'pitch' in col.lower() and 'diameter' in col.lower() and 'min' in col.lower()]
        if min_pitch_cols:
            pitch_dia_cols.extend(min_pitch_cols)
        
        # Second priority: Any pitch diameter columns
        general_pitch_cols = [col for col in df_thread.columns if 'pitch' in col.lower() and 'diameter' in col.lower()]
        if general_pitch_cols:
            pitch_dia_cols.extend([col for col in general_pitch_cols if col not in pitch_dia_cols])
        
        # Third priority: Any diameter column that might contain pitch diameter
        if not pitch_dia_cols:
            dia_cols = [col for col in df_thread.columns if 'diameter' in col.lower()]
            for col in dia_cols:
                if 'pitch' not in col.lower() and 'major' not in col.lower() and 'minor' not in col.lower():
                    pitch_dia_cols.append(col)
        
        if pitch_dia_cols:
            # Get the first pitch diameter value
            pitch_diameter = df_thread[pitch_dia_cols[0]].iloc[0]
            if pd.notna(pitch_diameter):
                return float(pitch_diameter)
        
        return None
        
    except Exception as e:
        st.warning(f"Could not retrieve pitch diameter: {str(e)}")
        return None

def calculate_weight_rectified(parameters):
    """FIXED: Enhanced weight calculation with proper data fetching for ALL products"""
    try:
        # Extract parameters
        product_type = parameters.get('product_type', 'Hex Bolt')
        diameter_type = parameters.get('diameter_type', 'Blank Diameter')
        diameter_value = parameters.get('diameter_value', 0.0)
        diameter_unit = parameters.get('diameter_unit', 'mm')
        length = parameters.get('length', 0.0)
        length_unit = parameters.get('length_unit', 'mm')
        material = parameters.get('material', 'Carbon Steel')
        standard = parameters.get('standard', 'ASME B18.2.1')
        size = parameters.get('size', 'All')
        grade = parameters.get('grade', 'All')
        
        # SPECIAL CASE: For Socket Head Products (ASME B18.3 and DIN-7991)
        socket_head_products = ["Hexagon Socket Head Cap Screws", "Hexagon Socket Countersunk Head Cap Screw"]
        
        if product_type in socket_head_products:
            # Get socket head dimensions from SEPARATE functions based on standard
            head_diameter, head_height, original_unit = get_socket_head_dimensions(standard, product_type, size, grade)
            
            # Store original dimensions for display
            original_head_diameter = head_diameter
            original_head_height = head_height
            
            # If dimensions not found in database, use default ratios
            if head_diameter is None:
                # Estimate head diameter based on shank diameter
                diameter_mm_temp = convert_to_mm(diameter_value, diameter_unit)
                head_diameter = diameter_mm_temp * 1.5  # Default ratio
                original_head_diameter = head_diameter
                original_unit = "mm"  # Default to mm for estimated values
            
            if head_height is None:
                # Estimate head height based on shank diameter
                diameter_mm_temp = convert_to_mm(diameter_value, diameter_unit)
                head_height = diameter_mm_temp * 0.65  # Default ratio
                original_head_height = head_height
            
            # Calculate using socket head formula - SAME FORMULA FOR BOTH
            return calculate_socket_product_weight_rectified(parameters, head_diameter, head_height, original_unit)
        
        # For hex products, get hex head dimensions
        width_across_flats, head_height, original_unit = get_hex_head_dimensions(standard, product_type, size, grade)
        
        # Store original dimensions for display
        original_width_across_flats = width_across_flats
        original_head_height = head_height
        
        # If dimensions not found in database, use default ratios
        if width_across_flats is None:
            # Estimate width across flats based on diameter
            diameter_mm_temp = convert_to_mm(diameter_value, diameter_unit)
            width_across_flats = diameter_mm_temp * 1.5  # Default ratio
            original_width_across_flats = width_across_flats
            original_unit = "mm"  # Default to mm for estimated values
        
        if head_height is None:
            # Estimate head height based on diameter
            diameter_mm_temp = convert_to_mm(diameter_value, diameter_unit)
            head_height = diameter_mm_temp * 0.65  # Default ratio
            original_head_height = head_height
        
        hex_products = ["Hex Bolt", "Heavy Hex Bolt", "Hex Cap Screws", "Heavy Hex Screws"]
        
        # Convert dimensions to mm only if needed
        diameter_mm = convert_to_mm(diameter_value, diameter_unit)
        length_mm = convert_to_mm(length, length_unit)
        
        # SPECIAL CASE: For Pitch Diameter from thread data in Inch series
        if (diameter_type == "Pitch Diameter" and 
            'pitch_diameter_value' in st.session_state and
            st.session_state.pitch_diameter_value is not None):
            
            # Use the pitch diameter from thread data and convert from inches to mm if needed
            pitch_diameter = st.session_state.pitch_diameter_value
            # Thread data for ASME B1.1 is in inches, so convert to mm
            if diameter_unit == 'inch':
                diameter_mm = pitch_diameter * 25.4
            else:
                diameter_mm = pitch_diameter

        # Get material density in g/cm³
        density_g_cm3 = get_material_density_rectified(material)
        
        # For Threaded Rod, use simple cylinder volume calculation in mm³
        if product_type == "Threaded Rod":
            # Simple cylinder volume for threaded rod in mm³
            shank_volume_mm3 = calculate_shank_volume_rectified(diameter_mm, length_mm)
            
            # Convert mm³ to cm³ for weight calculation
            volume_cm3 = shank_volume_mm3 / 1000
            weight_g = volume_cm3 * density_g_cm3
            weight_kg = weight_g / 1000
            weight_lb = weight_kg * 2.20462
            
            result = {
                'weight_kg': weight_kg,
                'weight_g': weight_g,
                'weight_lb': weight_lb,
                'volume_mm3': shank_volume_mm3,
                'volume_cm3': volume_cm3,
                'density_g_cm3': density_g_cm3,
                'diameter_mm': diameter_mm,
                'length_mm': length_mm,
                'original_diameter': f"{diameter_value} {diameter_unit}",
                'original_length': f"{length} {length_unit}",
                'calculation_method': 'Threaded Rod Cylinder Formula',
                'dimensions_used': {
                    'diameter_input': f"{diameter_value:.4f} {diameter_unit}",
                    'diameter_calculation_mm': f"{diameter_mm:.4f}",
                    'length_input': f"{length:.4f} {length_unit}",
                    'length_calculation_mm': f"{length_mm:.4f}",
                    'volume_mm3': f"{shank_volume_mm3:.4f}",
                    'volume_cm3': f"{volume_cm3:.4f}",
                    'density_g_cm3': f"{density_g_cm3:.4f}"
                }
            }
            
            LoadingManager.log_operation("Threaded Rod Weight Calculation", True, f"Weight: {weight_kg:.4f} kg")
            return result
        
        # For hex products, use rectified hex product formula in mm³
        elif product_type in hex_products:
            return calculate_hex_product_weight_rectified(parameters, width_across_flats, head_height, original_unit)
        
        # For other products, use simple cylinder calculation in mm³
        else:
            shank_volume_mm3 = calculate_shank_volume_rectified(diameter_mm, length_mm)
            volume_cm3 = shank_volume_mm3 / 1000
            weight_g = volume_cm3 * density_g_cm3
            weight_kg = weight_g / 1000
            weight_lb = weight_kg * 2.20462
            
            result = {
                'weight_kg': weight_kg,
                'weight_g': weight_g,
                'weight_lb': weight_lb,
                'volume_mm3': shank_volume_mm3,
                'volume_cm3': volume_cm3,
                'density_g_cm3': density_g_cm3,
                'diameter_mm': diameter_mm,
                'length_mm': length_mm,
                'original_diameter': f"{diameter_value:.4f} {diameter_unit}",
                'original_length': f"{length:.4f} {length_unit}",
                'calculation_method': 'Standard Cylinder Formula',
                'dimensions_used': {
                    'diameter_input': f"{diameter_value:.4f} {diameter_unit}",
                    'diameter_calculation_mm': f"{diameter_mm:.4f}",
                    'length_input': f"{length:.4f} {length_unit}",
                    'length_calculation_mm': f"{length_mm:.4f}",
                    'volume_mm3': f"{shank_volume_mm3:.4f}",
                    'volume_cm3': f"{volume_cm3:.4f}",
                    'density_g_cm3': f"{density_g_cm3:.4f}"
                }
            }
            
            LoadingManager.log_operation("Standard Product Weight Calculation", True, f"Weight: {weight_kg:.4f} kg")
            return result
            
    except Exception as e:
        st.error(f"Calculation error: {str(e)}")
        LoadingManager.log_operation("Weight Calculation", False, str(e))
        import traceback
        st.error(f"Detailed error: {traceback.format_exc()}")
        return None

# ======================================================
# MISSING FUNCTIONS IMPLEMENTATION - COMPLETED
# ======================================================

def get_filtered_dataframe(product, standard, grade="All"):
    """Get filtered dataframe based on product and standard selection"""
    if standard == "ASME B18.2.1":
        temp_df = df.copy()
    elif standard == "ISO 4014":
        temp_df = df_iso4014.copy()
    elif standard == "DIN-7991":
        temp_df = df_din7991.copy()
    elif standard == "ASME B18.3":
        temp_df = df_asme_b18_3.copy()
    else:
        return pd.DataFrame()
    
    # Apply product filter if specified
    if product != "All" and 'Product' in temp_df.columns:
        temp_df = temp_df[temp_df['Product'] == product]
    
    # Apply grade filter if specified (only for ISO 4014)
    if standard == "ISO 4014" and grade != "All" and 'Product Grade' in temp_df.columns:
        temp_df = temp_df[temp_df['Product Grade'] == grade]
    
    return temp_df

def apply_section_a_filters():
    """Apply filters for Section A - Dimensional Specifications"""
    filters = st.session_state.section_a_filters
    
    if not filters:
        return pd.DataFrame()
    
    product = filters.get('product', 'All')
    series = filters.get('series', 'All')
    standard = filters.get('standard', 'All')
    size = filters.get('size', 'All')
    grade = filters.get('grade', 'All')
    
    # Get appropriate dataframe
    if standard == "ASME B18.2.1":
        temp_df = df.copy()
    elif standard == "ISO 4014":
        temp_df = df_iso4014.copy()
    elif standard == "DIN-7991":
        temp_df = df_din7991.copy()
    elif standard == "ASME B18.3":
        temp_df = df_asme_b18_3.copy()
    else:
        return pd.DataFrame()
    
    # Apply filters
    if product != "All" and 'Product' in temp_df.columns:
        temp_df = temp_df[temp_df['Product'] == product]
    
    if size != "All" and 'Size' in temp_df.columns:
        temp_df = temp_df[temp_df['Size'].astype(str).str.strip() == str(size).strip()]
    
    # Apply grade filter if specified (only for ISO 4014)
    if standard == "ISO 4014" and grade != "All" and 'Product Grade' in temp_df.columns:
        temp_df = temp_df[temp_df['Product Grade'] == grade]
    
    return temp_df

def apply_section_b_filters():
    """Apply filters for Section B - Thread Specifications"""
    filters = st.session_state.section_b_filters
    
    if not filters:
        return pd.DataFrame()
    
    standard = filters.get('standard', 'All')
    size = filters.get('size', 'All')
    thread_class = filters.get('class', 'All')
    
    if standard == "All":
        return pd.DataFrame()
    
    return get_thread_data_enhanced(standard, size, thread_class)

def apply_section_c_filters():
    """Apply filters for Section C - Material Properties"""
    filters = st.session_state.section_c_filters
    
    if not filters or df_mechem.empty:
        return pd.DataFrame()
    
    property_class = filters.get('property_class', 'All')
    standard = filters.get('standard', 'All')
    
    if property_class == "All":
        return df_mechem.copy()
    
    # Find property class columns
    property_class_cols = []
    possible_class_cols = ['Grade', 'Class', 'Property Class', 'Material Grade', 'Type', 'Designation', 'Material']
    
    for col in df_mechem.columns:
        col_lower = str(col).lower()
        for possible in possible_class_cols:
            if possible.lower() in col_lower:
                property_class_cols.append(col)
                break
    
    # Try to find matching data
    filtered_data = pd.DataFrame()
    
    for prop_col in property_class_cols:
        if prop_col in df_mechem.columns:
            # Try exact match
            exact_match = df_mechem[df_mechem[prop_col] == property_class]
            if not exact_match.empty:
                filtered_data = exact_match
                break
            # Try string contains
            str_match = df_mechem[df_mechem[prop_col].astype(str).str.contains(str(property_class), na=False, case=False)]
            if not str_match.empty:
                filtered_data = str_match
                break
    
    # Apply standard filter if specified
    if standard != "All" and not filtered_data.empty:
        standard_cols = []
        possible_standard_cols = ['Standard', 'Specification', 'Norm', 'Type', 'Designation']
        
        for col in filtered_data.columns:
            col_lower = str(col).lower()
            for possible in possible_standard_cols:
                if possible.lower() in col_lower:
                    standard_cols.append(col)
                    break
        
        if standard_cols:
            for std_col in standard_cols:
                std_filtered = filtered_data[filtered_data[std_col].astype(str).str.contains(str(standard), na=False, case=False)]
                if not std_filtered.empty:
                    filtered_data = std_filtered
                    break
    
    return filtered_data

def show_section_a_results():
    """Show results for Section A"""
    if not st.session_state.section_a_results.empty:
        st.markdown('<div class="section-results">', unsafe_allow_html=True)
        st.markdown("### Section A Results - Dimensional Specifications")
        
        # Show professional card if requested
        if st.session_state.show_professional_card and st.session_state.selected_product_details:
            show_professional_product_card(st.session_state.selected_product_details)
        
        # Show data
        st.dataframe(
            st.session_state.section_a_results,
            use_container_width=True,
            height=400
        )
        
        # Show export options
        col1, col2 = st.columns(2)
        with col1:
            if st.button("Export Section A Results", key="export_section_a"):
                enhanced_export_data(st.session_state.section_a_results, "Excel")
        with col2:
            if st.button("Show Professional Card", key="show_pro_card_a"):
                if not st.session_state.section_a_results.empty:
                    # Extract first row for professional card
                    first_row = st.session_state.section_a_results.iloc[0].to_dict()
                    st.session_state.selected_product_details = extract_product_details(first_row)
                    st.session_state.show_professional_card = True
                    st.rerun()
        
        st.markdown('</div>', unsafe_allow_html=True)

def show_section_b_results():
    """Show results for Section B"""
    if not st.session_state.section_b_results.empty:
        st.markdown('<div class="section-results">', unsafe_allow_html=True)
        st.markdown("### Section B Results - Thread Specifications")
        
        st.dataframe(
            st.session_state.section_b_results,
            use_container_width=True,
            height=400
        )
        
        if st.button("Export Section B Results", key="export_section_b"):
            enhanced_export_data(st.session_state.section_b_results, "Excel")
        
        st.markdown('</div>', unsafe_allow_html=True)

def show_section_c_results():
    """Show results for Section C"""
    if not st.session_state.section_c_results.empty:
        st.markdown('<div class="section-results">', unsafe_allow_html=True)
        st.markdown("### Section C Results - Material Properties")
        
        st.dataframe(
            st.session_state.section_c_results,
            use_container_width=True,
            height=400
        )
        
        # Show detailed properties for selected property class
        if st.session_state.section_c_filters.get('property_class') and st.session_state.section_c_filters.get('property_class') != "All":
            show_mechanical_chemical_details(st.session_state.section_c_filters.get('property_class'))
        
        if st.button("Export Section C Results", key="export_section_c"):
            enhanced_export_data(st.session_state.section_c_results, "Excel")
        
        st.markdown('</div>', unsafe_allow_html=True)

def combine_all_results():
    """Combine results from all sections"""
    combined = pd.DataFrame()
    
    # Add section A results
    if not st.session_state.section_a_results.empty:
        section_a = st.session_state.section_a_results.copy()
        section_a['Section'] = 'A - Dimensional'
        combined = pd.concat([combined, section_a], ignore_index=True)
    
    # Add section B results  
    if not st.session_state.section_b_results.empty:
        section_b = st.session_state.section_b_results.copy()
        section_b['Section'] = 'B - Thread'
        combined = pd.concat([combined, section_b], ignore_index=True)
    
    # Add section C results
    if not st.session_state.section_c_results.empty:
        section_c = st.session_state.section_c_results.copy()
        section_c['Section'] = 'C - Material'
        combined = pd.concat([combined, section_c], ignore_index=True)
    
    return combined

def show_combined_results():
    """Show combined results from all sections"""
    if not st.session_state.combined_results.empty:
        st.markdown('<div class="combined-results">', unsafe_allow_html=True)
        st.markdown("### Combined Results - All Sections")
        
        st.dataframe(
            st.session_state.combined_results,
            use_container_width=True,
            height=500
        )
        
        col1, col2 = st.columns(2)
        with col1:
            if st.button("Export Combined Results", key="export_combined"):
                enhanced_export_data(st.session_state.combined_results, "Excel")
        with col2:
            if st.button("Clear Combined Results", key="clear_combined"):
                st.session_state.combined_results = pd.DataFrame()
                st.rerun()
        
        st.markdown('</div>', unsafe_allow_html=True)

def show_weight_calculator_rectified():
    """FIXED weight calculator with proper data fetching for ALL products"""
    
    st.markdown("""
    <div class="oracle11g-header">
        <h1>Weight Calculator - FIXED WORKFLOW</h1>
        <p>SEPARATE data fetching for Socket Head products with SAME formula</p>
        <div>
            <span class="oracle11g-badge">FIXED</span>
            <span class="oracle11g-badge-orange">Separate Data Fetching</span>
            <span class="oracle11g-badge-green">Same Formula</span>
            <span class="oracle11g-badge-yellow">Different Standards</span>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    st.info("""
    **SEPARATE DATA FETCHING FOR SOCKET HEAD PRODUCTS:** 
    - **Hexagon Socket Head Cap Screws (ASME B18.3):** Separate function `get_asme_b18_3_dimensions()`
    - **Hexagon Socket Countersunk Head Cap Screw (DIN-7991):** Separate function `get_din7991_dimensions()`  
    - **SAME FORMULA:** Both use cylinder volume formula: 0.7853 × d² × h
    - **DIFFERENT DATA FETCHING:** Each standard has its own dedicated function
    - **DETAILED CALCULATIONS:** Shows complete breakdown for all products
    """)
    
    # Initialize session state for form inputs
    if 'weight_form_submitted' not in st.session_state:
        st.session_state.weight_form_submitted = False
    
    # Main input form with enhanced workflow
    with st.form("weight_calculator_rectified"):
        st.markdown("### Product Standards Selection")
        
        col1, col2, col3, col4, col5 = st.columns(5)
        
        with col1:
            # A. Product Type
            product_options = get_available_products()
            selected_product = st.selectbox(
                "A. Product Type",
                product_options,
                key="weight_calc_product_select"
            )
            
            # Show product info
            if selected_product != "Select Product":
                st.caption(f"Selected: {selected_product}")
        
        with col2:
            # B. Series (Inch/Metric)
            series_options = get_series_for_product(selected_product)
            selected_series = st.selectbox(
                "B. Series",
                series_options,
                key="weight_calc_series_select"
            )
            
            # Show series info with unit conversion note
            if selected_series != "Select Series":
                st.caption(f"Series: {selected_series}")
                if selected_series == "Inch":
                    st.caption("⚠️ Inch dimensions will be converted to mm")
        
        with col3:
            # C. Standard (based on Product + Series) - DISABLED FOR THREADED ROD
            if selected_product == "Threaded Rod":
                st.info("Standard not required for Threaded Rod")
                selected_standard = "Not Required"
                st.session_state.weight_calc_standard_select = "Not Required"
            else:
                standard_options = get_standards_for_product_series(selected_product, selected_series)
                selected_standard = st.selectbox(
                    "C. Standard",
                    standard_options,
                    key="weight_calc_standard_select"
                )
            
            # Show standard info
            if selected_standard != "Select Standard" and selected_standard != "Not Required":
                st.caption(f"Standard: {selected_standard}")
        
        with col4:
            # D. Size (based on Standard + Product) - DISABLED FOR THREADED ROD
            if selected_product == "Threaded Rod":
                st.info("Size not required for Threaded Rod")
                selected_size = "Not Required"
                st.session_state.weight_calc_size_select = "Not Required"
            else:
                size_options = get_sizes_for_standard_product(selected_standard, selected_product)
                selected_size = st.selectbox(
                    "D. Size",
                    size_options,
                    key="weight_calc_size_select"
                )
            
            # Show size info
            if selected_size != "Select Size" and selected_size != "Not Required":
                st.caption(f"Size: {selected_size}")
        
        with col5:
            # E. Grade (only for ISO 4014 Hex Bolt)
            if selected_standard == "ISO 4014" and selected_product == "Hex Bolt":
                grade_options = get_available_grades_for_standard_product(selected_standard, selected_product)
                selected_grade = st.selectbox(
                    "E. Product Grade",
                    grade_options,
                    key="weight_calc_grade_select"
                )
                
                # Show grade info
                if selected_grade != "All":
                    st.caption(f"Grade: {selected_grade}")
                    st.caption("⚠️ Different dimensions for A/B grades")
            else:
                st.info("Grade not applicable")
                selected_grade = "Not Applicable"
                st.session_state.weight_calc_grade_select = "Not Applicable"
        
        st.markdown("---")
        st.markdown("### Diameter Specification")
        
        col1, col2 = st.columns(2)
        
        with col1:
            # F. Cylinder Diameter Type
            diameter_type_options = ["Blank Diameter", "Pitch Diameter"]
            selected_diameter_type = st.radio(
                "F. Cylinder Diameter Type",
                diameter_type_options,
                key="weight_calc_diameter_type_select"
            )
        
        with col2:
            # G. Conditional Input based on Diameter Type
            if selected_diameter_type == "Blank Diameter":
                st.markdown("**Blank Diameter Input**")
                dia_col1, dia_col2 = st.columns(2)
                with dia_col1:
                    blank_diameter = st.number_input(
                        "Blank Diameter Value",
                        min_value=0.1,
                        value=10.0,
                        step=0.1,
                        format="%.4f",
                        key="weight_calc_blank_diameter_input"
                    )
                with dia_col2:
                    blank_dia_unit = st.selectbox(
                        "Unit",
                        ["mm", "inch", "ft"],
                        key="weight_calc_blank_dia_unit_select"
                    )
                
                st.caption(f"Blank Diameter: {blank_diameter:.4f} {blank_dia_unit}")
                if blank_dia_unit != "mm":
                    st.caption(f"→ {convert_to_mm(blank_diameter, blank_dia_unit):.4f} mm (converted)")
                else:
                    st.caption("✓ Already in mm - no conversion needed")
            
            else:  # Pitch Diameter
                st.markdown("**Thread Specification**")
                
                # Thread Standard
                thread_std_options = get_thread_standards_for_series(selected_series)
                if selected_series == "Select Series":
                    thread_std_options = ["Select Thread Standard"]
                
                thread_standard = st.selectbox(
                    "Thread Standard",
                    thread_std_options,
                    key="weight_calc_thread_standard_select"
                )
                
                if thread_standard != "Select Thread Standard":
                    # Thread Size
                    thread_size_options = get_thread_sizes_enhanced(thread_standard)
                    thread_size = st.selectbox(
                        "Thread Size",
                        thread_size_options,
                        key="weight_calc_thread_size_select"
                    )
                    
                    # Thread Class - Only show for Inch series
                    if selected_series == "Inch" and thread_standard == "ASME B1.1":
                        thread_class_options = get_thread_classes_enhanced(thread_standard)
                        if len(thread_class_options) == 1:  # Only "All"
                            thread_class_options = ["2A", "3A", "1A"]
                        thread_class = st.selectbox(
                            "Tolerance Class",
                            thread_class_options,
                            key="weight_calc_thread_class_select"
                        )
                    else:
                        # For metric threads or non-ASME standards, don't show tolerance class
                        thread_class = "N/A"
                        st.caption("Tolerance Class: Not applicable for metric threads")
                    
                    st.caption(f"Thread: {thread_standard}, Size: {thread_size}, Class: {thread_class}")
                    
                    # Show pitch diameter information for ALL products using Pitch Diameter
                    if selected_diameter_type == "Pitch Diameter" and thread_size != "All":
                        with LoadingManager.show_loading_spinner("Fetching thread data..."):
                            pitch_diameter = get_pitch_diameter_from_thread_data(thread_standard, thread_size, thread_class)
                        if pitch_diameter is not None:
                            # Store the pitch diameter in session state for calculation
                            st.session_state.pitch_diameter_value = pitch_diameter
                            
                            # Display information
                            if selected_series == "Inch":
                                # For Inch series, pitch diameter from ASME B1.1 is in inches
                                pitch_diameter_mm = pitch_diameter * 25.4
                                st.success(f"Pitch Diameter (Min): {pitch_diameter:.4f} in → {pitch_diameter_mm:.4f} mm")
                                st.info(f"⚠️ For calculation: {pitch_diameter:.4f} in will be converted to mm")
                            else:
                                # For Metric series, pitch diameter is already in mm
                                st.success(f"Pitch Diameter (Min): {pitch_diameter:.4f} mm")
                                st.info(f"✓ For calculation: {pitch_diameter:.4f} mm will be used directly")
                        else:
                            st.warning("Pitch diameter not found in thread data")
        
        st.markdown("---")
        st.markdown("### Additional Parameters")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            # Length - UPDATED WITH FT UNIT AND 0.0000 FORMAT
            length_col1, length_col2 = st.columns(2)
            with length_col1:
                length = st.number_input(
                    "Length",
                    min_value=0.1,
                    value=50.0,
                    step=0.1,
                    format="%.4f",
                    key="weight_calc_length_input"
                )
            with length_col2:
                length_unit = st.selectbox(
                    "Unit",
                    ["mm", "inch", "ft", "meter"],
                    key="weight_calc_length_unit_select"
                )
            
            # Show length conversion info
            if length_unit != "mm":
                st.caption(f"→ {convert_to_mm(length, length_unit):.4f} mm (converted)")
            else:
                st.caption("✓ Already in mm - no conversion needed")
        
        with col2:
            # Material - UPDATED WITH MORE MATERIALS
            material_options = ["Carbon Steel", "Stainless Steel", "Alloy Steel", "Brass", "Aluminum", 
                              "Copper", "Titanium", "Bronze", "Inconel", "Monel", "Nickel"]
            material = st.selectbox(
                "Material",
                material_options,
                key="weight_calc_material_select"
            )
            
            # Show material density in g/cm³
            density = get_material_density_rectified(material)
            st.caption(f"Density: {density:.4f} g/cm³")
        
        with col3:
            # Calculate button space
            st.markdown("<br>", unsafe_allow_html=True)
            calculate_btn = st.form_submit_button("Calculate Weight", use_container_width=True, type="primary")
    
    # Handle form submission
    if calculate_btn:
        # Validate inputs
        validation_errors = []
        
        if selected_product == "Select Product":
            validation_errors.append("Please select a Product Type")
        if selected_series == "Select Series":
            validation_errors.append("Please select a Series")
        
        # Skip Standard and Size validation for Threaded Rod
        if selected_product != "Threaded Rod":
            if selected_standard == "Select Standard":
                validation_errors.append("Please select a Standard")
            if selected_size == "Select Size":
                validation_errors.append("Please select a Size")
        
        if selected_diameter_type == "Pitch Diameter" and thread_standard == "Select Thread Standard":
            validation_errors.append("Please select a Thread Standard for Pitch Diameter")
        
        if validation_errors:
            for error in validation_errors:
                st.error(error)
        else:
            # Prepare calculation parameters
            calculation_params = {
                'product_type': selected_product,
                'diameter_type': selected_diameter_type,
                'material': material,
                'length': length,
                'length_unit': length_unit,
                'standard': selected_standard,
                'size': selected_size,
                'grade': selected_grade if selected_standard == "ISO 4014" and selected_product == "Hex Bolt" else "All"
            }
            
            # Add diameter parameters based on type
            if selected_diameter_type == "Blank Diameter":
                calculation_params.update({
                    'diameter_value': blank_diameter,
                    'diameter_unit': blank_dia_unit
                })
            else:
                # For pitch diameter, use the pitch diameter value stored in session state
                if 'pitch_diameter_value' in st.session_state and st.session_state.pitch_diameter_value is not None:
                    pitch_diameter = st.session_state.pitch_diameter_value
                    
                    # Determine unit based on thread standard
                    if selected_series == "Inch":
                        calculation_params.update({
                            'diameter_value': pitch_diameter,
                            'diameter_unit': 'inch'  # ASME B1.1 data is in inches
                        })
                        st.success(f"Using Pitch Diameter: {pitch_diameter:.4f} inches for calculation")
                    else:
                        # For Metric series, pitch diameter is in mm
                        calculation_params.update({
                            'diameter_value': pitch_diameter,
                            'diameter_unit': 'mm'  # ISO thread data is in mm
                        })
                        st.success(f"Using Pitch Diameter: {pitch_diameter:.4f} mm for calculation")
                else:
                    st.error("Pitch diameter not available for calculation")
                    return
            
            # Perform calculation using FIXED function
            with LoadingManager.show_loading_spinner("Calculating weight..."):
                result = calculate_weight_rectified(calculation_params)
            
            if result:
                st.session_state.weight_calc_result = result
                st.session_state.weight_calculation_performed = True
                
                # Save to calculation history
                calculation_data = {
                    'product': selected_product,
                    'series': selected_series,
                    'size': selected_size if selected_product != "Threaded Rod" else "Threaded Rod",
                    'grade': selected_grade if selected_standard == "ISO 4014" and selected_product == "Hex Bolt" else "N/A",
                    'diameter': f"{blank_diameter:.4f} {blank_dia_unit}" if selected_diameter_type == 'Blank Diameter' else f"{thread_size} mm",
                    'length': f"{length:.4f} {length_unit}",
                    'material': material,
                    'weight_kg': result['weight_kg'],
                    'weight_lb': result['weight_lb'],
                    'timestamp': datetime.now().isoformat()
                }
                st.session_state.calculation_history.append(calculation_data)
                
                st.success("**Weight Calculation Completed Successfully!**")
    
    # Display current selection summary
    if selected_product != "Select Product":
        st.markdown("### Current Selection Summary")
        
        summary_col1, summary_col2 = st.columns(2)
        
        with summary_col1:
            st.markdown(f"""
            **Product Standards:**
            - **Product Type:** {selected_product}
            - **Series:** {selected_series}
            - **Standard:** {selected_standard if selected_product != "Threaded Rod" else "Not Required (Threaded Rod)"}
            - **Size:** {selected_size if selected_product != "Threaded Rod" else "Not Required (Threaded Rod)"}
            - **Grade:** {selected_grade if selected_standard == "ISO 4014" and selected_product == "Hex Bolt" else "Not Applicable"}
            """)
        
        with summary_col2:
            if selected_diameter_type == "Blank Diameter":
                st.markdown(f"""
                **Diameter Specification:**
                - **Type:** {selected_diameter_type}
                - **Value:** {blank_diameter:.4f} {blank_dia_unit}
                """)
            else:
                st.markdown(f"""
                **Diameter Specification:**
                - **Type:** {selected_diameter_type}
                - **Thread Standard:** {thread_standard}
                - **Thread Size:** {thread_size}
                - **Thread Class:** {thread_class}
                """)
        
        st.markdown(f"""
        **Additional Parameters:**
        - **Length:** {length:.4f} {length_unit}
        - **Material:** {material}
        - **Material Density:** {get_material_density_rectified(material):.4f} g/cm³
        - **Unit Handling:** {'Conversion to mm' if (selected_diameter_type == 'Blank Diameter' and blank_dia_unit != 'mm') or length_unit != 'mm' else 'All in mm - no conversion needed'}
        """)
    
    # Display calculation results - FIXED WITH DETAILED PARAMETERS FOR ALL PRODUCTS
    if st.session_state.weight_calculation_performed and st.session_state.weight_calc_result:
        result = st.session_state.weight_calc_result
        
        st.markdown("### Calculation Results")
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("Weight (kg)", f"{result['weight_kg']:.4f}")
        with col2:
            st.metric("Weight (grams)", f"{result['weight_g']:.4f}")
        with col3:
            st.metric("Weight (pounds)", f"{result['weight_lb']:.4f}")
        with col4:
            st.metric("Density", f"{result['density_g_cm3']:.4f} g/cm³")
        
        # FIXED: Enhanced detailed results for ALL PRODUCTS
        with st.expander("📐 Detailed Calculation Parameters - ALL PRODUCTS"):
            calculation_method = result.get('calculation_method', 'Standard Cylinder Formula')
            
            # Show calculation method
            st.markdown(f"**Calculation Method:** `{calculation_method}`")
            
            # Show ALL dimensions used in calculation
            if 'dimensions_used' in result:
                dimensions = result['dimensions_used']
                
                st.markdown("### 📏 Dimensions Used in Calculation")
                
                # Create columns for better layout
                dim_col1, dim_col2 = st.columns(2)
                
                with dim_col1:
                    st.markdown("#### Input Values")
                    for key, value in dimensions.items():
                        if 'input' in key.lower():
                            st.markdown(f"- **{key.replace('_', ' ').title()}:** `{value}`")
                
                with dim_col2:
                    st.markdown("#### Calculated Values (Millimeters)")
                    for key, value in dimensions.items():
                        if 'calculation' in key.lower() and 'mm' in key:
                            st.markdown(f"- **{key.replace('_', ' ').title()}:** `{value} mm`")
                
                # Show volume calculations
                st.markdown("#### Volume Calculations")
                vol_col1, vol_col2 = st.columns(2)
                with vol_col1:
                    if 'shank_volume_mm3' in dimensions:
                        st.markdown(f"- **Shank Volume:** `{dimensions['shank_volume_mm3']} mm³`")
                    if 'head_volume_mm3' in dimensions:
                        st.markdown(f"- **Head Volume:** `{dimensions['head_volume_mm3']} mm³`")
                    if 'total_volume_mm3' in dimensions:
                        st.markdown(f"- **Total Volume:** `{dimensions['total_volume_mm3']} mm³`")
                
                with vol_col2:
                    if 'total_volume_cm3' in dimensions:
                        st.markdown(f"- **Total Volume:** `{dimensions['total_volume_cm3']} cm³`")
                    if 'volume_cm3' in dimensions:
                        st.markdown(f"- **Volume:** `{dimensions['volume_cm3']} cm³`")
                    st.markdown(f"- **Density:** `{dimensions['density_g_cm3']} g/cm³`")
            
            # Show formula details for all product types
            if 'formula_details' in result:
                st.markdown("### 🧮 Formula Details")
                formulas = result['formula_details']
                for formula_name, formula in formulas.items():
                    st.markdown(f"- **{formula_name.replace('_', ' ').title()}:** `{formula}`")
            
            # Show specific product details
            if 'Socket Head Formula' in calculation_method:
                st.markdown("### 🔧 Socket Head Product Specific Details")
                st.markdown(f"""
                **Shank Volume Calculation:**
                - Formula: 0.7853 × (diameter)² × length
                - Diameter: `{result['diameter_mm']:.4f} mm`
                - Length: `{result['length_mm']:.4f} mm`
                - Result: `{result['shank_volume_mm3']:.4f} mm³`
                
                **Head Volume Calculation:**
                - Formula: 0.7853 × (head_diameter_min)² × head_height_min
                - Head Diameter: `{result['head_diameter_mm']:.4f} mm`
                - Head Height: `{result['head_height_mm']:.4f} mm`
                - Result: `{result['head_volume_mm3']:.4f} mm³`
                
                **Total Volume:**
                - Formula: shank_volume + head_volume
                - Result: `{result['total_volume_mm3']:.4f} mm³`
                - Converted to cm³: `{result['total_volume_cm3']:.4f} cm³`
                
                **Weight Calculation:**
                - Formula: total_volume_cm³ × density_g/cm³
                - Result: `{result['weight_g']:.4f} g` = `{result['weight_kg']:.4f} kg`
                """)
            
            elif 'Hex Product Formula' in calculation_method:
                st.markdown("### 🔧 Hex Product Specific Details")
                st.markdown(f"""
                **Shank Volume Calculation:**
                - Formula: 0.7853 × (diameter)² × length
                - Diameter: `{result['diameter_mm']:.4f} mm`
                - Length: `{result['length_mm']:.4f} mm`
                - Result: `{result['shank_volume_mm3']:.4f} mm³`
                
                **Head Volume Calculation:**
                - Formula: 0.65 × side_length² × head_height
                - Width Across Flats: `{result['width_across_flats_mm']:.4f} mm`
                - Side Length: `{result['side_length_mm']:.4f} mm`
                - Head Height: `{result['head_height_mm']:.4f} mm`
                - Result: `{result['head_volume_mm3']:.4f} mm³`
                
                **Total Volume:**
                - Formula: shank_volume + head_volume
                - Result: `{result['total_volume_mm3']:.4f} mm³`
                - Converted to cm³: `{result['total_volume_cm3']:.4f} cm³`
                
                **Weight Calculation:**
                - Formula: total_volume_cm³ × density_g/cm³
                - Result: `{result['weight_g']:.4f} g` = `{result['weight_kg']:.4f} kg`
                """)
            
            elif 'Threaded Rod Cylinder Formula' in calculation_method:
                st.markdown("### 🔧 Threaded Rod Specific Details")
                st.markdown(f"""
                **Volume Calculation:**
                - Formula: 0.7853 × (diameter)² × length
                - Diameter: `{result['diameter_mm']:.4f} mm`
                - Length: `{result['length_mm']:.4f} mm`
                - Result: `{result['volume_mm3']:.4f} mm³`
                - Converted to cm³: `{result['volume_cm3']:.4f} cm³`
                
                **Weight Calculation:**
                - Formula: volume_cm³ × density_g/cm³
                - Result: `{result['weight_g']:.4f} g` = `{result['weight_kg']:.4f} kg`
                """)

def show_batch_calculator_rectified():
    """FIXED batch calculator with proper data fetching"""
    
    st.markdown("### Batch Weight Calculator - FIXED WORKFLOW")
    
    st.info("""
    **FIXED BATCH PROCESSING:** 
    - Separate data fetching for Socket Head products
    - Same formula for all socket head products
    - Detailed calculation parameters for all products
    """)
    
    # Download template
    st.markdown("### Download FIXED Batch Template")
    template_df = ExportTemplateManager.get_weight_calc_template()
    csv_template = template_df.to_csv(index=False)
    st.download_button(
        label="Download FIXED Batch Template (CSV)",
        data=csv_template,
        file_name="fixed_batch_weight_template.csv",
        mime="text/csv",
        use_container_width=True
    )
    
    uploaded_file = st.file_uploader("Upload CSV/Excel file for batch processing", 
                                   type=["csv", "xlsx"],
                                   key="batch_upload_fixed")
    
    if uploaded_file:
        try:
            if uploaded_file.name.endswith('.xlsx'):
                batch_df = pd.read_excel(uploaded_file)
            else:
                batch_df = pd.read_csv(uploaded_file)
            
            st.success("File uploaded successfully!")
            st.write("Preview of uploaded data:")
            st.dataframe(batch_df.head())
            
            # Validate required columns
            required_cols = ['Product_Type', 'Series', 'Diameter_Type', 'Length']
            missing_cols = [col for col in required_cols if col not in batch_df.columns]
            
            if missing_cols:
                st.error(f"Missing required columns: {missing_cols}")
            else:
                if st.button("Process Batch Calculation", use_container_width=True, key="process_batch_fixed"):
                    with LoadingManager.show_loading_spinner("Processing batch calculations..."):
                        LoadingManager.show_progress_bar(1, 5, "Processing batch")
                        st.info("FIXED batch processing with separate data fetching ready for implementation")
                        st.write(f"Records to process: {len(batch_df)}")
                    
        except Exception as e:
            st.error(f"Error reading file: {str(e)}")
            LoadingManager.log_operation("Batch File Upload", False, str(e))

def show_rectified_calculations():
    """Fixed calculations page with proper data fetching for ALL products"""
    
    tab1, tab2, tab3 = st.tabs(["Single Calculator", "Batch Calculator", "Analytics"])
    
    with tab1:
        show_weight_calculator_rectified()
    
    with tab2:
        show_batch_weight_calculator()
    
    with tab3:
        st.markdown("### Calculation Analytics - FIXED")
        st.info("Analytics dashboard will show calculation history and trends after weight calculations are performed.")
        
        if 'calculation_history' in st.session_state and st.session_state.calculation_history:
            st.write("Calculation history will be displayed here")
        else:
            st.write("No calculation history yet. Perform calculations to see analytics here.")

# ======================================================
# ENHANCED DATA QUALITY INDICATORS
# ======================================================
def show_data_quality_indicators():
    """Show data quality and validation indicators"""
    st.sidebar.markdown("---")
    with st.sidebar.expander("Data Quality Status"):
        if not df.empty:
            total_rows = len(df)
            missing_data = df.isnull().sum().sum()
            completeness = ((total_rows * len(df.columns) - missing_data) / (total_rows * len(df.columns))) * 100
            st.markdown(f'<div class="data-quality-indicator quality-good">Main Data: {completeness:.1f}% Complete</div>', unsafe_allow_html=True)
        else:
            st.markdown('<div class="data-quality-indicator quality-error">Main Data: Not Loaded</div>', unsafe_allow_html=True)
        
        if not df_iso4014.empty:
            st.markdown(f'<div class="data-quality-indicator quality-good">ISO 4014: {len(df_iso4014)} Records</div>', unsafe_allow_html=True)
        else:
            st.markdown('<div class="data-quality-indicator quality-warning">ISO 4014: Limited Access</div>', unsafe_allow_html=True)
        
        if st.session_state.din7991_loaded:
            st.markdown(f'<div class="data-quality-indicator quality-good">DIN-7991: {len(df_din7991)} Records</div>', unsafe_allow_html=True)
        else:
            st.markdown('<div class="data-quality-indicator quality-warning">DIN-7991: Limited Access</div>', unsafe_allow_html=True)
        
        if st.session_state.asme_b18_3_loaded:
            st.markdown(f'<div class="data-quality-indicator quality-good">ASME B18.3: {len(df_asme_b18_3)} Records</div>', unsafe_allow_html=True)
        else:
            st.markdown('<div class="data-quality-indicator quality-warning">ASME B18.3: Limited Access</div>', unsafe_allow_html=True)
        
        if not df_mechem.empty:
            st.markdown(f'<div class="data-quality-indicator quality-good">Mech & Chem: {len(df_mechem)} Records</div>', unsafe_allow_html=True)
            st.markdown(f'<div style="font-size: 0.8rem; margin: 0.1rem 0;">Property Classes: {len(st.session_state.property_classes)}</div>', unsafe_allow_html=True)
        else:
            st.markdown('<div class="data-quality-indicator quality-warning">Mech & Chem: Limited Access</div>', unsafe_allow_html=True)
        
        thread_status = []
        for standard, url in thread_files.items():
            df_thread = load_thread_data_enhanced(standard)
            if not df_thread.empty:
                thread_status.append(f"{standard}: OK")
            else:
                thread_status.append(f"{standard}: Limited")
        
        st.markdown(f'<div class="data-quality-indicator quality-good">Thread Data: Available</div>', unsafe_allow_html=True)
        for status in thread_status:
            st.markdown(f'<div style="font-size: 0.8rem; margin: 0.1rem 0;">{status}</div>', unsafe_allow_html=True)

# ======================================================
# Enhanced Export Functionality
# ======================================================
def export_to_excel(df, filename_prefix):
    """Export dataframe to Excel with formatting"""
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            with pd.ExcelWriter(tmp.name, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Data', index=False)
                
                workbook = writer.book
                worksheet = writer.sheets['Data']
                
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            LoadingManager.log_operation("Export to Excel", True, f"Rows: {len(df)}")
            return tmp.name
    except Exception as e:
        st.error(f"Export error: {str(e)}")
        LoadingManager.log_operation("Export to Excel", False, str(e))
        return None

def enhanced_export_data(filtered_df, export_format):
    """Enhanced export with multiple format options"""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    if export_format == "Excel":
        with LoadingManager.show_loading_spinner("Generating Excel file..."):
            excel_file = export_to_excel(filtered_df, f"fastener_data_{timestamp}")
        if excel_file:
            with open(excel_file, 'rb') as f:
                st.download_button(
                    label="Download Excel File",
                    data=f,
                    file_name=f"fastener_data_{timestamp}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key=f"excel_export_{timestamp}"
                )
    else:
        csv_data = filtered_df.to_csv(index=False)
        st.download_button(
            label="Download CSV File",
            data=csv_data,
            file_name=f"fastener_data_{timestamp}.csv",
            mime="text/csv",
            use_container_width=True,
            key=f"csv_export_{timestamp}"
        )

# ======================================================
# Enhanced Calculation History
# ======================================================
def save_calculation_history(calculation_data):
    """Save calculation to history"""
    if 'calculation_history' not in st.session_state:
        st.session_state.calculation_history = []
    
    calculation_data['timestamp'] = datetime.now().isoformat()
    st.session_state.calculation_history.append(calculation_data)
    
    if len(st.session_state.calculation_history) > 20:
        st.session_state.calculation_history = st.session_state.calculation_history[-20:]
    
    LoadingManager.log_operation("Save Calculation History", True, f"Total calculations: {len(st.session_state.calculation_history)}")

def show_calculation_history():
    """Display calculation history"""
    if 'calculation_history' in st.session_state and st.session_state.calculation_history:
        st.markdown("### Recent Calculations")
        for calc in reversed(st.session_state.calculation_history[-5:]):
            with st.container():
                st.markdown(f"""
                <div class="calculation-card">
                    <strong>{calc.get('product', 'N/A')}</strong> | 
                    Size: {calc.get('size', 'N/A')} | 
                    Grade: {calc.get('grade', 'N/A')} |
                    Weight: {calc.get('weight_kg', 'N/A'):.4f} kg
                    <br><small>{calc.get('timestamp', '')}</small>
                </div>
                """, unsafe_allow_html=True)

# ======================================================
# COMPLETELY FIXED: PROFESSIONAL PRODUCT CARD DISPLAY
# ======================================================
def extract_product_details(row):
    """Extract product details from dataframe row and map to card format - COMPLETELY FIXED VERSION"""
    try:
        # Convert row to dictionary if it's a Series
        if hasattr(row, 'to_dict'):
            row_dict = row.to_dict()
        else:
            row_dict = dict(row)
        
        # Safely extract basic product information with multiple fallbacks
        details = {
            'Product': row_dict.get('Product', 'Hex Bolt'),
            'Size': row_dict.get('Size', 'N/A'),
            'Standards': row_dict.get('Standards', row_dict.get('Standard', 'ASME B18.2.1')),
            'Thread': row_dict.get('Thread', row_dict.get('Thread_Size', 'N/A')),
            'Product Grade': row_dict.get('Product Grade', row_dict.get('Grade', 'N/A')),
        }
        
        # Helper function to safely get values with multiple possible column names
        def get_value(row_dict, possible_keys, default='N/A'):
            for key in possible_keys:
                value = row_dict.get(key)
                if value is not None and pd.notna(value) and str(value).strip() != '':
                    return str(value).strip()
            return default
        
        # Map ALL possible dimensional specifications with comprehensive column name variations
        details.update({
            # Body Diameter - comprehensive column name variations
            'Body_Dia_Min': get_value(row_dict, [
                'Body_Diameter_Min', 'Body_Dia_Min', 'Basic_Major_Diameter_Min', 'Major_Dia_Min',
                'Body Diameter Min', 'Body Dia Min', 'Body_D_Min', 'Body_Dia_Min'
            ]),
            'Body_Dia_Max': get_value(row_dict, [
                'Body_Diameter_Max', 'Body_Dia_Max', 'Basic_Major_Diameter_Max', 'Major_Dia_Max',
                'Body Diameter Max', 'Body Dia Max', 'Body_D_Max', 'Body_Dia_Max'
            ]),
            
            # Width Across Flats - comprehensive column name variations
            'Width_Across_Flats_Min': get_value(row_dict, [
                'Width_Across_Flats_Min', 'W_Across_Flats_Min', 'Width_Min', 'W_Min',
                'Width Across Flats Min', 'W Across Flats Min', 'Width_Flats_Min', 'W_Flats_Min',
                'Width_Across_Flats', 'W_Across_Flats', 'Width_Flats', 'W_Flats'
            ]),
            'Width_Across_Flats_Max': get_value(row_dict, [
                'Width_Across_Flats_Max', 'W_Across_Flats_Max', 'Width_Max', 'W_Max',
                'Width Across Flats Max', 'W Across Flats Max', 'Width_Flats_Max', 'W_Flats_Max'
            ]),
            
            # Width Across Corners - comprehensive column name variations
            'Width_Across_Corners_Min': get_value(row_dict, [
                'Width_Across_Corners_Min', 'W_Across_Corners_Min', 'Width_Across_Corners_Min',
                'Width Across Corners Min', 'W Across Corners Min', 'Width_Corners_Min', 'W_Corners_Min',
                'Width_Across_Corners', 'W_Across_Corners', 'Width_Corners', 'W_Corners'
            ]),
            'Width_Across_Corners_Max': get_value(row_dict, [
                'Width_Across_Corners_Max', 'W_Across_Corners_Max', 'Width_Across_Corners_Max',
                'Width Across Corners Max', 'W Across Corners Max', 'Width_Corners_Max', 'W_Corners_Max'
            ]),
            
            # Head Height - comprehensive column name variations
            'Head_Height_Min': get_value(row_dict, [
                'Head_Height_Min', 'Head_Ht_Min', 'Head_Height', 'k_min', 'Head Height Min',
                'Head_H_Min', 'Head_Ht', 'Head_Height_Min', 'Head_H_Min', 'k'
            ]),
            'Head_Height_Max': get_value(row_dict, [
                'Head_Height_Max', 'Head_Ht_Max', 'Head_Height_Max', 'k_max',
                'Head Height Max', 'Head_H_Max', 'Head_Ht_Max'
            ]),
            
            # Radius of Fillet - comprehensive column name variations
            'Radius_Fillet_Min': get_value(row_dict, [
                'Radius_Fillet_Min', 'Fillet_Radius_Min', 'Radius_Fillet', 'Fillet_Radius',
                'Radius Fillet Min', 'Fillet Radius Min', 'Radius_F_Min', 'Fillet_R_Min',
                'Radius_Fillet', 'Fillet_Radius', 'Radius_F', 'Fillet_R'
            ]),
            'Radius_Fillet_Max': get_value(row_dict, [
                'Radius_Fillet_Max', 'Fillet_Radius_Max', 'Radius_Fillet_Max',
                'Radius Fillet Max', 'Fillet Radius Max', 'Radius_F_Max', 'Fillet_R_Max'
            ]),
            
            # Washer Face Thickness - comprehensive column name variations
            'Washer_Face_Thickness_Min': get_value(row_dict, [
                'Washer_Face_Thickness_Min', 'Washer_Face_Min', 'Washer_Thickness_Min',
                'Washer Face Thickness Min', 'Washer Face Min', 'Washer_Thickness_Min',
                'Washer_Face_Thickness', 'Washer_Face', 'Washer_Thickness'
            ]),
            'Washer_Face_Thickness_Max': get_value(row_dict, [
                'Washer_Face_Thickness_Max', 'Washer_Face_Max', 'Washer_Thickness_Max',
                'Washer Face Thickness Max', 'Washer Face Max', 'Washer_Thickness_Max'
            ]),
            
            # Wrenching Height - comprehensive column name variations
            'Wrenching_Height_Min': get_value(row_dict, [
                'Wrenching_Height_Min', 'Wrenching_Ht_Min', 'Wrenching_Height',
                'Wrenching Height Min', 'Wrenching_H_Min', 'Wrenching_Ht',
                'Wrenching_Height_Min', 'Wrenching_H_Min'
            ]),
            
            # Total Runout - comprehensive column name variations
            'Total_Runout_Max': get_value(row_dict, [
                'Total_Runout_Max', 'Runout_Max', 'Total_Runout',
                'Total Runout Max', 'Runout Max', 'Total_Runout_Max',
                'Runout', 'Total_Runout'
            ])
        })
        
        # Debug information
        if st.session_state.debug_mode:
            st.sidebar.write("Extracted Product Details:")
            for key, value in details.items():
                if value != 'N/A':
                    st.sidebar.write(f"  {key}: {value}")
        
        return details
        
    except Exception as e:
        st.error(f"Error extracting product details: {str(e)}")
        # Return basic details even if extraction fails
        return {
            'Product': 'Hex Bolt',
            'Size': 'N/A',
            'Standards': 'ASME B18.2.1',
            'Thread': 'N/A',
            'Product Grade': 'N/A',
            'Body_Dia_Min': 'N/A',
            'Body_Dia_Max': 'N/A',
            'Width_Across_Flats_Min': 'N/A',
            'Width_Across_Flats_Max': 'N/A',
            'Width_Across_Corners_Min': 'N/A',
            'Width_Across_Corners_Max': 'N/A',
            'Head_Height_Min': 'N/A',
            'Head_Height_Max': 'N/A',
            'Radius_Fillet_Min': 'N/A',
            'Radius_Fillet_Max': 'N/A',
            'Washer_Face_Thickness_Min': 'N/A',
            'Washer_Face_Thickness_Max': 'N/A',
            'Wrenching_Height_Min': 'N/A',
            'Total_Runout_Max': 'N/A'
        }

def show_professional_product_card(product_details):
    """Display a beautiful professional product specification card - COMPLETELY FIXED"""
    
    # Extract product details with safe defaults
    product_name = product_details.get('Product', 'Hex Bolt')
    size = product_details.get('Size', 'N/A')
    standard = product_details.get('Standards', 'ASME B18.2.1')
    thread = product_details.get('Thread', 'N/A')
    grade = product_details.get('Product Grade', 'N/A')
    
    # Get current date and user info
    current_date = datetime.now().strftime('%d/%m/%Y')
    generated_by = "Partha Sharma"
    
    # Create the professional card HTML
    card_html = f"""
    <div class="professional-card">
        <div class="card-header">
            <div>
                <h1 class="card-title">{product_name}</h1>
                <p class="card-subtitle">Size: {size} | Standard: {standard} | Grade: {grade}</p>
            </div>
            <div class="card-company">JSC India</div>
        </div>
        
        <div class="specification-grid">
            <!-- Dimensional Specifications Group -->
            <div class="oracle11g-card">
                <div class="filter-header">Dimensional Specifications</div>
                
                <!-- Body Diameter -->
                <div class="spec-row">
                    <div class="spec-label-min">Body Dia (Min)</div>
                    <div class="spec-dimension">Body Diameter</div>
                    <div class="spec-label-max">Body Dia (Max)</div>
                </div>
                <div class="spec-row">
                    <div class="spec-value">{product_details.get('Body_Dia_Min', 'N/A')}</div>
                    <div class="spec-dimension"></div>
                    <div class="spec-value">{product_details.get('Body_Dia_Max', 'N/A')}</div>
                </div>
                
                <!-- Width Across Flats -->
                <div class="spec-row">
                    <div class="spec-label-min">Width Across Flats (Min)</div>
                    <div class="spec-dimension">Width Across Flats</div>
                    <div class="spec-label-max">Width Across Flats (Max)</div>
                </div>
                <div class="spec-row">
                    <div class="spec-value">{product_details.get('Width_Across_Flats_Min', 'N/A')}</div>
                    <div class="spec-dimension"></div>
                    <div class="spec-value">{product_details.get('Width_Across_Flats_Max', 'N/A')}</div>
                </div>
                
                <!-- Width Across Corners -->
                <div class="spec-row">
                    <div class="spec-label-min">Width Across Corners (Min)</div>
                    <div class="spec-dimension">Width Across Corners</div>
                    <div class="spec-label-max">Width Across Corners (Max)</div>
                </div>
                <div class="spec-row">
                    <div class="spec-value">{product_details.get('Width_Across_Corners_Min', 'N/A')}</div>
                    <div class="spec-dimension"></div>
                    <div class="spec-value">{product_details.get('Width_Across_Corners_Max', 'N/A')}</div>
                </div>
                
                <!-- Head Height -->
                <div class="spec-row">
                    <div class="spec-label-min">Head Height (Min)</div>
                    <div class="spec-dimension">Head Height</div>
                    <div class="spec-label-max">Head Height (Max)</div>
                </div>
                <div class="spec-row">
                    <div class="spec-value">{product_details.get('Head_Height_Min', 'N/A')}</div>
                    <div class="spec-dimension"></div>
                    <div class="spec-value">{product_details.get('Head_Height_Max', 'N/A')}</div>
                </div>
                
                <!-- Radius of Fillet -->
                <div class="spec-row">
                    <div class="spec-label-min">Radius of Fillet (Min)</div>
                    <div class="spec-dimension">Radius of Fillet</div>
                    <div class="spec-label-max">Radius of Fillet (Max)</div>
                </div>
                <div class="spec-row">
                    <div class="spec-value">{product_details.get('Radius_Fillet_Min', 'N/A')}</div>
                    <div class="spec-dimension"></div>
                    <div class="spec-value">{product_details.get('Radius_Fillet_Max', 'N/A')}</div>
                </div>
                
                <!-- Washer Face Thickness -->
                <div class="spec-row">
                    <div class="spec-label-min">Washer Face Thickness (Min)</div>
                    <div class="spec-dimension">Washer Face Thickness</div>
                    <div class="spec-label-max">Washer Face Thickness (Max)</div>
                </div>
                <div class="spec-row">
                    <div class="spec-value">{product_details.get('Washer_Face_Thickness_Min', 'N/A')}</div>
                    <div class="spec-dimension"></div>
                    <div class="spec-value">{product_details.get('Washer_Face_Thickness_Max', 'N/A')}</div>
                </div>
                
                <!-- Wrenching Height -->
                <div class="spec-row">
                    <div class="spec-label-min">Wrenching Height (Min)</div>
                    <div class="spec-dimension">Wrenching Height</div>
                    <div class="spec-label-max">Total Runout (Max)</div>
                </div>
                <div class="spec-row">
                    <div class="spec-value">{product_details.get('Wrenching_Height_Min', 'N/A')}</div>
                    <div class="spec-dimension"></div>
                    <div class="spec-value">{product_details.get('Total_Runout_Max', 'N/A')}</div>
                </div>
            </div>
            
            <!-- Head Specifications Group -->
            <div class="oracle11g-card">
                <div class="filter-header">Head Specifications</div>
                
                <!-- Head Height -->
                <div class="spec-row">
                    <div class="spec-label-min">Head Height (Min)</div>
                    <div class="spec-dimension">Head Height</div>
                    <div class="spec-label-max">Head Height (Max)</div>
                </div>
                <div class="spec-row">
                    <div class="spec-value">{product_details.get('Head_Height_Min', 'N/A')}</div>
                    <div class="spec-dimension"></div>
                    <div class="spec-value">{product_details.get('Head_Height_Max', 'N/A')}</div>
                </div>
                
                <!-- Radius of Fillet -->
                <div class="spec-row">
                    <div class="spec-label-min">Radius of Fillet (Min)</div>
                    <div class="spec-dimension">Radius of Fillet</div>
                    <div class="spec-label-max">Radius of Fillet (Max)</div>
                </div>
                <div class="spec-row">
                    <div class="spec-value">{product_details.get('Radius_Fillet_Min', 'N/A')}</div>
                    <div class="spec-dimension"></div>
                    <div class="spec-value">{product_details.get('Radius_Fillet_Max', 'N/A')}</div>
                </div>
                
                <!-- Washer Face Thickness -->
                <div class="spec-row">
                    <div class="spec-label-min">Washer Face Thickness (Min)</div>
                    <div class="spec-dimension">Washer Face Thickness</div>
                    <div class="spec-label-max">Washer Face Thickness (Max)</div>
                </div>
                <div class="spec-row">
                    <div class="spec-value">{product_details.get('Washer_Face_Thickness_Min', 'N/A')}</div>
                    <div class="spec-dimension"></div>
                    <div class="spec-value">{product_details.get('Washer_Face_Thickness_Max', 'N/A')}</div>
                </div>
            </div>
            
            <!-- Additional Specifications Group -->
            <div class="oracle11g-card">
                <div class="filter-header">Additional Specifications</div>
                
                <!-- Wrenching Height -->
                <div class="spec-row">
                    <div class="spec-label-min">Wrenching Height (Min)</div>
                    <div class="spec-dimension">Wrenching Height</div>
                    <div class="spec-label-max">Total Runout (Max)</div>
                </div>
                <div class="spec-row">
                    <div class="spec-value">{product_details.get('Wrenching_Height_Min', 'N/A')}</div>
                    <div class="spec-dimension"></div>
                    <div class="spec-value">{product_details.get('Total_Runout_Max', 'N/A')}</div>
                </div>
                
                <!-- Thread Information -->
                <div class="spec-row">
                    <div class="spec-dimension" style="grid-column: 1 / span 3; text-align: center; background: var(--oracle11g-blue); color: white; padding: 0.8rem;">
                        <strong>Thread: {thread}</strong>
                    </div>
                </div>
            </div>
        </div>
        
        <div class="card-footer">
            <div>
                <strong>Generation Date:</strong> {current_date}<br>
                <strong>Generated By:</strong> {generated_by}
            </div>
            <div class="oracle11g-badge">
                Professional Specification
            </div>
        </div>
        
        <div class="card-actions">
            <button class="action-button" onclick="window.print()">Print Specification</button>
            <button class="action-button secondary">Email Specification</button>
            <button class="action-button secondary">Save as PDF</button>
        </div>
    </div>
    """
    
    # Display the card
    st.markdown(card_html, unsafe_allow_html=True)
    
    # Add some action buttons below the card
    col1, col2, col3 = st.columns(3)
    with col1:
        if st.button("View Raw Data", use_container_width=True, key="view_raw_pro_card"):
            st.dataframe(pd.DataFrame([product_details]))
    with col2:
        if st.button("Compare Products", use_container_width=True, key="compare_pro_card"):
            st.info("Product comparison feature coming soon!")
    with col3:
        if st.button("Close Card", use_container_width=True, key="close_pro_card"):
            st.session_state.show_professional_card = False
            st.rerun()

# ======================================================
# FIXED SECTION A - PROPER PRODUCT-SERIES-STANDARD-SIZE-GRADE RELATIONSHIP
# ======================================================

def get_available_standards_for_product_series(product, series):
    """Get available standards based on selected product and series"""
    available_standards = ["All"]
    
    if product == "All" and series == "All":
        # Show all standards
        for standard in st.session_state.available_products.keys():
            available_standards.append(standard)
    elif product == "All" and series != "All":
        # Filter by series only
        for standard, std_series in st.session_state.available_series.items():
            if std_series == series:
                available_standards.append(standard)
    elif product != "All" and series == "All":
        # Filter by product only
        for standard, products in st.session_state.available_products.items():
            if product in products:
                available_standards.append(standard)
    else:
        # Filter by both product and series
        for standard, products in st.session_state.available_products.items():
            if product in products:
                std_series = st.session_state.available_series.get(standard, "")
                if std_series == series:
                    available_standards.append(standard)
    
    return available_standards

def get_available_sizes_for_standard_product(standard, product, grade="All"):
    """Get available sizes based on selected standard and product"""
    size_options = ["All"]
    
    if standard == "All" or product == "All":
        return size_options
    
    temp_df = get_filtered_dataframe(product, standard, grade)
    size_options = get_safe_size_options(temp_df)
    
    return size_options

# ======================================================
# FIXED SECTION B - THREAD SPECIFICATIONS WITH PROPER DATA HANDLING
# ======================================================
def show_enhanced_product_database():
    """Enhanced Product Intelligence Center with COMPLETELY FIXED Section C material properties"""
    
    st.markdown("""
    <div class="oracle11g-header">
        <h1>Product Intelligence Center - Independent Sections</h1>
        <p>Each section works completely independently - No dependencies</p>
        <div>
            <span class="oracle11g-badge">Enhanced Calculator</span>
            <span class="oracle11g-badge-orange">Product-Based Workflow</span>
            <span class="oracle11g-badge-green">Dynamic Standards</span>
            <span class="oracle11g-badge-yellow">Professional Grade</span>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    if df.empty and df_mechem.empty and df_iso4014.empty and not st.session_state.din7991_loaded and not st.session_state.asme_b18_3_loaded:
        st.error("No data sources available. Please check your data connections.")
        return
    
    # Section toggles
    st.markdown("### Section Controls")
    col1, col2, col3 = st.columns(3)
    
    with col1:
        section_a_active = st.checkbox("Section A - Dimensional Specifications", value=st.session_state.section_a_view, key="section_a_toggle")
        st.session_state.section_a_view = section_a_active
    
    with col2:
        section_b_active = st.checkbox("Section B - Thread Specifications", value=st.session_state.section_b_view, key="section_b_toggle")
        st.session_state.section_b_view = section_b_active
    
    with col3:
        section_c_active = st.checkbox("Section C - Material Properties", value=st.session_state.section_c_view, key="section_c_toggle")
        st.session_state.section_c_view = section_c_active
    
    st.markdown("---")
    
    # SECTION A - DIMENSIONAL SPECIFICATIONS (FIXED RELATIONSHIPS)
    if st.session_state.section_a_view:
        st.markdown("""
        <div class="independent-section">
            <h3 class="filter-header">Section A - Dimensional Specifications</h3>
            <p><strong>Relationship:</strong> Product -> Series -> Standards -> Size -> Grade (ISO 4014 only)</p>
        </div>
        """, unsafe_allow_html=True)
        
        col1, col2, col3, col4, col5 = st.columns(5)
        
        with col1:
            # 1. Product List - Get all unique products from all standards
            all_products = set()
            for standard_products_list in st.session_state.available_products.values():
                all_products.update(standard_products_list)
            all_products = ["All"] + sorted([p for p in all_products if p != "All"])
            
            dimensional_product = st.selectbox(
                "Product List", 
                all_products, 
                key="section_a_product",
                index=all_products.index(st.session_state.section_a_current_product) if st.session_state.section_a_current_product in all_products else 0
            )
            st.session_state.section_a_current_product = dimensional_product
        
        with col2:
            # 2. Series System - Always show both options
            series_options = ["All", "Inch", "Metric"]
            dimensional_series = st.selectbox(
                "Series System", 
                series_options, 
                key="section_a_series",
                index=series_options.index(st.session_state.section_a_current_series) if st.session_state.section_a_current_series in series_options else 0
            )
            st.session_state.section_a_current_series = dimensional_series
        
        with col3:
            # 3. Standards - Filtered based on Product and Series
            available_standards = get_available_standards_for_product_series(dimensional_product, dimensional_series)
            
            dimensional_standard = st.selectbox(
                "Standards", 
                available_standards, 
                key="section_a_standard",
                index=available_standards.index(st.session_state.section_a_current_standard) if st.session_state.section_a_current_standard in available_standards else 0
            )
            st.session_state.section_a_current_standard = dimensional_standard
            
            # Show info about available standards
            if dimensional_standard != "All":
                std_series = st.session_state.available_series.get(dimensional_standard, "Unknown")
                st.caption(f"Series: {std_series}")
        
        with col4:
            # 4. Size - Filtered based on Standard and Product
            available_sizes = get_available_sizes_for_standard_product(dimensional_standard, dimensional_product)
            
            dimensional_size = st.selectbox(
                "Size", 
                available_sizes, 
                key="section_a_size",
                index=available_sizes.index(st.session_state.section_a_current_size) if st.session_state.section_a_current_size in available_sizes else 0
            )
            st.session_state.section_a_current_size = dimensional_size
            
            # Show info about available sizes
            if dimensional_size != "All":
                st.caption(f"Sizes available: {len(available_sizes)-1}")
        
        with col5:
            # 5. Grade - Only for ISO 4014 Hex Bolt
            if dimensional_standard == "ISO 4014" and dimensional_product == "Hex Bolt":
                grade_options = get_available_grades_for_standard_product(dimensional_standard, dimensional_product)
                dimensional_grade = st.selectbox(
                    "Product Grade", 
                    grade_options, 
                    key="section_a_grade",
                    index=grade_options.index(st.session_state.section_a_current_grade) if st.session_state.section_a_current_grade in grade_options else 0
                )
                st.session_state.section_a_current_grade = dimensional_grade
                
                # Show info about available grades
                if dimensional_grade != "All":
                    st.caption(f"Grade: {dimensional_grade}")
                    st.caption("⚠️ Different dimensions for A/B")
            else:
                st.info("Grade not applicable")
                dimensional_grade = "Not Applicable"
                st.session_state.section_a_current_grade = "All"
        
        # Debug information
        if st.session_state.debug_mode:
            st.info(f"""
            **Debug Info - Section A:**
            - Product: {dimensional_product}
            - Series: {dimensional_series} 
            - Standards Available: {len(available_standards)-1}
            - Sizes Available: {len(available_sizes)-1}
            - Grade: {dimensional_grade}
            - Selected Standard: {dimensional_standard}
            - Selected Size: {dimensional_size}
            """)
        
        # Apply Section A Filters Button
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            if st.button("APPLY SECTION A FILTERS", use_container_width=True, type="primary", key="apply_section_a"):
                with LoadingManager.show_loading_spinner("Applying filters..."):
                    st.session_state.section_a_filters = {
                        'product': dimensional_product,
                        'series': dimensional_series,
                        'standard': dimensional_standard,
                        'size': dimensional_size,
                        'grade': dimensional_grade if dimensional_standard == "ISO 4014" and dimensional_product == "Hex Bolt" else "All"
                    }
                    # Apply filters and store results
                    st.session_state.section_a_results = apply_section_a_filters()
                st.rerun()
        
        # Show Section A Results
        show_section_a_results()
    
    # SECTION B - THREAD SPECIFICATIONS (FIXED DATA TYPES)
    if st.session_state.section_b_view:
        st.markdown("""
        <div class="independent-section">
            <h3 class="filter-header">Section B - Thread Specifications</h3>
            <p><strong>FIXED:</strong> Proper data loading from Excel files with correct tolerance classes</p>
        </div>
        """, unsafe_allow_html=True)
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            # Thread standards
            thread_standards = ["All", "ASME B1.1", "ISO 965-2-98 Coarse", "ISO 965-2-98 Fine"]
            thread_standard = st.selectbox(
                "Thread Standard", 
                thread_standards, 
                key="section_b_standard",
                index=thread_standards.index(st.session_state.section_b_current_standard) if st.session_state.section_b_current_standard in thread_standards else 0
            )
            st.session_state.section_b_current_standard = thread_standard
            
            # Show thread data info
            if thread_standard != "All":
                df_thread = load_thread_data_enhanced(thread_standard)
                if not df_thread.empty:
                    st.caption(f"Threads available: {len(df_thread)}")
                    if st.session_state.debug_mode:
                        st.caption(f"Columns: {df_thread.columns.tolist()}")
        
        with col2:
            # Thread sizes - FIXED: Get from actual Excel data
            thread_size_options = get_thread_sizes_enhanced(thread_standard)
            
            thread_size = st.selectbox(
                "Thread Size", 
                thread_size_options, 
                key="section_b_size",
                index=thread_size_options.index(st.session_state.section_b_current_size) if st.session_state.section_b_current_size in thread_size_options else 0
            )
            st.session_state.section_b_current_size = thread_size
            
            if thread_size != "All":
                st.caption(f"Sizes available: {len(thread_size_options)-1}")
        
        with col3:
            # Tolerance classes - FIXED: Get ACTUAL classes from Excel data
            if thread_standard == "ASME B1.1":
                # Get actual tolerance classes from Excel data
                tolerance_options = get_thread_classes_enhanced(thread_standard)
                
                # If no specific classes found, use default
                if len(tolerance_options) == 1:  # Only "All"
                    tolerance_options = ["All", "1A", "2A", "3A"]
                elif len(tolerance_options) > 1 and "All" in tolerance_options:
                    # If "All" is present, prioritize it
                    tolerance_options = ["All"] + [cls for cls in tolerance_options if cls != "All"]
                
                tolerance_class = st.selectbox(
                    "Tolerance Class", 
                    tolerance_options, 
                    key="section_b_class",
                    index=tolerance_options.index(st.session_state.section_b_current_class) if st.session_state.section_b_current_class in tolerance_options else 0
                )
                st.session_state.section_b_current_class = tolerance_class
                
                if tolerance_class != "All":
                    st.caption(f"Classes available: {len(tolerance_options)-1}")
            else:
                # For metric threads, show available classes from data
                tolerance_options = get_thread_classes_enhanced(thread_standard)
                tolerance_class = st.selectbox(
                    "Tolerance Class", 
                    tolerance_options, 
                    key="section_b_class",
                    index=tolerance_options.index(st.session_state.section_b_current_class) if st.session_state.section_b_current_class in tolerance_options else 0
                )
                st.session_state.section_b_current_class = tolerance_class
                
                if tolerance_class != "All":
                    st.caption(f"Classes available: {len(tolerance_options)-1}")
        
        # Debug information for Section B
        if st.session_state.debug_mode and thread_standard != "All":
            df_thread_sample = load_thread_data_enhanced(thread_standard)
            if not df_thread_sample.empty:
                st.info(f"""
                **Debug Info - Section B ({thread_standard}):**
                - Total Records: {len(df_thread_sample)}
                - Columns: {df_thread_sample.columns.tolist()}
                - Unique Sizes: {len(get_thread_sizes_enhanced(thread_standard))-1}
                - Unique Classes: {len(get_thread_classes_enhanced(thread_standard))-1}
                - Sample Data: {df_thread_sample[['Thread', 'Class']].head(3).to_dict() if 'Thread' in df_thread_sample.columns and 'Class' in df_thread_sample.columns else 'No Thread/Class columns'}
                """)
        
        # Apply Section B Filters Button
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            if st.button("APPLY SECTION B FILTERS", use_container_width=True, type="primary", key="apply_section_b"):
                with LoadingManager.show_loading_spinner("Applying thread filters..."):
                    st.session_state.section_b_filters = {
                        'standard': thread_standard,
                        'size': thread_size,
                        'class': tolerance_class
                    }
                    # Apply filters and store results
                    st.session_state.section_b_results = apply_section_b_filters()
                st.rerun()
        
        # Show Section B Results
        show_section_b_results()
    
    # SECTION C - MATERIAL PROPERTIES (COMPLETELY INDEPENDENT) - COMPLETELY FIXED VERSION
    if st.session_state.section_c_view:
        st.markdown("""
        <div class="independent-section">
            <h3 class="filter-header">Section C - Material Properties</h3>
            <p><strong>COMPLETELY FIXED:</strong> Works with ALL property classes including 10.9, 6.8, 8.8, 304, A, B, B7</p>
        </div>
        """, unsafe_allow_html=True)
        
        col1, col2 = st.columns(2)
        
        with col1:
            # Property classes - FIXED: Get ALL property classes from Mechanical & Chemical data
            property_classes = ["All"]
            if st.session_state.property_classes:
                property_classes.extend(sorted(st.session_state.property_classes))
            else:
                # If no property classes found, show a message
                st.info("No property classes found in Mechanical & Chemical data")
                property_classes = ["All", "No data available"]
            
            property_class = st.selectbox(
                "Property Class (Grade)", 
                property_classes, 
                key="section_c_class",
                index=property_classes.index(st.session_state.section_c_current_class) if st.session_state.section_c_current_class in property_classes else 0
            )
            st.session_state.section_c_current_class = property_class
            
            # Show info about selected property class
            if property_class != "All" and property_class != "No data available":
                st.caption(f"Selected: {property_class}")
                # Show available standards for this property class
                available_standards = get_standards_for_property_class(property_class)
                if available_standards:
                    st.caption(f"Available standards: {len(available_standards)}")
        
        with col2:
            # Material standards - FIXED: Get standards based on selected property class
            material_standards = ["All"]
            if property_class != "All" and property_class != "No data available":
                mechem_standards = get_standards_for_property_class(property_class)
                if mechem_standards:
                    material_standards.extend(sorted(mechem_standards))
                else:
                    st.caption("No specific standards found for this property class")
                    # Add some common standards as fallback
                    material_standards.extend(["ASTM A193", "ASTM A320", "ISO 898-1", "ASME B18.2.1"])
            
            material_standard = st.selectbox(
                "Material Standard", 
                material_standards, 
                key="section_c_standard",
                index=material_standards.index(st.session_state.section_c_current_standard) if st.session_state.section_c_current_standard in material_standards else 0
            )
            st.session_state.section_c_current_standard = material_standard
            
            # Show info about available standards
            if material_standard != "All":
                st.caption(f"Standard: {material_standard}")
        
        # Debug information for Section C
        if st.session_state.debug_mode:
            st.info(f"""
            **Debug Info - Section C:**
            - Property Classes Available: {len(property_classes)-1}
            - Selected Property Class: {property_class}
            - Standards Available: {len(material_standards)-1}
            - Selected Standard: {material_standard}
            - Mechanical & Chemical Data: {len(df_mechem)} records
            - Sample Property Classes: {st.session_state.property_classes[:5] if st.session_state.property_classes else 'None'}
            """)
        
        # Apply Section C Filters Button
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            if st.button("APPLY SECTION C FILTERS", use_container_width=True, type="primary", key="apply_section_c"):
                if property_class == "All" or property_class == "No data available":
                    st.warning("Please select a valid property class")
                else:
                    with LoadingManager.show_loading_spinner("Applying material filters..."):
                        st.session_state.section_c_filters = {
                            'property_class': property_class,
                            'standard': material_standard
                        }
                        # Apply filters and store results
                        st.session_state.section_c_results = apply_section_c_filters()
                    
                    # Show immediate feedback
                    if st.session_state.section_c_results.empty:
                        st.warning(f"No data found for Property Class: {property_class} and Standard: {material_standard}")
                    else:
                        st.success(f"Found {len(st.session_state.section_c_results)} records for {property_class}")
                    
                    st.rerun()
        
        # Show Section C Results
        show_section_c_results()
    
    # COMBINE ALL RESULTS SECTION
    st.markdown("---")
    st.markdown("### Combine All Sections")
    
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        if st.button("COMBINE ALL SECTION RESULTS", use_container_width=True, type="secondary", key="combine_all"):
            with LoadingManager.show_loading_spinner("Combining results..."):
                st.session_state.combined_results = combine_all_results()
            st.rerun()
    
    # Show Combined Results
    show_combined_results()
    
    # Quick actions
    st.markdown("---")
    st.markdown("### Quick Actions")
    
    quick_col1, quick_col2, quick_col3, quick_col4 = st.columns(4)
    
    with quick_col1:
        if st.button("Clear All Filters", use_container_width=True, key="clear_all"):
            st.session_state.section_a_filters = {}
            st.session_state.section_b_filters = {}
            st.session_state.section_c_filters = {}
            st.session_state.section_a_results = pd.DataFrame()
            st.session_state.section_b_results = pd.DataFrame()
            st.session_state.section_c_results = pd.DataFrame()
            st.session_state.combined_results = pd.DataFrame()
            st.session_state.show_professional_card = False
            # Reset current selections
            st.session_state.section_a_current_product = "All"
            st.session_state.section_a_current_series = "All"
            st.session_state.section_a_current_standard = "All"
            st.session_state.section_a_current_size = "All"
            st.session_state.section_a_current_grade = "All"
            st.session_state.section_b_current_standard = "All"
            st.session_state.section_b_current_size = "All"
            st.session_state.section_b_current_class = "All"
            st.session_state.section_c_current_class = "All"
            st.session_state.section_c_current_standard = "All"
            st.rerun()
    
    with quick_col2:
        if st.button("View All Data", use_container_width=True, key="view_all"):
            # Show all available data
            st.session_state.section_a_results = df.copy()
            # Load thread data for ASME B1.1
            st.session_state.section_b_results = get_thread_data_enhanced("ASME B1.1")
            if not df_mechem.empty:
                st.session_state.section_c_results = df_mechem.copy()
            st.rerun()
    
    with quick_col3:
        if st.button("Export Everything", use_container_width=True, key="export_all"):
            # Combine current results and export
            combined = combine_all_results()
            if not combined.empty:
                enhanced_export_data(combined, "Excel")
            else:
                st.warning("No data to export")
    
    with quick_col4:
        if st.button("Reset Sections", use_container_width=True, key="reset_sections"):
            st.session_state.section_a_view = True
            st.session_state.section_b_view = True
            st.session_state.section_c_view = True
            st.rerun()

# ======================================================
# FIXED HOME DASHBOARD
# ======================================================
def show_rectified_home():
    """Show professional engineering dashboard"""
    
    st.markdown("""
    <div class="oracle11g-header">
        <h1>JSC Industries</h1>
        <p>Professional Fastener Intelligence Platform v4.0 - FIXED</p>
        <div>
            <span class="oracle11g-badge">FIXED Calculator</span>
            <span class="oracle11g-badge-orange">Separate Data Fetching</span>
            <span class="oracle11g-badge-green">Same Formula</span>
            <span class="oracle11g-badge-yellow">Different Standards</span>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    col1, col2, col3, col4 = st.columns(4)
    
    total_products = len(df) + (len(df_iso4014) if not df_iso4014.empty else 0) + (len(df_din7991) if st.session_state.din7991_loaded else 0) + (len(df_asme_b18_3) if st.session_state.asme_b18_3_loaded else 0)
    total_dimensional_standards = st.session_state.dimensional_standards_count
    total_threads = len(thread_files)
    total_mecert = len(df_mechem) if not df_mechem.empty else 0
    
    with col1:
        st.markdown(f"""
        <div class="metric-card">
            <h3 style="color: var(--oracle11g-blue); margin:0;">Products</h3>
            <h2 style="color: var(--oracle11g-blue-dark); margin:0.5rem 0;">{total_products}</h2>
            <p style="color: var(--oracle11g-gray); margin:0;">Total Records</p>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        st.markdown(f"""
        <div class="metric-card">
            <h3 style="color: var(--oracle11g-blue); margin:0;">Dimensional Standards</h3>
            <h2 style="color: var(--oracle11g-blue-dark); margin:0.5rem 0;">{total_dimensional_standards}</h2>
            <p style="color: var(--oracle11g-gray); margin:0;">ASME B18.2.1, ASME B18.3, ISO 4014, DIN-7991</p>
        </div>
        """, unsafe_allow_html=True)
    
    with col3:
        st.markdown(f"""
        <div class="metric-card">
            <h3 style="color: var(--oracle11g-blue); margin:0;">Thread Types</h3>
            <h2 style="color: var(--oracle11g-blue-dark); margin:0.5rem 0;">{total_threads}</h2>
            <p style="color: var(--oracle11g-gray); margin:0;">Available</p>
        </div>
        """, unsafe_allow_html=True)
    
    with col4:
        st.markdown(f"""
        <div class="metric-card">
            <h3 style="color: var(--oracle11g-blue); margin:0;">ME&CERT</h3>
            <h2 style="color: var(--oracle11g-blue-dark); margin:0.5rem 0;">{total_mecert}</h2>
            <p style="color: var(--oracle11g-gray); margin:0;">Properties</p>
        </div>
        """, unsafe_allow_html=True)
    
    st.markdown('<h2 class="section-header">Engineering Tools - FIXED</h2>', unsafe_allow_html=True)
    
    cols = st.columns(3)
    actions = [
        ("Product Database", "Professional product discovery with engineering filters", "database"),
        ("Engineering Calculator", "FIXED weight calculations with separate data fetching", "calculator"),
        ("Batch Calculator", "Industrial-scale processing for 1000+ products", "batch"),
        ("Compare Products", "Side-by-side technical comparison", "compare"),
        ("Export Reports", "Generate professional engineering reports", "export")
    ]
    
    for idx, (title, description, key) in enumerate(actions):
        with cols[idx % 3]:
            if st.button(f"**{title}**\n\n{description}", key=f"home_{key}"):
                section_map = {
                    "database": "Product Database",
                    "calculator": "Calculations",
                    "batch": "Batch Calculator"
                }
                st.session_state.selected_section = section_map.get(key, "Product Database")
                st.rerun()
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown('<h3 class="section-header">System Status - FIXED</h3>', unsafe_allow_html=True)
        
        status_items = [
            ("ASME B18.2.1 Data", not df.empty, "oracle11g-badge"),
            ("ISO 4014 Data", not df_iso4014.empty, "oracle11g-badge-orange"),
            ("DIN-7991 Data", st.session_state.din7991_loaded, "oracle11g-badge-green"),
            ("ASME B18.3 Data", st.session_state.asme_b18_3_loaded, "oracle11g-badge-yellow"),
            ("ME&CERT Data", not df_mechem.empty, "oracle11g-badge"),
            ("Thread Data", any(not load_thread_data_enhanced(url).empty for url in thread_files.values()), "oracle11g-badge-orange"),
            ("Weight Calculations", True, "oracle11g-badge-green"),
            ("FIXED Calculator", True, "oracle11g-badge-yellow"),
            ("Batch Calculator", True, "oracle11g-badge"),
        ]
        
        for item_name, status, badge_class in status_items:
            if status:
                st.markdown(f'<div class="{badge_class}" style="margin: 0.3rem 0;">{item_name} - Active</div>', unsafe_allow_html=True)
            else:
                st.markdown(f'<div class="{badge_class}" style="margin: 0.3rem 0; background: var(--oracle11g-gray);">{item_name} - Limited</div>', unsafe_allow_html=True)
    
    with col2:
        st.markdown('<h3 class="section-header">FIXED Features</h3>', unsafe_allow_html=True)
        
        features = [
            "Socket Head Products: Same formula, separate data fetching",
            "Hexagon Socket Head Cap Screws (ASME B18.3): Separate function get_asme_b18_3_dimensions()",
            "Hexagon Socket Countersunk Head Cap Screw (DIN-7991): Separate function get_din7991_dimensions()",
            "Same formula: 0.7853 × d² × h for all socket head products",
            "Detailed calculation parameters for ALL products",
            "Hex Bolt, Heavy Hex Bolt detailed parameters",
            "Hex Cap Screws, Heavy Hex Screws detailed parameters",
            "Threaded Rod detailed parameters",
            "Professional reporting",
            "Carbon steel density: 7.85 g/cm³",
            "Batch processing capabilities",
            "Database-connected calculations",
            "Enhanced user interface",
            "ISO 4014 Product Grade selection"
        ]
        
        for feature in features:
            st.markdown(f'<div class="oracle11g-card" style="padding: 0.5rem; margin: 0.2rem 0;">• {feature}</div>', unsafe_allow_html=True)
    
    show_calculation_history()

# ======================================================
# HELP SYSTEM
# ======================================================
def show_help_system():
    """Show contextual help system"""
    with st.sidebar:
        st.markdown("---")
        with st.expander("FIXED Weight Calculator Guide"):
            st.markdown("""
            **SEPARATE DATA FETCHING FOR SOCKET HEAD PRODUCTS:**
            
            **ASME B18.3 - Hexagon Socket Head Cap Screws:**
            - Function: `get_asme_b18_3_dimensions()`
            - Data: Inches (converted to mm)
            - Specific column mapping for ASME B18.3
            
            **DIN-7991 - Hexagon Socket Countersunk Head Cap Screw:**
            - Function: `get_din7991_dimensions()`  
            - Data: Millimeters (no conversion needed)
            - Specific column mapping for DIN-7991 (dk, k columns)
            
            **SAME FORMULA FOR BOTH:**
            - **Shank Volume:** 0.7853 × (diameter)² × length (mm³)
            - **Socket Head Volume:** 0.7853 × (head_diameter_min)² × head_height_min (mm³)
            - **Volume Conversion:** mm³ to cm³ = divide by 1000
            - **Weight Calculation:** volume_cm³ × density_g/cm³
            
            **DENSITY VALUES (g/cm³):**
            - Carbon Steel: 7.85
            - Stainless Steel: 8.00
            - Alloy Steel: 7.85
            - Brass: 8.50
            - Aluminum: 2.70
            - Copper: 8.96
            - Titanium: 4.50
            
            **DETAILED PARAMETERS:**
            - Now shows for ALL products including Socket Head products
            - Complete dimension breakdown
            - Volume calculations for each component
            """)

        with st.expander("Batch Calculator Guide"):
            st.markdown("""
            **BATCH CALCULATOR FEATURES:**
            
            **Two Input Modes:**
            - **Basic Mode:** Provide only Size + Length → System auto-detects other parameters
            - **Advanced Mode:** Provide complete specifications for precise control
            
            **Auto-Detection Logic:**
            - **Metric Sizes (M10, M12):** Auto-detects as ISO 4014 Hex Bolt
            - **Inch Sizes (1/4, 5/16):** Auto-detects as ASME B18.2.1 Hex Bolt
            - **Diameter Calculation:** Automatically calculates from size
            - **Material Default:** Carbon Steel (can be overridden)
            
            **Processing Capabilities:**
            - Process 1000+ records simultaneously
            - Real-time progress tracking
            - Continue processing on errors
            - Comprehensive error reporting
            - Multiple export formats
            
            **Output Features:**
            - Individual item weights
            - Total weight summaries
            - Error diagnostics
            - Professional Excel reports
            """)

# ======================================================
# SECTION DISPATCHER
# ======================================================
def show_section(title):
    if title == "Product Database":
        show_enhanced_product_database()
    elif title == "Calculations":
        show_rectified_calculations()
    elif title == "Batch Calculator":
        show_batch_weight_calculator()
    else:
        st.info(f"Section {title} is coming soon!")
    
    st.markdown("---")
    if st.button("Back to Dashboard", use_container_width=True):
        st.session_state.selected_section = None
        st.rerun()

# ======================================================
# MAIN APPLICATION
# ======================================================
def main():
    """Main application entry point"""
    
    show_help_system()
    
    show_data_quality_indicators()
    
    # Sidebar navigation
    with st.sidebar:
        st.markdown("## Navigation")
        
        sections = [
            "Home Dashboard",
            "Product Database", 
            "Calculations",
            "Batch Calculator"
        ]
        
        for section in sections:
            if st.button(section, use_container_width=True, key=f"nav_{section}"):
                if section == "Home Dashboard":
                    st.session_state.selected_section = None
                else:
                    st.session_state.selected_section = section
                st.rerun()
        
        # Debug mode toggle
        st.markdown("---")
        st.session_state.debug_mode = st.checkbox("Debug Mode", value=st.session_state.debug_mode)
        
        # Mobile optimization toggle
        st.session_state.mobile_view_optimized = st.checkbox("Mobile Optimized View", value=st.session_state.mobile_view_optimized)
        
        # Export templates section
        st.markdown("---")
        with st.expander("Export Templates"):
            st.markdown("### Download Templates")
            
            if st.button("Weight Calculator Template"):
                template_df = ExportTemplateManager.get_weight_calc_template()
                csv_template = template_df.to_csv(index=False)
                st.download_button(
                    label="Download CSV Template",
                    data=csv_template,
                    file_name="weight_calculator_template.csv",
                    mime="text/csv",
                    use_container_width=True
                )
            
            if st.button("Product Database Template"):
                template_df = ExportTemplateManager.get_product_database_template()
                csv_template = template_df.to_csv(index=False)
                st.download_button(
                    label="Download CSV Template",
                    data=csv_template,
                    file_name="product_database_template.csv",
                    mime="text/csv",
                    use_container_width=True
                )
            
            if st.button("Batch Calculator Template"):
                col1, col2 = st.columns(2)
                with col1:
                    basic_template = BatchTemplateManager.get_basic_template()
                    csv_basic = basic_template.to_csv(index=False)
                    st.download_button(
                        label="Basic Template",
                        data=csv_basic,
                        file_name="batch_calculator_basic_template.csv",
                        mime="text/csv",
                        use_container_width=True
                    )
                with col2:
                    advanced_template = BatchTemplateManager.get_advanced_template()
                    csv_advanced = advanced_template.to_csv(index=False)
                    st.download_button(
                        label="Advanced Template",
                        data=csv_advanced,
                        file_name="batch_calculator_advanced_template.csv",
                        mime="text/csv",
                        use_container_width=True
                    )
    
    if st.session_state.selected_section is None:
        show_rectified_home()
    else:
        show_section(st.session_state.selected_section)
    
    st.markdown("""
        <hr>
        <div class="oracle11g-footer">
            <div style="display: flex; justify-content: center; gap: 2rem; margin-bottom: 1rem;">
                <span class="oracle11g-badge">FIXED Calculator</span>
                <span class="oracle11g-badge-orange">Separate Data Fetching</span>
                <span class="oracle11g-badge-green">Same Formula</span>
                <span class="oracle11g-badge-yellow">Batch Processing</span>
            </div>
            <p><strong>© 2024 JSC Industries Pvt Ltd</strong> | Born to Perform • Engineered for Excellence</p>
            <p style="font-size: 0.8rem;">Professional Fastener Intelligence Platform v4.0 - FIXED Weight Calculator with SEPARATE data fetching and SAME formula for socket head products</p>
        </div>
    """, unsafe_allow_html=True)

# Run the application
if __name__ == "__main__":
    main()